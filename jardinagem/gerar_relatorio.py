"""
Garra Terraplenagem — Gerador de Relatórios Jardinagem
Gera Excel com fotos + KM e envia por email ao cliente
"""
import os, io, requests, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

AZUL  = "1A2A5E"
LARJ  = "E8820C"
BRNCO = "FFFFFF"

def fill(h): return PatternFill("solid", fgColor=h)
def brd(c="CBD5E1"):
    s=Side(style="thin",color=c)
    return Border(left=s,right=s,top=s,bottom=s)
def aln(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

# ── Logo ──────────────────────────────────────────────────────
LOGO_PATH = os.path.join(os.path.dirname(__file__), "static", "icons", "logo-garra.jpg")

def make_logo(w=160):
    img = PILImage.open(LOGO_PATH)
    h   = int(w * img.height / img.width)
    img = img.resize((w, h), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    buf.seek(0)
    return buf, w, h

# ── Download foto do Supabase Storage ────────────────────────
def baixar_foto(storage_path, supabase_url, service_key,
                max_px=800, qualidade=80):
    """Baixa foto e retorna BytesIO redimensionada para o Excel."""
    url = f"{supabase_url}/storage/v1/object/{storage_path}"
    headers = {"Authorization": f"Bearer {service_key}"}
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code != 200:
        return None
    img = PILImage.open(io.BytesIO(r.content))
    if img.mode not in ("RGB","L"):
        img = img.convert("RGB")
    img.thumbnail((max_px, max_px), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=qualidade)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════
def gerar_relatorio_fotos(semana: dict, pares: list,
                           supabase_url: str, service_key: str) -> io.BytesIO:
    """
    Gera Excel do Relatório Fotográfico.
    semana: {label, data_ini, data_fim}
    pares:  [{codigo_a, codigo_d, local_nome, foto_antes{storage_path}, foto_depois{storage_path}}]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Fotográfico"
    ws.sheet_view.showGridLines = False

    for col in range(1,10):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["A"].width = 2

    # Logo
    buf_logo, lw, lh = make_logo(160)
    logo = XLImage(buf_logo)
    logo.width=lw; logo.height=lh; logo.anchor="B1"
    ws.add_image(logo)
    logo_rows = max(5, lh//14+1)
    for r in range(1, logo_rows+1):
        ws.row_dimensions[r].height = 14

    ws.merge_cells(f"E1:I{logo_rows}")
    ws["E1"] = "RELATÓRIO FOTOGRÁFICO\nSERVIÇOS DE JARDINAGEM"
    ws["E1"].font      = Font(name="Arial",size=16,bold=True,color=AZUL)
    ws["E1"].alignment = aln("center","center",wrap=True)
    ws["E1"].fill      = fill("EFF6FF")

    ri = logo_rows+1
    ws.merge_cells(f"A{ri}:I{ri}")
    ws[f"A{ri}"] = (f"Empresa: Garra Terraplenagem e Caçambas  |  "
                    f"Equipe: Jardinagem 1  |  Período: {semana['data_ini']} a {semana['data_fim']}")
    ws[f"A{ri}"].font=Font(name="Arial",size=10,bold=True,color=BRNCO)
    ws[f"A{ri}"].fill=fill(LARJ); ws[f"A{ri}"].alignment=aln("center")
    ws.row_dimensions[ri].height=18

    rs=ri+1
    ws.merge_cells(f"A{rs}:I{rs}")
    ws[f"A{rs}"] = semana["label"]
    ws[f"A{rs}"].font=Font(name="Arial",size=11,bold=True,color=BRNCO)
    ws[f"A{rs}"].fill=fill(AZUL); ws[f"A{rs}"].alignment=aln("center")
    ws.row_dimensions[rs].height=18

    row = rs+2

    # Dimensão de cada foto no Excel (pixels)
    FOTO_W, FOTO_H = 280, 210

    for i in range(0, len(pares), 2):
        pa = pares[i]
        pb = pares[i+1] if i+1 < len(pares) else None

        ws.row_dimensions[row].height = 18

        # ── Labels códigos ────────────────────────────────────
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = f"{pa['codigo_a']}ª FOTO: INICIO LIMPEZA {pa['local_nome']}"
        ws[f"B{row}"].font=Font(name="Arial",size=9,bold=True,color=AZUL)
        ws[f"B{row}"].fill=fill("DBEAFE"); ws[f"B{row}"].alignment=aln("center",wrap=True); ws[f"B{row}"].border=brd()

        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"] = f"{pa['codigo_d']}ª FOTO: FIM LIMPEZA {pa['local_nome']}"
        ws[f"D{row}"].font=Font(name="Arial",size=9,bold=True,color="92400E")
        ws[f"D{row}"].fill=fill("FFF7ED"); ws[f"D{row}"].alignment=aln("center",wrap=True); ws[f"D{row}"].border=brd()

        if pb:
            ws.merge_cells(f"F{row}:G{row}")
            ws[f"F{row}"] = f"{pb['codigo_a']}ª FOTO: INICIO LIMPEZA {pb['local_nome']}"
            ws[f"F{row}"].font=Font(name="Arial",size=9,bold=True,color=AZUL)
            ws[f"F{row}"].fill=fill("DBEAFE"); ws[f"F{row}"].alignment=aln("center",wrap=True); ws[f"F{row}"].border=brd()

            ws.merge_cells(f"H{row}:I{row}")
            ws[f"H{row}"] = f"{pb['codigo_d']}ª FOTO: FIM LIMPEZA {pb['local_nome']}"
            ws[f"H{row}"].font=Font(name="Arial",size=9,bold=True,color="92400E")
            ws[f"H{row}"].fill=fill("FFF7ED"); ws[f"H{row}"].alignment=aln("center",wrap=True); ws[f"H{row}"].border=brd()

        # ── Área das fotos ────────────────────────────────────
        row += 1
        ROWS_FOTO = 14
        for r2 in range(row, row+ROWS_FOTO):
            ws.row_dimensions[r2].height = 14
            for cl in ["B","C","D","E","F","G","H","I"]:
                ws[f"{cl}{r2}"].border=brd("E2E8F0")
                ws[f"{cl}{r2}"].fill=fill("F8FAFC")

        # Insere foto ANTES do par A
        foto_a_antes = pa.get("foto_antes")
        if foto_a_antes and foto_a_antes.get("storage_path"):
            buf_img = baixar_foto(foto_a_antes["storage_path"], supabase_url, service_key)
            if buf_img:
                xl_img = XLImage(buf_img)
                xl_img.width=FOTO_W; xl_img.height=FOTO_H
                xl_img.anchor=f"B{row}"
                ws.add_image(xl_img)
        else:
            ws.merge_cells(f"B{row}:C{row+ROWS_FOTO-1}")
            ws[f"B{row}"].value="📷 Foto Antes"
            ws[f"B{row}"].font=Font(name="Arial",size=9,italic=True,color="94A3B8")
            ws[f"B{row}"].alignment=aln("center")

        # Insere foto DEPOIS do par A
        foto_a_dep = pa.get("foto_depois")
        if foto_a_dep and foto_a_dep.get("storage_path"):
            buf_img2 = baixar_foto(foto_a_dep["storage_path"], supabase_url, service_key)
            if buf_img2:
                xl_img2 = XLImage(buf_img2)
                xl_img2.width=FOTO_W; xl_img2.height=FOTO_H
                xl_img2.anchor=f"D{row}"
                ws.add_image(xl_img2)
        else:
            ws.merge_cells(f"D{row}:E{row+ROWS_FOTO-1}")
            ws[f"D{row}"].value="📷 Foto Depois"
            ws[f"D{row}"].font=Font(name="Arial",size=9,italic=True,color="94A3B8")
            ws[f"D{row}"].alignment=aln("center")

        if pb:
            foto_b_antes = pb.get("foto_antes")
            if foto_b_antes and foto_b_antes.get("storage_path"):
                buf_img3 = baixar_foto(foto_b_antes["storage_path"], supabase_url, service_key)
                if buf_img3:
                    xl_img3 = XLImage(buf_img3)
                    xl_img3.width=FOTO_W; xl_img3.height=FOTO_H
                    xl_img3.anchor=f"F{row}"
                    ws.add_image(xl_img3)
            else:
                ws.merge_cells(f"F{row}:G{row+ROWS_FOTO-1}")
                ws[f"F{row}"].value="📷 Foto Antes"
                ws[f"F{row}"].font=Font(name="Arial",size=9,italic=True,color="94A3B8")
                ws[f"F{row}"].alignment=aln("center")

            foto_b_dep = pb.get("foto_depois")
            if foto_b_dep and foto_b_dep.get("storage_path"):
                buf_img4 = baixar_foto(foto_b_dep["storage_path"], supabase_url, service_key)
                if buf_img4:
                    xl_img4 = XLImage(buf_img4)
                    xl_img4.width=FOTO_W; xl_img4.height=FOTO_H
                    xl_img4.anchor=f"H{row}"
                    ws.add_image(xl_img4)
            else:
                ws.merge_cells(f"H{row}:I{row+ROWS_FOTO-1}")
                ws[f"H{row}"].value="📷 Foto Depois"
                ws[f"H{row}"].font=Font(name="Arial",size=9,italic=True,color="94A3B8")
                ws[f"H{row}"].alignment=aln("center")

        row += ROWS_FOTO + 1

    # Rodapé
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "Garra Terraplenagem e Caçambas  ·  Jardinagem  ·  Águas de Pará de Minas"
    ws[f"A{row}"].font=Font(name="Arial",size=8,italic=True,color="64748B")
    ws[f"A{row}"].alignment=aln("center"); ws[f"A{row}"].fill=fill("F0F4FF")
    ws.row_dimensions[row].height=14

    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    return buf_out


# ══════════════════════════════════════════════════════════════
def gerar_relatorio_km(semana: dict, relatorios: list) -> io.BytesIO:
    """
    Gera Excel do Relatório Diário de KM.
    relatorios: [{data, local, km_ini, km_fin, hr_ini, hr_fim, obs, responsavel}]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Diário"
    ws.sheet_view.showGridLines = False

    for col,w in zip("ABCDEFGHI",[12,35,11,11,9,8,8,22,15]):
        ws.column_dimensions[col].width = w

    # Logo
    buf_logo, lw, lh = make_logo(120)
    logo = XLImage(buf_logo)
    logo.width=lw; logo.height=lh; logo.anchor="A1"
    ws.add_image(logo)
    lr = max(3, lh//14+1)
    for r in range(1, lr+1):
        ws.row_dimensions[r].height = 14

    ws.merge_cells(f"C1:I{lr}")
    ws["C1"] = "RELATÓRIO DIÁRIO — SERVIÇO JARDINAGEM"
    ws["C1"].font=Font(name="Arial",size=14,bold=True,color=AZUL)
    ws["C1"].alignment=aln("center","center",wrap=True); ws["C1"].fill=fill("EFF6FF")

    ri = lr+1
    ws.merge_cells(f"A{ri}:I{ri}")
    ws[f"A{ri}"] = f"Empresa: Garra Terraplenagem e Caçambas   |   Período: {semana['data_ini']} a {semana['data_fim']}"
    ws[f"A{ri}"].font=Font(name="Arial",size=10,bold=True,color=BRNCO)
    ws[f"A{ri}"].fill=fill(LARJ); ws[f"A{ri}"].alignment=aln("center")
    ws.row_dimensions[ri].height=18; ri+=1

    for col,hdr in zip("ABCDEFGHI",["Data","Local","KM Inicial","KM Final","KM Total","Início","Fim","Observação","Responsável"]):
        c=ws[f"{col}{ri}"]; c.value=hdr
        c.font=Font(name="Arial",size=10,bold=True,color=BRNCO)
        c.fill=fill("1E3A8A"); c.alignment=aln("center"); c.border=brd("1A2A5E")
    ws.row_dimensions[ri].height=20; ri+=1

    km_ini_total = relatorios[0]["km_ini"] if relatorios else 0
    km_fim_total = relatorios[-1]["km_fin"] if relatorios else 0

    for i,r in enumerate(relatorios):
        bg="FFFFFF" if i%2==0 else "F0F4FF"
        km_t = r["km_fin"]-r["km_ini"]
        dados=[r["data"],r["local"],r["km_ini"],r["km_fin"],km_t,
               r["hr_ini"],r["hr_fim"],r.get("obs",""),r["responsavel"]]
        for col,val,al in zip("ABCDEFGHI",dados,["center","left","center","center","center","center","center","left","center"]):
            c=ws[f"{col}{ri}"]; c.value=val
            c.font=Font(name="Arial",size=10,bold=(col=="I"),
                        color=AZUL if col=="I" else "1E293B")
            c.fill=fill(bg); c.alignment=aln(al); c.border=brd()
        ws.row_dimensions[ri].height=18; ri+=1

    ws.merge_cells(f"A{ri}:B{ri}")
    ws[f"A{ri}"]="TOTAL KM"
    ws[f"A{ri}"].font=Font(name="Arial",size=10,bold=True,color=BRNCO)
    ws[f"A{ri}"].fill=fill(AZUL); ws[f"A{ri}"].alignment=aln("center"); ws[f"A{ri}"].border=brd(AZUL)
    for col,val in zip("CDE",[km_ini_total, km_fim_total, f"=D{ri}-C{ri}"]):
        c=ws[f"{col}{ri}"]; c.value=val
        c.font=Font(name="Arial",size=10,bold=True,color=LARJ if col=="E" else BRNCO)
        c.fill=fill(AZUL); c.alignment=aln("center"); c.border=brd(AZUL)
    ws.row_dimensions[ri].height=22; ri+=2

    ws.merge_cells(f"A{ri}:I{ri}")
    ws[f"A{ri}"]="Garra Terraplenagem e Caçambas  ·  Jardinagem  ·  Águas de Pará de Minas"
    ws[f"A{ri}"].font=Font(name="Arial",size=8,italic=True,color="64748B")
    ws[f"A{ri}"].alignment=aln("center"); ws[f"A{ri}"].fill=fill("F0F4FF")
    ws.row_dimensions[ri].height=14

    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    return buf_out


# ══════════════════════════════════════════════════════════════
def enviar_relatorios_email(semana: dict, buf_fotos: io.BytesIO,
                             buf_km: io.BytesIO,
                             mail_user: str, mail_pass: str,
                             mail_destino: str, mail_cc: str = ""):
    """Envia os dois relatórios por email ao cliente."""
    periodo = f"{semana['data_ini']} a {semana['data_fim']}"
    assunto = f"Relatório de Jardinagem — {semana['label']} | {periodo}"

    msg = MIMEMultipart()
    msg["From"]    = mail_user
    msg["To"]      = mail_destino
    msg["Subject"] = assunto
    if mail_cc:
        msg["Cc"] = mail_cc

    corpo = f"""
Prezados,

Segue em anexo o relatório de serviços de jardinagem referente ao período de {periodo}.

• Relatório Fotográfico — registro visual dos serviços realizados
• Relatório Diário — registro de deslocamentos e horas trabalhadas

Atenciosamente,
Garra Terraplenagem e Caçambas
"""
    msg.attach(MIMEText(corpo, "plain", "utf-8"))

    # Anexo 1: Fotos
    nome_fotos = f"Relatorio-Fotos-{semana['data_ini'].replace('/','')}-{semana['data_fim'].replace('/','')}.xlsx"
    part1 = MIMEBase("application","vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part1.set_payload(buf_fotos.read())
    encoders.encode_base64(part1)
    part1.add_header("Content-Disposition","attachment",filename=nome_fotos)
    msg.attach(part1)

    # Anexo 2: KM
    nome_km = f"Relatorio-KM-{semana['data_ini'].replace('/','')}-{semana['data_fim'].replace('/','')}.xlsx"
    part2 = MIMEBase("application","vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part2.set_payload(buf_km.read())
    encoders.encode_base64(part2)
    part2.add_header("Content-Disposition","attachment",filename=nome_km)
    msg.attach(part2)

    destinatarios = [mail_destino]
    if mail_cc:
        destinatarios += [e.strip() for e in mail_cc.split(",") if e.strip()]

    with smtplib.SMTP("smtp.gmail.com", 587) as s:
        s.starttls()
        s.login(mail_user, mail_pass)
        s.sendmail(mail_user, destinatarios, msg.as_string())

    return True
