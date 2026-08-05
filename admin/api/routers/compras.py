# ══════════════════════════════════════════════════════════════
# MÓDULO COMPRAS — Ordens de Compra (v28)
# Fluxo: rascunho → solicitada → aprovada → enviada
#        → recebida_parcial → recebida | rejeitada | cancelada
# Regras:
#   - Toda OC tem SETOR (compras.setores, parametrizável — Regra 63)
#   - Alçadas por usuário (compras.alcadas): valor_limite NULL = sem limite;
#     OC dentro da alçada do criador → aprovada em 1 passo
#   - Aprovação/rejeição = assinatura digital (usuario + timestamp na trilha)
#   - Rejeição exige motivo
#   - Recebimento confere item a item (qtd_recebida acumulada)
#   - OC vinculada a OT: ao receber, o valor recebido soma no custo_total da OT
#   - peca_id em oc_itens é opcional — fase 2 liga ao manutencao.pecas/estoque
# Permissões (public.permissoes_colaborador):
#   'compras_solicitar' → criar/editar/enviar/receber OC
#   'compras_aprovar'   → aprovar/rejeitar (limitado pela alçada)
# ══════════════════════════════════════════════════════════════
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
import os

from core.auth import verificar_token
from core.db import ajard_query, ajard_query_id
from core.storage import storage_upload, storage_url

router = APIRouter()

_PERFIS_COMPRAS = {"admin", "gestor", "luana"}

_TRANSICOES = {
    "rascunho":         {"solicitada", "cancelada"},
    "solicitada":       {"aprovada", "rejeitada", "cancelada"},
    "aprovada":         {"enviada", "cancelada"},
    "enviada":          {"recebida_parcial", "recebida", "cancelada"},
    "recebida_parcial": {"recebida", "cancelada"},
    "recebida":         set(),
    "rejeitada":        set(),
    "cancelada":        set(),
}


async def _usuario_id(payload):
    u = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE login=%s",
        (payload.get("sub", ""),), fetch="one")
    return u["id"] if u else None


async def _tem_permissao(payload, modulo):
    """Perfil liberado OU permissão concedida em permissoes_colaborador."""
    if (payload.get("perfil") or "").lower() in _PERFIS_COMPRAS:
        return True
    uid = await _usuario_id(payload)
    if not uid:
        return False
    p = await ajard_query(
        """SELECT permitido FROM public.permissoes_colaborador
           WHERE usuario_id=%s AND modulo=%s""",
        (uid, modulo), fetch="one")
    return bool(p and p.get("permitido"))


async def verificar_compras(payload=Depends(verificar_token)):
    """Gate geral do módulo: solicitar, aprovar OU visão geral dão acesso.
    'compras_ver_todas' sozinha atende o perfil recebedor/financeiro
    (ex.: atendimento que confere mercadoria sem abrir OCs)."""
    if await _tem_permissao(payload, "compras_solicitar"):
        return payload
    if await _tem_permissao(payload, "compras_aprovar"):
        return payload
    if await _tem_permissao(payload, "compras_ver_todas"):
        return payload
    raise HTTPException(status_code=403, detail="Sem permissão para o módulo Compras")


async def verificar_compras_solicitante(payload=Depends(verificar_token)):
    """Criar OC exige a permissão de Solicitar (ou perfil de gestão).
    'Visão geral' sozinha vê e confere — não abre OCs."""
    if (payload.get("perfil") or "").lower() in _PERFIS_COMPRAS:
        return payload
    if await _tem_permissao(payload, "compras_solicitar"):
        return payload
    raise HTTPException(status_code=403, detail="Sem permissão para criar OCs — peça à central de cotações")


async def verificar_compras_gestor(payload=Depends(verificar_token)):
    """Parametrização do módulo (alçadas, setores, condições).
    Permissão controla o acesso, não o perfil (princípio do sistema):
    perfil de gestão passa por padrão, e o checkbox 'Compras — Alçadas'
    no painel de Permissões concede/nega por exceção. Quem só aprova OC
    não configura alçada — inclusive a própria."""
    if await _tem_permissao(payload, "compras_alcadas"):
        return payload
    raise HTTPException(status_code=403, detail="Somente a gestão configura o módulo Compras")


async def verificar_compras_aprovador(payload=Depends(verificar_token)):
    if await _tem_permissao(payload, "compras_aprovar"):
        return payload
    raise HTTPException(status_code=403, detail="Sem permissão para aprovar compras")


async def _ve_todas(payload):
    """Quem enxerga TODAS as OCs (e pode conferir recebimento/devolução delas):
    - perfil de gestão;
    - quem aprova (compras_aprovar) — precisa ver para decidir;
    - quem tem 'compras_ver_todas' (Visão geral): recebedor de entregas e
      financeiro — vê e confere tudo, SEM poder aprovar (alçada continua
      mandando) e sem editar/excluir OC alheia.
    Solicitante comum segue vendo apenas as próprias."""
    if (payload.get("perfil") or "").lower() in _PERFIS_COMPRAS:
        return True
    if await _tem_permissao(payload, "compras_aprovar"):
        return True
    return await _tem_permissao(payload, "compras_ver_todas")


async def _alcada_efetiva(payload):
    """Alçada usada nos gates de aprovação. Perfil ADMIN (Master) aprova
    livre, sem exigir cadastro (decisão 29/07/2026). Demais usuários:
    somente com linha ativa em compras.alcadas.
    Retorna (tem, limite_por_compra, limite_mensal)."""
    if (payload.get("perfil") or "").lower() == "admin":
        return (True, None, None)
    uid = await _usuario_id(payload)
    return await _alcada_do_usuario(uid)


async def _alcada_do_usuario(uid):
    """(tem_alcada, valor_limite, limite_mensal). None = sem limite.
    Sem linha em compras.alcadas = sem alçada nenhuma."""
    a = await ajard_query(
        """SELECT valor_limite, limite_mensal
           FROM compras.alcadas WHERE usuario_id=%s AND ativo=true""",
        (uid,), fetch="one")
    if not a:
        return (False, None, None)
    return (True, a.get("valor_limite"), a.get("limite_mensal"))


async def _aprovado_no_mes(uid):
    """Soma do que o aprovador já comprometeu no mês corrente (OCs aprovadas
    por ele, não rejeitadas/canceladas). Base do teto mensal: crédito que
    renova a cada mês — aprovou R$ 50 de um teto de R$ 300, restam R$ 250."""
    row = await ajard_query(
        """SELECT COALESCE(SUM(valor_total),0) AS total
           FROM compras.ordens_compra
           WHERE aprovador_id=%s AND ativo=true
             AND status NOT IN ('rejeitada','cancelada')
             AND date_trunc('month', data_aprovacao) = date_trunc('month', now())""",
        (uid,), fetch="one")
    return float(row["total"] or 0)


def _valor_dentro_alcada(valor, tem_alcada, limite):
    """Checagem POR COMPRA (limite individual de cada OC)."""
    if not tem_alcada:
        return False
    if limite is None:
        return True
    try:
        return float(valor or 0) <= float(limite)
    except (TypeError, ValueError):
        return False


async def _cabe_no_teto_mensal(uid, valor, limite_mensal):
    """Checagem MENSAL ACUMULADA. limite_mensal None = sem teto (padrão,
    comportamento anterior). Retorna (cabe, ja_aprovado, saldo)."""
    if limite_mensal is None:
        return (True, None, None)
    ja = await _aprovado_no_mes(uid)
    saldo = float(limite_mensal) - ja
    return (float(valor or 0) <= saldo + 1e-9, ja, saldo)


def _validar_itens(itens):
    """Sanidade dos itens: quantidade > 0 e valor >= 0. Sem isso, um item
    negativo derruba o total e burla alçada/teto (achado da auditoria)."""
    if not itens:
        raise HTTPException(status_code=400, detail="Informe ao menos 1 item")
    for i in itens:
        try:
            q = float(i.get("quantidade") or 0)
            v = float(i.get("valor_unit") or 0)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Quantidade/valor inválidos")
        if q <= 0:
            raise HTTPException(status_code=400,
                                detail=f"Quantidade deve ser maior que zero ({i.get('descricao','item')})")
        if v < 0:
            raise HTTPException(status_code=400,
                                detail=f"Valor não pode ser negativo ({i.get('descricao','item')})")
        if not (i.get("descricao") or "").strip():
            raise HTTPException(status_code=400, detail="Item sem descrição")


async def _trilha(oc_id, status_de, status_para, observacao, uid):
    await ajard_query(
        """INSERT INTO compras.oc_historico (oc_id, status_de, status_para, observacao, usuario_id)
           VALUES (%s,%s,%s,%s,%s)""",
        (oc_id, status_de, status_para, observacao, uid), fetch="none")


async def _recalcular_total(oc_id):
    t = await ajard_query(
        """SELECT COALESCE(SUM(quantidade*valor_unit),0) AS total
           FROM compras.oc_itens WHERE oc_id=%s AND ativo=true""",
        (oc_id,), fetch="one")
    total = t["total"] if t else 0
    await ajard_query(
        "UPDATE compras.ordens_compra SET valor_total=%s, atualizado_em=now() WHERE id=%s",
        (total, oc_id), fetch="none")
    return total


# ── SETORES (parametrização — Regra 63) ───────────────────────

@router.get("/compras/api/setores")
async def listar_setores(_auth=Depends(verificar_compras)):
    rows = await ajard_query(
        "SELECT * FROM compras.setores WHERE ativo=true ORDER BY nome")
    return [dict(r) for r in rows]


@router.post("/compras/api/setores")
async def criar_setor(request: Request, _auth=Depends(verificar_compras_gestor)):
    d = await request.json()
    codigo = (d.get("codigo") or "").strip().upper()
    nome = (d.get("nome") or "").strip()
    if not codigo or not nome:
        raise HTTPException(status_code=400, detail="Código e nome são obrigatórios")
    await ajard_query(
        """INSERT INTO compras.setores (codigo, nome, cor)
           VALUES (%s,%s,%s) ON CONFLICT (codigo) DO NOTHING""",
        (codigo, nome, d.get("cor")), fetch="none")
    return {"ok": True, "codigo": codigo}


@router.patch("/compras/api/setores/{codigo}")
async def editar_setor(codigo: str, request: Request,
                       _auth=Depends(verificar_compras_gestor)):
    d = await request.json()
    updates, params = [], []
    for c in ("nome", "cor", "ativo"):
        if c in d:
            updates.append(f"{c}=%s")
            params.append(d[c])
    if not updates:
        return {"ok": True}
    params.append(codigo)
    await ajard_query(
        f"UPDATE compras.setores SET {', '.join(updates)} WHERE codigo=%s",
        params, fetch="none")
    return {"ok": True}


# ── CONDIÇÕES DE PAGAMENTO (parametrização — Regra 63) ───────

@router.get("/compras/api/condicoes")
async def listar_condicoes(_auth=Depends(verificar_compras)):
    rows = await ajard_query(
        "SELECT * FROM compras.condicoes_pagamento WHERE ativo=true ORDER BY nome")
    return [dict(r) for r in rows]


@router.post("/compras/api/condicoes")
async def criar_condicao(request: Request, _auth=Depends(verificar_compras_gestor)):
    d = await request.json()
    codigo = (d.get("codigo") or "").strip().upper().replace(" ", "_")
    nome = (d.get("nome") or "").strip()
    if not codigo or not nome:
        raise HTTPException(status_code=400, detail="Código e nome são obrigatórios")
    await ajard_query(
        """INSERT INTO compras.condicoes_pagamento (codigo, nome)
           VALUES (%s,%s) ON CONFLICT (codigo) DO UPDATE SET nome=EXCLUDED.nome, ativo=true""",
        (codigo, nome), fetch="none")
    return {"ok": True, "codigo": codigo}


@router.patch("/compras/api/condicoes/{codigo}")
async def editar_condicao(codigo: str, request: Request,
                          _auth=Depends(verificar_compras_gestor)):
    d = await request.json()
    updates, params = [], []
    for c in ("nome", "ativo"):
        if c in d:
            updates.append(f"{c}=%s")
            params.append(d[c])
    if not updates:
        return {"ok": True}
    params.append(codigo)
    await ajard_query(
        f"UPDATE compras.condicoes_pagamento SET {', '.join(updates)} WHERE codigo=%s",
        params, fetch="none")
    return {"ok": True}


# ── ALÇADAS (parametrização) ──────────────────────────────────

@router.get("/compras/api/alcadas")
async def listar_alcadas(_auth=Depends(verificar_compras_gestor)):
    rows = await ajard_query(
        """SELECT a.*, u.nome AS usuario_nome, u.login AS usuario_login, u.perfil,
                  COALESCE((SELECT SUM(oc.valor_total) FROM compras.ordens_compra oc
                            WHERE oc.aprovador_id=a.usuario_id AND oc.ativo=true
                              AND oc.status NOT IN ('rejeitada','cancelada')
                              AND date_trunc('month', oc.data_aprovacao)=date_trunc('month', now())),0)
                    AS aprovado_mes
           FROM compras.alcadas a
           JOIN public.usuarios_garra u ON u.id = a.usuario_id
           ORDER BY u.nome""")
    return [dict(r) for r in rows]


@router.post("/compras/api/alcadas")
async def salvar_alcada(request: Request, _auth=Depends(verificar_compras_gestor)):
    """Cria/atualiza a alçada de um usuário.
    valor_limite null no body = sem limite; ativo=false desativa."""
    d = await request.json()
    uid = d.get("usuario_id")
    if not uid:
        raise HTTPException(status_code=400, detail="usuario_id é obrigatório")
    await ajard_query(
        """INSERT INTO compras.alcadas (usuario_id, valor_limite, limite_mensal, ativo)
           VALUES (%s,%s,%s,%s)
           ON CONFLICT (usuario_id)
           DO UPDATE SET valor_limite=EXCLUDED.valor_limite,
                         limite_mensal=EXCLUDED.limite_mensal,
                         ativo=EXCLUDED.ativo, atualizado_em=now()""",
        (uid, d.get("valor_limite"), d.get("limite_mensal"), d.get("ativo", True)), fetch="none")
    return {"ok": True}


# ── ORDENS DE COMPRA ──────────────────────────────────────────

@router.post("/compras/api/ocs")
async def criar_oc(request: Request, payload=Depends(verificar_compras_solicitante)):
    """Cria OC em rascunho (cotação). Itens no body: lista de
    {descricao, quantidade, unidade, valor_unit, peca_id?}."""
    d = await request.json()
    setor = (d.get("setor_codigo") or "").strip().upper()
    if not setor:
        raise HTTPException(status_code=400, detail="Setor é obrigatório")
    s = await ajard_query(
        "SELECT codigo FROM compras.setores WHERE codigo=%s AND ativo=true",
        (setor,), fetch="one")
    if not s:
        raise HTTPException(status_code=404, detail="Setor não encontrado")

    itens = d.get("itens") or []
    _validar_itens(itens)

    seq = await ajard_query(
        """SELECT COALESCE(MAX(sequencia),0)+1 AS n FROM compras.ordens_compra
           WHERE ano = EXTRACT(YEAR FROM now())::int""", fetch="one")
    from datetime import date as _date
    ano = _date.today().year
    numero = f"OC-{ano}-{int(seq['n']):04d}"

    uid = await _usuario_id(payload)
    row = await ajard_query_id(
        """INSERT INTO compras.ordens_compra
              (numero, ano, sequencia, setor_codigo, fornecedor_id, fornecedor_avulso, fornecedor_avulso_contato,
               ot_id, equipamento_id, prioridade, condicao_pagamento, observacao,
               solicitante_id)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (numero, ano, int(seq["n"]), setor, d.get("fornecedor_id"),
         (d.get("fornecedor_avulso") or "").strip() or None,
         (d.get("fornecedor_avulso_contato") or "").strip() or None,
         d.get("ot_id"), d.get("equipamento_id"),
         d.get("prioridade", "normal"), d.get("condicao_pagamento"),
         d.get("observacao"), uid))

    for i, it in enumerate(itens):
        desc = (it.get("descricao") or "").strip()
        if not desc:
            continue
        await ajard_query(
            """INSERT INTO compras.oc_itens
                  (oc_id, peca_id, descricao, quantidade, unidade, valor_unit, ordem)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (row["id"], it.get("peca_id"), desc,
             it.get("quantidade") or 1, it.get("unidade", "UN"),
             it.get("valor_unit") or 0, i), fetch="none")

    total = await _recalcular_total(row["id"])
    await _trilha(row["id"], None, "rascunho", "OC criada (cotação)", uid)
    out = dict(row)
    out["valor_total"] = total
    return out


@router.get("/compras/api/ocs")
async def listar_ocs(status: str = None, setor: str = None,
                     fornecedor_id: str = None, mes: str = None,
                     _auth=Depends(verificar_compras)):
    where, params = ["oc.ativo=true"], []
    if not await _ve_todas(_auth):
        uid = await _usuario_id(_auth)
        params.append(uid)
        where.append("oc.solicitante_id=%s")
    if status:
        params.append(status)
        where.append("oc.status=%s")
    if setor:
        params.append(setor.upper())
        where.append("oc.setor_codigo=%s")
    if fornecedor_id:
        params.append(fornecedor_id)
        where.append("oc.fornecedor_id=%s")
    if mes:  # 'YYYY-MM'
        params.append(mes)
        where.append("to_char(oc.criado_em,'YYYY-MM')=%s")
    rows = await ajard_query(
        f"""SELECT oc.*, se.nome AS setor_nome, se.cor AS setor_cor,
                  COALESCE(fo.nome, oc.fornecedor_avulso) AS fornecedor_nome,
                  us.nome AS solicitante_nome, ua.nome AS aprovador_nome,
                  ot.numero AS ot_numero,
                  eq.codigo AS equipamento_codigo
           FROM compras.ordens_compra oc
           JOIN compras.setores se ON se.codigo = oc.setor_codigo
           LEFT JOIN public.fornecedores fo ON fo.id = oc.fornecedor_id
           LEFT JOIN public.usuarios_garra us ON us.id = oc.solicitante_id
           LEFT JOIN public.usuarios_garra ua ON ua.id = oc.aprovador_id
           LEFT JOIN manutencao.ot ot ON ot.id = oc.ot_id
           LEFT JOIN operacional.equipamentos eq ON eq.id = oc.equipamento_id
           WHERE {' AND '.join(where)}
           ORDER BY oc.criado_em DESC""", params)
    return [dict(r) for r in rows]


@router.get("/compras/api/resumo")
async def resumo_compras(_auth=Depends(verificar_compras)):
    filtro, params = "", []
    if not await _ve_todas(_auth):
        uid = await _usuario_id(_auth)
        filtro = " AND solicitante_id=%s"
        params = [uid]
    r = await ajard_query(f"""
        SELECT
          COUNT(*) FILTER (WHERE status='solicitada')                    AS aguardando_aprovacao,
          COUNT(*) FILTER (WHERE status='aprovada')                      AS aprovadas,
          COUNT(*) FILTER (WHERE status IN ('enviada','recebida_parcial')) AS aguardando_entrega,
          COUNT(*) FILTER (WHERE status='recebida'
                           AND date_trunc('month',criado_em)=date_trunc('month',now())) AS recebidas_mes,
          COALESCE(SUM(valor_total) FILTER (WHERE status IN ('aprovada','enviada','recebida_parcial','recebida')
                           AND date_trunc('month',criado_em)=date_trunc('month',now())),0) AS valor_mes
        FROM compras.ordens_compra WHERE ativo=true{filtro}""", params, fetch="one")
    return dict(r)


@router.get("/compras/api/pendentes-aprovacao")
async def fila_aprovacao(payload=Depends(verificar_compras_aprovador)):
    """Fila do aprovador: OCs solicitadas dentro da alçada dele.
    Permissão de aprovar SEM alçada cadastrada = não aprova nada, então o
    gate nega (403) e a aba Aprovar nem aparece no app — em vez de exibir
    uma fila eternamente vazia (incoerência apontada em produção)."""
    tem, limite, limite_mensal = await _alcada_efetiva(payload)
    if not tem:
        raise HTTPException(
            status_code=403,
            detail="Você tem permissão de aprovar, mas ainda não tem alçada cadastrada — peça à gestão em Compras → Alçadas")
    where, params = ["oc.ativo=true", "oc.status='solicitada'"], []
    if limite is not None:
        params.append(limite)
        where.append("oc.valor_total<=%s")
    rows = await ajard_query(
        f"""SELECT oc.*, se.nome AS setor_nome, se.cor AS setor_cor,
                  COALESCE(fo.nome, oc.fornecedor_avulso) AS fornecedor_nome, us.nome AS solicitante_nome,
                  ot.numero AS ot_numero, eq.codigo AS equipamento_codigo
           FROM compras.ordens_compra oc
           JOIN compras.setores se ON se.codigo = oc.setor_codigo
           LEFT JOIN public.fornecedores fo ON fo.id = oc.fornecedor_id
           LEFT JOIN public.usuarios_garra us ON us.id = oc.solicitante_id
           LEFT JOIN manutencao.ot ot ON ot.id = oc.ot_id
           LEFT JOIN operacional.equipamentos eq ON eq.id = oc.equipamento_id
           WHERE {' AND '.join(where)}
           ORDER BY CASE oc.prioridade WHEN 'urgente' THEN 0 ELSE 1 END, oc.criado_em""",
        params)
    return [dict(r) for r in rows]


@router.get("/compras/api/ocs/{oc_id}")
async def detalhe_oc(oc_id: str, _auth=Depends(verificar_compras)):
    oc = await ajard_query(
        """SELECT oc.*, se.nome AS setor_nome, se.cor AS setor_cor,
                  COALESCE(fo.nome, oc.fornecedor_avulso) AS fornecedor_nome, fo.telefone AS fornecedor_telefone,
                  us.nome AS solicitante_nome, ua.nome AS aprovador_nome,
                  ot.numero AS ot_numero, eq.codigo AS equipamento_codigo
           FROM compras.ordens_compra oc
           JOIN compras.setores se ON se.codigo = oc.setor_codigo
           LEFT JOIN public.fornecedores fo ON fo.id = oc.fornecedor_id
           LEFT JOIN public.usuarios_garra us ON us.id = oc.solicitante_id
           LEFT JOIN public.usuarios_garra ua ON ua.id = oc.aprovador_id
           LEFT JOIN manutencao.ot ot ON ot.id = oc.ot_id
           LEFT JOIN operacional.equipamentos eq ON eq.id = oc.equipamento_id
           WHERE oc.id=%s AND oc.ativo=true""", (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if not await _ve_todas(_auth):
        uid = await _usuario_id(_auth)
        if str(oc.get("solicitante_id") or "") != str(uid):
            raise HTTPException(status_code=404, detail="OC não encontrada")
    itens = await ajard_query(
        """SELECT * FROM compras.oc_itens
           WHERE oc_id=%s AND ativo=true ORDER BY ordem""", (oc_id,))
    hist = await ajard_query(
        """SELECT h.*, u.nome AS usuario_nome
           FROM compras.oc_historico h
           LEFT JOIN public.usuarios_garra u ON u.id = h.usuario_id
           WHERE h.oc_id=%s ORDER BY h.criado_em""", (oc_id,))
    anexos = await ajard_query(
        """SELECT a.*, u.nome AS enviado_por_nome
           FROM compras.oc_anexos a
           LEFT JOIN public.usuarios_garra u ON u.id = a.enviado_por
           WHERE a.oc_id=%s AND a.ativo=true ORDER BY a.criado_em""", (oc_id,))
    d = dict(oc)
    d["itens"] = [dict(i) for i in itens]
    d["historico"] = [dict(h) for h in hist]
    # telefone para WhatsApp: cadastrado usa o do fornecedor; avulso usa o contato digitado
    if not d.get("fornecedor_telefone") and d.get("fornecedor_avulso_contato"):
        d["fornecedor_telefone"] = d["fornecedor_avulso_contato"]
    d["anexos"] = []
    for a in anexos:
        ax = dict(a)
        ax["url"] = storage_url(ax.get("caminho") or "")
        d["anexos"].append(ax)
    uid = await _usuario_id(_auth)
    eh_dono = bool(uid) and str(oc.get("solicitante_id") or "") == str(uid)
    eh_gestor = await _tem_permissao(_auth, "compras_aprovar")
    editavel = oc["status"] in ("rascunho", "solicitada")
    d["pode_editar"] = editavel and (eh_dono or eh_gestor)
    d["pode_excluir"] = editavel and (eh_dono or eh_gestor)
    return d


@router.patch("/compras/api/ocs/{oc_id}")
async def editar_oc(oc_id: str, request: Request, payload=Depends(verificar_compras)):
    """Edição só em rascunho ou solicitada. Se vier 'itens', substitui a lista."""
    d = await request.json()
    oc = await ajard_query(
        "SELECT id, status, solicitante_id FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if oc["status"] not in ("rascunho", "solicitada"):
        raise HTTPException(status_code=400,
                            detail=f"OC em '{oc['status']}' não pode mais ser editada")
    uid = await _usuario_id(payload)
    eh_dono = bool(uid) and str(oc.get("solicitante_id") or "") == str(uid)
    if not eh_dono and not await _tem_permissao(payload, "compras_aprovar"):
        raise HTTPException(status_code=403,
                            detail="Só quem solicitou (ou a gestão) pode editar esta OC")

    campos = ["setor_codigo", "fornecedor_id", "fornecedor_avulso", "ot_id",
              "equipamento_id", "prioridade", "condicao_pagamento", "observacao"]
    updates, params = [], []
    for c in campos:
        if c in d:
            updates.append(f"{c}=%s")
            params.append(d[c])
    if updates:
        params.append(oc_id)
        await ajard_query(
            f"UPDATE compras.ordens_compra SET {', '.join(updates)}, atualizado_em=now() WHERE id=%s",
            params, fetch="none")

    if "itens" in d:
        _validar_itens([i for i in (d["itens"] or []) if (i.get("descricao") or "").strip()])
        await ajard_query(
            "UPDATE compras.oc_itens SET ativo=false WHERE oc_id=%s",
            (oc_id,), fetch="none")
        for i, it in enumerate(d["itens"] or []):
            desc = (it.get("descricao") or "").strip()
            if not desc:
                continue
            await ajard_query(
                """INSERT INTO compras.oc_itens
                      (oc_id, peca_id, descricao, quantidade, unidade, valor_unit, ordem)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (oc_id, it.get("peca_id"), desc,
                 it.get("quantidade") or 1, it.get("unidade", "UN"),
                 it.get("valor_unit") or 0, i), fetch="none")
    await _recalcular_total(oc_id)
    return {"ok": True}


@router.post("/compras/api/ocs/{oc_id}/solicitar")
async def solicitar_aprovacao(oc_id: str, payload=Depends(verificar_compras)):
    """Rascunho → solicitada. Se o valor está dentro da alçada de quem
    solicita → aprova direto (1 passo), registrando as duas etapas na trilha."""
    oc = await ajard_query(
        "SELECT id, status, valor_total FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if "solicitada" not in _TRANSICOES.get(oc["status"], set()):
        raise HTTPException(status_code=400,
                            detail=f"Transição inválida: {oc['status']} → solicitada")
    uid = await _usuario_id(payload)
    await ajard_query(
        "UPDATE compras.ordens_compra SET status='solicitada', atualizado_em=now() WHERE id=%s",
        (oc_id,), fetch="none")
    await _trilha(oc_id, oc["status"], "solicitada", "Aprovação solicitada", uid)

    tem, limite, limite_mensal = await _alcada_efetiva(payload)
    if _valor_dentro_alcada(oc["valor_total"], tem, limite):
        cabe, _ja, _saldo = await _cabe_no_teto_mensal(uid, oc["valor_total"], limite_mensal)
        if cabe:
            await ajard_query(
                """UPDATE compras.ordens_compra
                   SET status='aprovada', aprovador_id=%s, data_aprovacao=now(),
                       atualizado_em=now() WHERE id=%s""",
                (uid, oc_id), fetch="none")
            await _trilha(oc_id, "solicitada", "aprovada",
                          "Aprovada em 1 passo (dentro da alçada)", uid)
            return {"ok": True, "status": "aprovada", "auto_aprovada": True}
        # Teto mensal esgotado → segue pra fila de quem pode (sem erro)
        await _trilha(oc_id, "solicitada", "solicitada",
                      "Teto mensal do solicitante atingido — encaminhada para aprovação", uid)
    return {"ok": True, "status": "solicitada", "auto_aprovada": False}


@router.post("/compras/api/ocs/{oc_id}/aprovar")
async def aprovar_oc(oc_id: str, request: Request,
                     payload=Depends(verificar_compras_aprovador)):
    d = {}
    try:
        d = await request.json()
    except Exception:
        pass
    oc = await ajard_query(
        "SELECT id, status, valor_total FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if "aprovada" not in _TRANSICOES.get(oc["status"], set()):
        raise HTTPException(status_code=400,
                            detail=f"Transição inválida: {oc['status']} → aprovada")
    uid = await _usuario_id(payload)
    tem, limite, limite_mensal = await _alcada_efetiva(payload)
    if not tem:
        raise HTTPException(
            status_code=403,
            detail="Você ainda não tem alçada cadastrada — cadastre em ⚙️ Alçadas (gestão)")
    if not _valor_dentro_alcada(oc["valor_total"], tem, limite):
        raise HTTPException(
            status_code=403,
            detail=f"Valor acima da sua alçada de aprovação (limite R$ {float(limite):,.2f} por compra)")
    cabe, ja, saldo = await _cabe_no_teto_mensal(uid, oc["valor_total"], limite_mensal)
    if not cabe:
        raise HTTPException(
            status_code=403,
            detail=(f"Teto mensal atingido: seu limite é R$ {float(limite_mensal):,.2f}/mês, "
                    f"já aprovado R$ {ja:,.2f} — saldo R$ {max(saldo,0):,.2f}. "
                    "Esta OC segue para aprovação da gestão."))
    await ajard_query(
        """UPDATE compras.ordens_compra
           SET status='aprovada', aprovador_id=%s, data_aprovacao=now(),
               atualizado_em=now() WHERE id=%s""",
        (uid, oc_id), fetch="none")
    await _trilha(oc_id, oc["status"], "aprovada",
                  (d.get("observacao") or "").strip() or "OC aprovada", uid)
    return {"ok": True, "status": "aprovada"}


@router.post("/compras/api/ocs/{oc_id}/rejeitar")
async def rejeitar_oc(oc_id: str, request: Request,
                      payload=Depends(verificar_compras_aprovador)):
    d = await request.json()
    motivo = (d.get("motivo") or "").strip()
    if not motivo:
        raise HTTPException(status_code=400, detail="Motivo da rejeição é obrigatório")
    oc = await ajard_query(
        "SELECT id, status FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if "rejeitada" not in _TRANSICOES.get(oc["status"], set()):
        raise HTTPException(status_code=400,
                            detail=f"Transição inválida: {oc['status']} → rejeitada")
    uid = await _usuario_id(payload)
    await ajard_query(
        """UPDATE compras.ordens_compra
           SET status='rejeitada', motivo_rejeicao=%s, atualizado_em=now() WHERE id=%s""",
        (motivo, oc_id), fetch="none")
    await _trilha(oc_id, oc["status"], "rejeitada", motivo, uid)
    return {"ok": True, "status": "rejeitada"}


@router.post("/compras/api/ocs/{oc_id}/enviar")
async def marcar_enviada(oc_id: str, payload=Depends(verificar_compras)):
    """Marca a OC como enviada ao fornecedor (após gerar PDF / link WhatsApp)."""
    oc = await ajard_query(
        "SELECT id, status, solicitante_id FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if not await _pode_mexer_na_oc(payload, oc):
        raise HTTPException(status_code=403, detail="Só quem solicitou (ou a gestão) envia esta OC")
    if "enviada" not in _TRANSICOES.get(oc["status"], set()):
        raise HTTPException(status_code=400,
                            detail=f"Transição inválida: {oc['status']} → enviada")
    uid = await _usuario_id(payload)
    await ajard_query(
        """UPDATE compras.ordens_compra
           SET status='enviada', enviado_por=%s, enviado_em=now(),
               atualizado_em=now() WHERE id=%s""",
        (uid, oc_id), fetch="none")
    await _trilha(oc_id, oc["status"], "enviada", "Enviada ao fornecedor", uid)
    return {"ok": True, "status": "enviada"}


@router.post("/compras/api/ocs/{oc_id}/receber")
async def receber_oc(oc_id: str, request: Request, payload=Depends(verificar_compras)):
    """Recebimento item a item. Body:
    { nf_numero?, itens: [{item_id, qtd_recebida}] } — qtd é o RECEBIDO AGORA
    (acumula). Todas completas → 'recebida'; senão → 'recebida_parcial'.
    OC com ot_id: o valor recebido agora soma no custo_total da OT."""
    d = await request.json()
    oc = await ajard_query(
        "SELECT id, status, ot_id, solicitante_id FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if not await _pode_mexer_na_oc(payload, oc):
        raise HTTPException(status_code=403, detail="Só quem solicitou (ou a gestão) registra o recebimento")
    if oc["status"] not in ("enviada", "recebida_parcial"):
        raise HTTPException(status_code=400,
                            detail=f"OC em '{oc['status']}' não está aguardando entrega")

    uid = await _usuario_id(payload)

    # ── Itens complementares (vieram fora da OC) — régua do limite ──
    extras = d.get("itens_extras") or []
    valor_extras = 0.0
    extras_validos = []
    for ex in extras:
        desc = (ex.get("descricao") or "").strip()
        qtd = float(ex.get("quantidade") or 0)
        vu = float(ex.get("valor_unit") or 0)
        if not desc or qtd <= 0:
            continue
        valor_extras += qtd * vu
        extras_validos.append((desc, qtd, (ex.get("unidade") or "UN").strip() or "UN", vu))
    if extras_validos:
        oc_val = await ajard_query(
            "SELECT COALESCE(valor_total,0) AS v FROM compras.ordens_compra WHERE id=%s",
            (oc_id,), fetch="one")
        limite = max(_COMPLEMENTO_MIN, float(oc_val["v"]) * _COMPLEMENTO_PCT)
        if valor_extras > limite + 1e-9:
            raise HTTPException(
                status_code=400,
                detail=(f"Complemento de R$ {valor_extras:,.2f} acima do limite "
                        f"(R$ {limite:,.2f} = maior entre R$ {_COMPLEMENTO_MIN:,.0f} "
                        f"e {int(_COMPLEMENTO_PCT*100)}% da OC) — abra uma OC nova para estes itens"))

    valor_recebido_agora = 0.0
    for it in d.get("itens") or []:
        item = await ajard_query(
            """SELECT id, quantidade, qtd_recebida, valor_unit
               FROM compras.oc_itens WHERE id=%s AND oc_id=%s AND ativo=true""",
            (it.get("item_id"), oc_id), fetch="one")
        if not item:
            continue
        qtd_agora = float(it.get("qtd_recebida") or 0)
        if qtd_agora <= 0:
            continue
        await ajard_query(
            """UPDATE compras.oc_itens
               SET qtd_recebida = COALESCE(qtd_recebida,0)+%s WHERE id=%s""",
            (qtd_agora, item["id"]), fetch="none")
        valor_recebido_agora += qtd_agora * float(item.get("valor_unit") or 0)

    if extras_validos:
        maxo = await ajard_query(
            "SELECT COALESCE(MAX(ordem),0) AS m FROM compras.oc_itens WHERE oc_id=%s",
            (oc_id,), fetch="one")
        ordem = int(maxo["m"])
        for desc, qtd, un, vu in extras_validos:
            ordem += 1
            await ajard_query(
                """INSERT INTO compras.oc_itens
                     (oc_id, descricao, quantidade, unidade, valor_unit, qtd_recebida, ordem)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (oc_id, desc, qtd, un, vu, qtd, ordem), fetch="none")
            valor_recebido_agora += qtd * vu
        # valor_total da OC recalculado com os complementos
        await ajard_query(
            """UPDATE compras.ordens_compra
               SET valor_total = (SELECT COALESCE(SUM(quantidade*valor_unit),0)
                                    FROM compras.oc_itens
                                   WHERE oc_id=%s AND ativo=true)
             WHERE id=%s""", (oc_id, oc_id), fetch="none")
        detalhe_ext = "; ".join(f"{q:g} {u} {dsc} (R$ {q*v:,.2f})" for dsc, q, u, v in extras_validos)
        await _trilha(oc_id, oc["status"], oc["status"],
                      f"Complemento no recebimento (R$ {valor_extras:,.2f}): {detalhe_ext}", uid)

    if d.get("nf_numero"):
        await ajard_query(
            "UPDATE compras.ordens_compra SET nf_numero=%s WHERE id=%s",
            ((d.get("nf_numero") or "").strip(), oc_id), fetch="none")

    pend = await ajard_query(
        """SELECT COUNT(*) AS n FROM compras.oc_itens
           WHERE oc_id=%s AND ativo=true
             AND COALESCE(qtd_recebida,0) < quantidade""",
        (oc_id,), fetch="one")
    novo = "recebida" if int(pend["n"]) == 0 else "recebida_parcial"
    await ajard_query(
        "UPDATE compras.ordens_compra SET status=%s, atualizado_em=now() WHERE id=%s",
        (novo, oc_id), fetch="none")
    obs_receb = (d.get("observacao") or "").strip()
    msg_trilha = f"Recebimento registrado (R$ {valor_recebido_agora:.2f})"
    if obs_receb:
        msg_trilha += f" — {obs_receb}"
    await _trilha(oc_id, oc["status"], novo, msg_trilha, uid)

    if oc.get("ot_id") and valor_recebido_agora > 0:
        await ajard_query(
            """UPDATE manutencao.ot
               SET custo_total = COALESCE(custo_total,0)+%s, atualizado_em=now()
               WHERE id=%s""",
            (valor_recebido_agora, oc["ot_id"]), fetch="none")

    return {"ok": True, "status": novo, "valor_recebido": valor_recebido_agora}


@router.post("/compras/api/ocs/{oc_id}/cancelar")
async def cancelar_oc(oc_id: str, request: Request,
                      payload=Depends(verificar_compras_aprovador)):
    d = {}
    try:
        d = await request.json()
    except Exception:
        pass
    oc = await ajard_query(
        "SELECT id, status FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if "cancelada" not in _TRANSICOES.get(oc["status"], set()):
        raise HTTPException(status_code=400,
                            detail=f"Transição inválida: {oc['status']} → cancelada")
    uid = await _usuario_id(payload)
    await ajard_query(
        "UPDATE compras.ordens_compra SET status='cancelada', atualizado_em=now() WHERE id=%s",
        (oc_id,), fetch="none")
    await _trilha(oc_id, oc["status"], "cancelada",
                  (d.get("motivo") or "").strip() or "OC cancelada", uid)
    return {"ok": True, "status": "cancelada"}


@router.delete("/compras/api/ocs/{oc_id}")
async def excluir_oc(oc_id: str, payload=Depends(verificar_compras)):
    """Exclusão (soft delete) pelo próprio solicitante enquanto a OC ainda
    não foi aprovada (rascunho/solicitada), ou pela gestão (compras_aprovar).
    A OC some das listas mas permanece no banco com trilha."""
    oc = await ajard_query(
        "SELECT id, numero, status, solicitante_id FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if oc["status"] not in ("rascunho", "solicitada"):
        raise HTTPException(status_code=400,
                            detail=f"OC em '{oc['status']}' não pode ser excluída — use Cancelar")
    uid = await _usuario_id(payload)
    eh_dono = bool(uid) and str(oc.get("solicitante_id") or "") == str(uid)
    if not eh_dono and not await _tem_permissao(payload, "compras_aprovar"):
        raise HTTPException(status_code=403,
                            detail="Só quem solicitou (ou a gestão) pode excluir esta OC")
    await ajard_query(
        "UPDATE compras.ordens_compra SET ativo=false, atualizado_em=now() WHERE id=%s",
        (oc_id,), fetch="none")
    await _trilha(oc_id, oc["status"], "excluida", "OC excluída pelo solicitante/gestão", uid)
    return {"ok": True}


# ── DEVOLUÇÃO (parte do ciclo da compra, não sai do âmbito da OC) ─
# "Pedi X, recebi X, devolvi Y, fiquei com Z" — a OC responde tudo.
# O crédito/estorno financeiro é assunto de contas a pagar; aqui fica
# o registro: itens devolvidos, motivo obrigatório, NF de devolução,
# e o estorno automático do custo na OT vinculada.

@router.post("/compras/api/ocs/{oc_id}/devolver")
async def devolver_itens(oc_id: str, request: Request, payload=Depends(verificar_compras)):
    d = await request.json()
    motivo = (d.get("motivo") or "").strip()
    if not motivo:
        raise HTTPException(status_code=400, detail="Motivo da devolução é obrigatório")
    oc = await ajard_query(
        """SELECT id, numero, status, ot_id, solicitante_id
           FROM compras.ordens_compra WHERE id=%s AND ativo=true""",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if oc["status"] not in ("recebida", "recebida_parcial"):
        raise HTTPException(status_code=400,
                            detail="Devolução só se registra em OC recebida — antes disso, basta não receber o item")
    if not await _pode_mexer_na_oc(payload, oc):
        raise HTTPException(status_code=403, detail="Só quem solicitou (ou a gestão) registra devolução")

    uid = await _usuario_id(payload)
    valor_devolvido = 0.0
    alguma = False
    for it in d.get("itens") or []:
        item = await ajard_query(
            """SELECT id, qtd_recebida, qtd_devolvida, valor_unit
               FROM compras.oc_itens WHERE id=%s AND oc_id=%s AND ativo=true""",
            (it.get("item_id"), oc_id), fetch="one")
        if not item:
            continue
        qtd = float(it.get("qtd_devolvida") or 0)
        if qtd <= 0:
            continue
        disponivel = float(item.get("qtd_recebida") or 0) - float(item.get("qtd_devolvida") or 0)
        if qtd > disponivel + 1e-9:
            raise HTTPException(status_code=400,
                                detail=f"Devolução maior que o recebido disponível ({disponivel:g})")
        await ajard_query(
            """UPDATE compras.oc_itens
               SET qtd_devolvida = COALESCE(qtd_devolvida,0)+%s WHERE id=%s""",
            (qtd, item["id"]), fetch="none")
        valor_devolvido += qtd * float(item.get("valor_unit") or 0)
        alguma = True
    if not alguma:
        raise HTTPException(status_code=400, detail="Informe ao menos 1 item com quantidade devolvida")

    nf_dev = (d.get("nf_devolucao") or "").strip() or None
    if nf_dev:
        await ajard_query(
            "UPDATE compras.ordens_compra SET nf_devolucao=%s WHERE id=%s",
            (nf_dev, oc_id), fetch="none")

    # Tudo que foi recebido foi devolvido? → status final 'devolvida'
    resto = await ajard_query(
        """SELECT COUNT(*) AS n FROM compras.oc_itens
           WHERE oc_id=%s AND ativo=true
             AND COALESCE(qtd_recebida,0) - COALESCE(qtd_devolvida,0) > 0""",
        (oc_id,), fetch="one")
    novo = "devolvida" if int(resto["n"]) == 0 else oc["status"]
    await ajard_query(
        "UPDATE compras.ordens_compra SET status=%s, atualizado_em=now() WHERE id=%s",
        (novo, oc_id), fetch="none")
    await _trilha(oc_id, oc["status"], novo,
                  f"Devolução registrada (R$ {valor_devolvido:.2f}): {motivo}"
                  + (f" · NF dev. {nf_dev}" if nf_dev else ""), uid)

    # Estorno automático do custo na OT vinculada
    if oc.get("ot_id") and valor_devolvido > 0:
        await ajard_query(
            """UPDATE manutencao.ot
               SET custo_total = GREATEST(COALESCE(custo_total,0)-%s, 0), atualizado_em=now()
               WHERE id=%s""",
            (valor_devolvido, oc["ot_id"]), fetch="none")

    return {"ok": True, "status": novo, "valor_devolvido": valor_devolvido}


# ── ANEXOS (orçamentos de fornecedores, fotos, PDFs) ──────────
# Cotação anexada à OC vira evidência com data: quem aprova vê o orçamento
# real; divergência de faturamento se resolve com documento na mão.

# ── RÉGUA DO COMPLEMENTO NO RECEBIMENTO (decisão gestão 03/08/2026) ──
# Item que veio fora da OC pode entrar no recebimento como COMPLEMENTO
# até um limite: R$ 100 OU 20% do valor da OC (o que for MAIOR).
# Acima disso → OC nova (é outra decisão de compra, não complemento).
_COMPLEMENTO_MIN = 100.0    # piso em reais (dá folga em OCs pequenas)
_COMPLEMENTO_PCT = 0.20     # percentual do valor da OC (folga em OCs grandes)

_ANEXO_TIPOS = {"application/pdf", "image/jpeg", "image/png", "image/webp"}
_ANEXO_MAX = 10 * 1024 * 1024  # 10 MB


async def _pode_mexer_na_oc(payload, oc):
    uid = await _usuario_id(payload)
    if bool(uid) and str(oc.get("solicitante_id") or "") == str(uid):
        return True
    return await _ve_todas(payload)


@router.post("/compras/api/ocs/{oc_id}/anexos")
async def anexar_arquivo(oc_id: str, arquivo: UploadFile = File(...),
                         payload=Depends(verificar_compras)):
    oc = await ajard_query(
        "SELECT id, status, solicitante_id FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if oc["status"] in ("rejeitada", "cancelada"):
        raise HTTPException(status_code=400, detail="OC encerrada não recebe anexos")
    if not await _pode_mexer_na_oc(payload, oc):
        raise HTTPException(status_code=403, detail="Só quem solicitou (ou a gestão) anexa nesta OC")

    ctype = (arquivo.content_type or "").lower()
    if ctype not in _ANEXO_TIPOS:
        raise HTTPException(status_code=400,
                            detail="Tipo não aceito — envie PDF ou imagem (JPG/PNG/WebP)")
    dados = await arquivo.read()
    if len(dados) > _ANEXO_MAX:
        raise HTTPException(status_code=400, detail="Arquivo acima de 10 MB")
    if not dados:
        raise HTTPException(status_code=400, detail="Arquivo vazio")

    import re as _re, uuid as _uuid
    nome = (arquivo.filename or "anexo").strip()
    nome_seguro = _re.sub(r"[^A-Za-z0-9._-]", "_", nome)[:80]
    caminho = f"compras/{oc_id}/{_uuid.uuid4().hex[:8]}-{nome_seguro}"
    try:
        caminho_salvo = storage_upload(dados, caminho, content_type=ctype)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Falha no upload do anexo: {e}")

    uid = await _usuario_id(payload)
    row = await ajard_query_id(
        """INSERT INTO compras.oc_anexos (oc_id, nome, caminho, content_type, tamanho, enviado_por)
           VALUES (%s,%s,%s,%s,%s,%s)""",
        (oc_id, nome, caminho_salvo, ctype, len(dados), uid))
    await _trilha(oc_id, oc["status"], oc["status"], f"Anexo adicionado: {nome}", uid)
    out = dict(row)
    out["url"] = storage_url(caminho_salvo)
    return out


@router.delete("/compras/api/ocs/{oc_id}/anexos/{anexo_id}")
async def excluir_anexo(oc_id: str, anexo_id: str, payload=Depends(verificar_compras)):
    oc = await ajard_query(
        "SELECT id, status, solicitante_id FROM compras.ordens_compra WHERE id=%s AND ativo=true",
        (oc_id,), fetch="one")
    if not oc:
        raise HTTPException(status_code=404, detail="OC não encontrada")
    if oc["status"] == "recebida":
        raise HTTPException(status_code=400,
                            detail="OC recebida está congelada — anexos não podem ser removidos")
    ax = await ajard_query(
        "SELECT id, enviado_por, nome FROM compras.oc_anexos WHERE id=%s AND oc_id=%s AND ativo=true",
        (anexo_id, oc_id), fetch="one")
    if not ax:
        raise HTTPException(status_code=404, detail="Anexo não encontrado")
    uid = await _usuario_id(payload)
    eh_autor = bool(uid) and str(ax.get("enviado_por") or "") == str(uid)
    if not eh_autor and not await _tem_permissao(payload, "compras_aprovar"):
        raise HTTPException(status_code=403, detail="Só quem enviou (ou a gestão) remove o anexo")
    await ajard_query(
        "UPDATE compras.oc_anexos SET ativo=false WHERE id=%s", (anexo_id,), fetch="none")
    await _trilha(oc_id, oc["status"], oc["status"], f"Anexo removido: {ax['nome']}", uid)
    return {"ok": True}


# ── INTELIGÊNCIA: último preço pago ───────────────────────────

@router.get("/compras/api/ultimo-preco")
async def ultimo_preco(q: str = "", _auth=Depends(verificar_compras)):
    """Histórico de preço por descrição (OCs recebidas). Alimenta a dica
    'Último preço: R$ X · FORNECEDOR · data' ao digitar o item."""
    q = (q or "").strip()
    if len(q) < 3:
        return []
    rows = await ajard_query(
        """SELECT i.descricao, i.valor_unit, i.unidade,
                  COALESCE(fo.nome, oc.fornecedor_avulso) AS fornecedor_nome, oc.criado_em::date AS data_compra
           FROM compras.oc_itens i
           JOIN compras.ordens_compra oc ON oc.id = i.oc_id
           LEFT JOIN public.fornecedores fo ON fo.id = oc.fornecedor_id
           WHERE i.ativo=true AND oc.ativo=true
             AND oc.status IN ('recebida','recebida_parcial')
             AND i.descricao ILIKE %s
           ORDER BY oc.criado_em DESC LIMIT 5""",
        (f"%{q}%",))
    return [dict(r) for r in rows]


# ── RELATÓRIOS (período flexível × status × setor → Excel/JSON) ─
# PDF é gerado no cliente pela janela de impressão (mesmo motor do
# documento da OC) — sem dependência nova no deploy.

async def _dados_relatorio(payload, inicio, fim, status, setor):
    from datetime import date as _date
    hoje = _date.today()
    if not inicio:
        inicio = hoje.replace(day=1).isoformat()
    if not fim:
        fim = hoje.isoformat()
    where, params = ["oc.ativo=true", "oc.criado_em::date BETWEEN %s AND %s"], [inicio, fim]
    if status:
        params.append(status)
        where.append("oc.status=%s")
    if setor:
        params.append(setor.upper())
        where.append("oc.setor_codigo=%s")
    if not await _ve_todas(payload):
        uid = await _usuario_id(payload)
        params.append(uid)
        where.append("oc.solicitante_id=%s")
    linhas = await ajard_query(
        f"""SELECT oc.numero, oc.criado_em::date AS data, se.nome AS setor,
                  oc.status, oc.prioridade,
                  COALESCE(fo.nome, oc.fornecedor_avulso, '-') AS fornecedor,
                  us.nome AS solicitante, ua.nome AS aprovador,
                  oc.nf_numero, ot.numero AS ot_numero, oc.valor_total
           FROM compras.ordens_compra oc
           JOIN compras.setores se ON se.codigo = oc.setor_codigo
           LEFT JOIN public.fornecedores fo ON fo.id = oc.fornecedor_id
           LEFT JOIN public.usuarios_garra us ON us.id = oc.solicitante_id
           LEFT JOIN public.usuarios_garra ua ON ua.id = oc.aprovador_id
           LEFT JOIN manutencao.ot ot ON ot.id = oc.ot_id
           WHERE {' AND '.join(where)}
           ORDER BY oc.criado_em""", params)
    linhas = [dict(l) for l in linhas]
    por_setor, por_status = {}, {}
    total = 0.0
    for l in linhas:
        v = float(l.get("valor_total") or 0)
        total += v
        por_setor[l["setor"]] = por_setor.get(l["setor"], 0) + v
        por_status[l["status"]] = por_status.get(l["status"], 0) + v
    return {"inicio": inicio, "fim": fim, "linhas": linhas, "total": total,
            "por_setor": por_setor, "por_status": por_status}


@router.get("/compras/api/relatorio")
async def relatorio_compras(inicio: str = None, fim: str = None,
                            status: str = None, setor: str = None,
                            formato: str = "json",
                            payload=Depends(verificar_compras)):
    d = await _dados_relatorio(payload, inicio, fim, status, setor)
    if formato != "excel":
        return d

    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1A2A5E"
    fino = Side(style="thin", color="CBD5E1")
    borda = Border(left=fino, right=fino, top=fino, bottom=fino)
    hfill = PatternFill("solid", fgColor=NAVY)
    hfont = Font(bold=True, color="FFFFFF", size=10)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCs"
    ws.append([f"ORDENS DE COMPRA — {d['inicio']} a {d['fim']}"
               + (f" · Status: {status}" if status else "")
               + (f" · Setor: {setor}" if setor else "")])
    ws["A1"].font = Font(bold=True, size=12, color=NAVY)
    cab = ["Nº OC", "Data", "Setor", "Status", "Prioridade", "Fornecedor",
           "Solicitante", "Aprovador", "NF", "OT", "Valor (R$)"]
    ws.append(cab)
    for i in range(1, len(cab) + 1):
        c = ws.cell(row=2, column=i)
        c.fill = hfill; c.font = hfont; c.border = borda
        c.alignment = Alignment(horizontal="center")
    for l in d["linhas"]:
        ws.append([l["numero"], str(l["data"]), l["setor"], l["status"],
                   l["prioridade"], l["fornecedor"], l["solicitante"],
                   l["aprovador"] or "", l["nf_numero"] or "", l["ot_numero"] or "",
                   float(l["valor_total"] or 0)])
    ult = ws.max_row + 1
    ws.cell(row=ult, column=10, value="TOTAL").font = Font(bold=True, color=NAVY)
    tc = ws.cell(row=ult, column=11, value=d["total"])
    tc.font = Font(bold=True, color=NAVY)
    for row in ws.iter_rows(min_row=3, max_row=ult, max_col=11):
        for c in row:
            c.border = borda
            if c.column == 11:
                c.number_format = '#,##0.00'
    larguras = [15, 11, 14, 15, 11, 24, 16, 16, 12, 14, 13]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Resumo")
    ws2.append(["RESUMO POR SETOR"]); ws2["A1"].font = Font(bold=True, color=NAVY)
    ws2.append(["Setor", "Valor (R$)"])
    for i in (1, 2):
        c = ws2.cell(row=2, column=i); c.fill = hfill; c.font = hfont; c.border = borda
    for s, v in sorted(d["por_setor"].items()):
        ws2.append([s, v])
    ws2.append([])
    ini = ws2.max_row + 1
    ws2.append(["RESUMO POR STATUS"]); ws2.cell(row=ini, column=1).font = Font(bold=True, color=NAVY)
    ws2.append(["Status", "Valor (R$)"])
    for i in (1, 2):
        c = ws2.cell(row=ini + 1, column=i); c.fill = hfill; c.font = hfont; c.border = borda
    for s, v in sorted(d["por_status"].items()):
        ws2.append([s, v])
    for row in ws2.iter_rows(min_row=2, max_col=2):
        for c in row:
            if c.value is not None:
                c.border = borda
                if c.column == 2 and isinstance(c.value, (int, float)):
                    c.number_format = '#,##0.00'
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"relatorio-compras-{d['inicio']}-a-{d['fim']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'})


# ── PÁGINA DO MÓDULO (desktop + mobile, SSO — Regra 60) ───────

@router.get("/compras", response_class=HTMLResponse)
async def pagina_compras():
    raiz = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "compras"))
    candidatos = [
        os.path.join(raiz, "static", "compras.html"),
        os.path.join(raiz, "compras.html"),
    ]
    for p in candidatos:
        if os.path.isfile(p):
            # no-store: navegador NUNCA cacheia o HTML da página — atualização
            # de permissão/gates chega ao aparelho na hora (lição do PWA:
            # "fix não chega ao aparelho" por cache heurístico do Chrome).
            return HTMLResponse(
                open(p, encoding="utf-8").read(),
                headers={"Cache-Control": "no-store, no-cache, must-revalidate",
                         "Pragma": "no-cache"})
    raise HTTPException(status_code=404, detail="Página do módulo Compras não encontrada")
