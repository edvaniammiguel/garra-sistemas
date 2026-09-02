# -*- coding: utf-8 -*-
"""routers.abastecimentos — FASE A (31/08/2026): a NOTA é a mãe.

Fluxo real da Garra: o colaborador fotografa a notinha do posto (às vezes a
foto do horímetro e/ou um texto como mandaria no grupo). O sistema lê tudo e
monta o registro; ele confere e confirma.

Modelo:
  operacional.abastecimento_notas  — cabeçalho (posto=public.fornecedores por
      CNPJ, cupom único por posto, combustível→rubrica herdada, litros/valor,
      fotos, texto livre, JSON bruto da leitura)
  operacional.abastecimentos       — itens: um por equipamento OU galão.
      Rateio fecha com o total da nota (400 se não fechar). Leitura por item
      (cascata digitada → foto → última → cadastro), divergência nunca
      bloqueia, cadastro nunca recua. Galão = equipamento da frota única com
      categoria 'galao'; item destino galão não tem leitura; saída do galão
      para máquina nasce como item origem='galao' e debita o saldo.
  manutencao.combustiveis          — domínio parametrizado (codigo, nome,
      rubrica, ativo). Rubrica NUNCA é digitada: herdada do combustível.

Inteligência embutida no registrar (itens 1–4 da conversa de 31/08):
  consumo_medio do item (L/h ou L/100 km desde o abastecimento anterior),
  alerta_consumo (> 25% acima do CMD-R do próprio equipamento, mín. 3
  abastecimentos), sem_funcionamento (nenhuma parte diária/checklist entre o
  abastecimento anterior e este).

Rodapé ManWinWin (Fase 1 = réplica fiel) em GET .../resumo/{eq_id}:
  Último Registo · FMD-R · CMD-R · Distância — janela padrão 90 dias.

Montagem no main.py (inalterada): include_router(abastecimentos_router).
Env: ANTHROPIC_API_KEY (sem ela o módulo funciona 100% manual).
"""

import os, re, json, base64, asyncio, uuid as uuid_lib
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse

from core.db import ajard_query
from core.auth import verificar_token
from core.storage import storage_upload, storage_url as _storage_url_core


def storage_url(path, segundos=None):
    """Compatibilidade: notas gravadas entre a virada e 01/09/2026 têm o caminho
    sem o prefixo do bucket ('abastecimentos/2026-08/nota-x.jpg'). Elas estão no
    bucket unificado — prefixa antes de assinar. `segundos` opcional (Excel usa 7 dias)."""
    if path and ":" not in path and not path.startswith("http"):
        path = "garra-fotos:" + path
    return _storage_url_core(path, segundos) if segundos else _storage_url_core(path)

router = APIRouter()

_PERFIS_LIVRES = {"admin", "gestor", "bruna", "luana"}

async def verificar_abastecimento(payload=Depends(verificar_token)):
    """Gate do módulo 'abastecimento' — matriz Por Perfil + exceções
    individuais (mesmo resolvedor oficial do Manutenção)."""
    if (payload.get("perfil") or "").lower() in _PERFIS_LIVRES:
        return payload
    from routers.manutencao import _tem_modulo
    if await _tem_modulo(payload, {"abastecimento"}):
        return payload
    raise HTTPException(status_code=403, detail="Sem permissão para Abastecimento")

async def verificar_notas(payload=Depends(verificar_token)):
    """(01/09/2026) Gate da conferência financeira das notas — o balcão da
    Luana: vê fotos e valores e marca 'lançada na MAIS' SEM entrar na
    Manutenção. Perfis livres + módulo abastecimento_notas (ou manutencao —
    Bruna também enxerga)."""
    if (payload.get("perfil") or "").lower() in _PERFIS_LIVRES:
        return payload
    from routers.manutencao import _tem_modulo
    if await _tem_modulo(payload, {"abastecimento_notas", "manutencao"}):
        return payload
    raise HTTPException(status_code=403, detail="Sem permissão para Notas de Abastecimento")


ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODELO_VISAO = os.environ.get("ABASTECIMENTO_MODELO", "claude-haiku-4-5-20251001")
JANELA_DIAS = int(os.environ.get("ABASTECIMENTO_JANELA_DIAS", "90") or 90)
TOLERANCIA_RATEIO_L = 0.05
FATOR_ALERTA_CONSUMO = 1.25
MIN_ABAST_PARA_ALERTA = 3

# ── DDL idempotente ─────────────────────────────────────────────────
_DDL_OK = False

_COLUNAS_ITEM = [
    ("equipamento_id", "UUID"),
    ("data", "TIMESTAMPTZ DEFAULT now()"),
    ("litros", "NUMERIC(10,2)"),
    ("valor_total", "NUMERIC(12,2)"),
    ("leitura", "NUMERIC(12,1)"),
    ("leitura_fonte", "TEXT"),
    ("medicao", "TEXT"),
    ("leitura_digitada", "NUMERIC(12,1)"),
    ("leitura_foto", "NUMERIC(12,1)"),
    ("litros_foto", "NUMERIC(10,2)"),
    ("valor_foto", "NUMERIC(12,2)"),
    ("placa_foto", "TEXT"),
    ("divergencia_leitura", "BOOLEAN DEFAULT false"),
    ("divergencia_placa", "BOOLEAN DEFAULT false"),
    ("foto_painel", "TEXT"),
    ("foto_bomba", "TEXT"),
    ("usuario_id", "UUID"),
    ("usuario_nome", "TEXT"),
    ("observacao", "TEXT"),
    ("criado_em", "TIMESTAMPTZ DEFAULT now()"),
    ("ativo", "BOOLEAN DEFAULT true"),
    # fase A
    ("nota_id", "UUID"),
    ("origem", "TEXT DEFAULT 'nota'"),          # nota | galao
    ("galao_id", "UUID"),                        # quando origem='galao'
    ("destino_tipo", "TEXT DEFAULT 'equipamento'"),  # equipamento | galao
    ("combustivel", "TEXT"),
    ("rubrica", "TEXT"),
    ("centro_custo", "TEXT"),
    ("preco_litro", "NUMERIC(10,3)"),
    ("consumo_medio", "NUMERIC(10,3)"),          # L/h ou L/100km
    ("alerta_consumo", "BOOLEAN DEFAULT false"),
    ("sem_funcionamento", "BOOLEAN DEFAULT false"),
]

_COLUNAS_NOTA = [
    ("numero_cupom", "TEXT"),
    ("fornecedor_id", "UUID"),
    ("fornecedor_nome_lido", "TEXT"),
    ("cnpj_lido", "TEXT"),
    ("data", "TIMESTAMPTZ DEFAULT now()"),
    ("combustivel", "TEXT"),
    ("rubrica", "TEXT"),
    ("litros_total", "NUMERIC(10,2)"),
    ("preco_litro", "NUMERIC(10,3)"),
    ("valor_total", "NUMERIC(12,2)"),
    ("foto_nota", "TEXT"),
    ("foto_leitura", "TEXT"),
    ("texto_livre", "TEXT"),
    ("extracao", "JSONB"),
    ("status_conferencia", "TEXT DEFAULT 'pendente'"),  # fase B: conferida | divergente
    ("usuario_id", "UUID"),
    ("usuario_nome", "TEXT"),
    ("criado_em", "TIMESTAMPTZ DEFAULT now()"),
    ("ativo", "BOOLEAN DEFAULT true"),
]

_SEED_COMBUSTIVEIS = [
    ("DIESEL_S10", "Diesel S10", "6.04.01"),
    ("DIESEL_S500", "Diesel S500", "6.04.01"),
    ("GASOLINA", "Gasolina", "6.04.02"),
    ("ETANOL", "Etanol", "6.04.03"),
]

async def _ddl():
    global _DDL_OK
    if _DDL_OK:
        return
    await ajard_query(
        """CREATE TABLE IF NOT EXISTS operacional.abastecimentos (
             id UUID PRIMARY KEY DEFAULT gen_random_uuid())""", fetch="none")
    for col, tipo in _COLUNAS_ITEM:
        await ajard_query(
            f"ALTER TABLE operacional.abastecimentos ADD COLUMN IF NOT EXISTS {col} {tipo}",
            fetch="none")
    await ajard_query(
        """CREATE INDEX IF NOT EXISTS idx_abast_equip
           ON operacional.abastecimentos (equipamento_id, data DESC)""", fetch="none")
    await ajard_query(
        """CREATE TABLE IF NOT EXISTS operacional.abastecimento_notas (
             id UUID PRIMARY KEY DEFAULT gen_random_uuid())""", fetch="none")
    for col, tipo in _COLUNAS_NOTA:
        await ajard_query(
            f"ALTER TABLE operacional.abastecimento_notas ADD COLUMN IF NOT EXISTS {col} {tipo}",
            fetch="none")
    await ajard_query(
        """CREATE INDEX IF NOT EXISTS idx_abast_nota_cupom
           ON operacional.abastecimento_notas (fornecedor_id, numero_cupom)""", fetch="none")
    # domínio de combustíveis (aparece na Parametrização via _DOMINIOS_PARAM)
    await ajard_query(
        """CREATE TABLE IF NOT EXISTS manutencao.combustiveis (
             codigo TEXT PRIMARY KEY, nome TEXT, ativo BOOLEAN DEFAULT true)""", fetch="none")
    await ajard_query(
        "ALTER TABLE manutencao.combustiveis ADD COLUMN IF NOT EXISTS rubrica TEXT", fetch="none")
    for cod, nome, rub in _SEED_COMBUSTIVEIS:
        await ajard_query(
            """INSERT INTO manutencao.combustiveis (codigo, nome, rubrica)
               VALUES (%s,%s,%s) ON CONFLICT (codigo) DO NOTHING""", (cod, nome, rub), fetch="none")
    # galão: capacidade na frota única
    await ajard_query(
        "ALTER TABLE operacional.equipamentos ADD COLUMN IF NOT EXISTS capacidade_l NUMERIC(10,1)",
        fetch="none")
    # parser numérico pt-BR usado pela v_leituras (KM do checklist vem em texto)
    await ajard_query(r"""
        CREATE OR REPLACE FUNCTION operacional.f_num_br(t text) RETURNS numeric
        LANGUAGE plpgsql IMMUTABLE AS $$
        DECLARE s text;
        BEGIN
          IF t IS NULL THEN RETURN NULL; END IF;
          s := regexp_replace(t, '[^0-9,\.]', '', 'g');
          IF s = '' THEN RETURN NULL; END IF;
          IF position(',' in s) > 0 THEN
            s := replace(replace(s, '.', ''), ',', '.');
          ELSIF s ~ '^[0-9]{1,3}(\.[0-9]{3})+$' THEN
            s := replace(s, '.', '');
          END IF;
          BEGIN
            RETURN s::numeric;
          EXCEPTION WHEN others THEN
            RETURN NULL;
          END;
        END $$""", fetch="none")
    await ajard_query("ALTER TABLE public.fornecedores ADD COLUMN IF NOT EXISTS eh_posto BOOLEAN DEFAULT false", fetch="none")
    _DDL_OK = True


# ── Helpers ─────────────────────────────────────────────────────────
def _uid(payload):
    """usuario_id só se for UUID válido — o sub do JWT pode ser o login (texto)."""
    for chave in ("usuario_id", "sub", "id"):
        v = payload.get(chave)
        if v:
            try:
                return str(uuid_lib.UUID(str(v)))
            except (ValueError, AttributeError, TypeError):
                continue
    return None


def _norm_placa(p):
    v = re.sub(r"[^A-Z0-9]", "", (p or "").upper())
    return v or None


def _norm_cnpj(c):
    v = re.sub(r"\D", "", c or "")
    return v if len(v) == 14 else None


def _chave_cod(s):
    """Chave tolerante de identificador: 'CB-0006' ≡ 'CB 06' ≡ 'cb6' → ('CB', 6).
    Letras + número sem zeros à esquerda. Identificador sem número → só letras."""
    t = re.sub(r"[^A-Z0-9]", "", (s or "").upper())
    if not t:
        return None
    m = re.match(r"^([A-Z]+)0*([0-9]+)$", t)
    if m:
        return (m.group(1), int(m.group(2)))
    return (t, None)


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    s = re.sub(r"[^0-9,\.\-]", "", s)
    if not s:
        return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif re.match(r"^\d{1,3}(\.\d{3})+$", s):
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def _f(v):
    return float(v) if v is not None else None


def _dt(v):
    """'31/08/2026 10:42' | ISO | None → datetime (fallback agora)."""
    if not v:
        return datetime.now()
    s = str(v).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            d = datetime.strptime(s[:19], fmt)
            if d.year < 2000 or d > datetime.now() + timedelta(days=1):
                return datetime.now()
            return d
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", ""))
    except ValueError:
        return datetime.now()


def _is_galao(eq):
    return (eq.get("categoria") or "").lower().startswith("gal")


async def _equip(eq_id):
    r = await ajard_query(
        """SELECT id, codigo, descricao, medicao, placa, categoria, centro_custo,
                  horimetro_atual, km_atual, capacidade_l
           FROM operacional.equipamentos WHERE id=%s""", (eq_id,))
    if not r:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    return dict(r[0])


async def _frota():
    r = await ajard_query(
        """SELECT id, codigo, descricao, medicao, placa, categoria, centro_custo, capacidade_l
           FROM operacional.equipamentos
           WHERE ativo=true AND COALESCE(categoria,'') NOT IN ('componente')
           ORDER BY codigo""")
    return [dict(x) for x in (r or [])]


async def _ultima_leitura(eq):
    """Nível 3 da cascata: maior leitura conhecida (fonte única v_leituras +
    partes diárias como rede)."""
    campo_parte = "km_final" if (eq.get("medicao") == "km") else "horimetro_final"
    r = await ajard_query(
        f"""SELECT GREATEST(
              COALESCE((SELECT MAX(leitura) FROM operacional.abastecimentos
                        WHERE equipamento_id=%s AND ativo=true
                          AND COALESCE(divergencia_leitura,false)=false), 0),
              COALESCE((SELECT MAX({campo_parte}) FROM operacional.partes_diarias
                        WHERE equipamento_id=%s), 0)) AS ultima""",
        (eq["id"], eq["id"]))
    v = r and r[0]["ultima"]
    return float(v) if v else None


def _cascata(digitada, foto, ultima, cadastro):
    if digitada is not None:
        return digitada, "digitada"
    if foto is not None:
        return foto, "foto"
    if ultima:
        return ultima, "ultima"
    if cadastro is not None:
        return float(cadastro), "equipamento"
    return None, None


async def _posto_resolver(nome, cnpj):
    """Devolve o id do posto: por CNPJ, senão por nome (exato, sem caixa),
    senão cria. Sempre marca tipo='posto' e completa o CNPJ se faltava."""
    nome = (nome or "").strip()
    cnpj = _norm_cnpj(cnpj) or (cnpj or None)
    if not nome and not cnpj:
        raise HTTPException(status_code=400, detail="Informe o nome do posto")
    r = None
    if cnpj:
        r = await ajard_query("SELECT id FROM public.fornecedores WHERE cnpj=%s AND ativo=true LIMIT 1", (cnpj,))
    if not r and nome:
        r = await ajard_query("SELECT id FROM public.fornecedores WHERE lower(nome)=lower(%s) AND ativo=true LIMIT 1", (nome,))
    if r:
        fid = str(r[0]["id"])
        await ajard_query("UPDATE public.fornecedores SET eh_posto=true, cnpj=COALESCE(cnpj,%s) WHERE id=%s", (cnpj, fid), fetch="none")
        return fid
    r = await ajard_query(
        """INSERT INTO public.fornecedores (nome, cnpj, tipo, eh_posto, observacao)
           VALUES (%s,%s,'servico',true,'criado pelo abastecimento') RETURNING id""", (nome or cnpj, cnpj))
    return str(r[0]["id"])


async def _garantir_lancada():
    await ajard_query("""ALTER TABLE operacional.abastecimento_notas
        ADD COLUMN IF NOT EXISTS lancada_mais BOOLEAN DEFAULT false,
        ADD COLUMN IF NOT EXISTS lancada_por UUID,
        ADD COLUMN IF NOT EXISTS lancada_em TIMESTAMPTZ""", fetch="none")


@router.get("/operacional/api/abastecimentos/notas")
async def notas_listar(mes: str = None, posto: str = None, pendentes: int = 0,
                       payload=Depends(verificar_notas)):
    """Notas por mês (competência da fatura do posto, que fecha dia 15).
    Cada nota: posto, cupom, combustível, litros, R$, equipamentos do rateio,
    fotos assinadas (1 h) e o carimbo 'lançada na MAIS'."""
    await _ddl(); await _garantir_lancada()
    from datetime import date as _date
    mes = (mes or _date.today().strftime("%Y-%m")).strip()[:7]
    where, params = ["n.ativo=true", "to_char(n.data,'YYYY-MM')=%s"], [mes]
    if posto:
        where.append("n.fornecedor_id=%s"); params.append(posto)
    if pendentes:
        where.append("COALESCE(n.lancada_mais,false)=false")
    rows = await ajard_query(
        f"""SELECT n.id, n.numero_cupom, n.data, n.combustivel, n.rubrica, n.litros_total,
                   n.preco_litro, n.valor_total, n.foto_nota, n.foto_leitura, n.usuario_nome,
                   n.lancada_mais, n.lancada_em, n.fornecedor_id, fo.nome AS posto, fo.cnpj,
                   (SELECT nome FROM public.usuarios_garra u WHERE u.id=n.lancada_por) AS lancada_por_nome,
                   (SELECT string_agg(e.codigo || ' (' || trim(to_char(a.litros,'FM999G999D99')) || ' L)', ' + ' ORDER BY e.codigo)
                      FROM operacional.abastecimentos a JOIN operacional.equipamentos e ON e.id=a.equipamento_id
                     WHERE a.nota_id=n.id AND a.ativo=true) AS equipamentos
            FROM operacional.abastecimento_notas n
            LEFT JOIN public.fornecedores fo ON fo.id=n.fornecedor_id
            WHERE {' AND '.join(where)}
            ORDER BY fo.nome NULLS LAST, n.data""", tuple(params)) or []
    notas, tot = [], {}
    for r in rows:
        d = dict(r)
        d["id"] = str(d["id"]); d["fornecedor_id"] = str(d["fornecedor_id"]) if d.get("fornecedor_id") else None
        for k in ("litros_total", "preco_litro", "valor_total"):
            d[k] = float(d[k]) if d.get(k) is not None else None
        d["data"] = d["data"].isoformat() if d.get("data") else None
        d["lancada_em"] = d["lancada_em"].isoformat() if d.get("lancada_em") else None
        d["foto_nota_url"] = storage_url(d.pop("foto_nota")) if d.get("foto_nota") else None
        d["foto_leitura_url"] = storage_url(d.pop("foto_leitura")) if d.get("foto_leitura") else None
        p = d.get("posto") or "(sem posto)"
        t = tot.setdefault(p, {"posto": p, "notas": 0, "litros": 0.0, "valor": 0.0, "pendentes": 0})
        t["notas"] += 1; t["litros"] += d["litros_total"] or 0; t["valor"] += d["valor_total"] or 0
        if not d.get("lancada_mais"): t["pendentes"] += 1
        notas.append(d)
    for t in tot.values():
        t["litros"] = round(t["litros"], 2); t["valor"] = round(t["valor"], 2)
    return {"mes": mes, "notas": notas, "totais": sorted(tot.values(), key=lambda x: x["posto"])}


@router.patch("/operacional/api/abastecimentos/notas/{nota_id}/lancada")
async def nota_lancada(nota_id: str, request: Request, payload=Depends(verificar_notas)):
    """Carimbo 'lançada na MAIS' — a conciliação viva entre os dois sistemas."""
    await _garantir_lancada()
    d = await request.json()
    lanc = bool(d.get("lancada"))
    uid = _uid(payload)
    await ajard_query(
        """UPDATE operacional.abastecimento_notas SET lancada_mais=%s,
             lancada_por=CASE WHEN %s THEN %s ELSE NULL END,
             lancada_em=CASE WHEN %s THEN now() ELSE NULL END
           WHERE id=%s""", (lanc, lanc, uid, lanc, nota_id), fetch="none")
    return {"ok": True, "lancada": lanc}


@router.get("/operacional/api/abastecimentos/notas/excel")
async def notas_excel(mes: str = None, posto: str = None, payload=Depends(verificar_notas)):
    """Excel da competência — uma linha por nota, link da foto válido por 7 dias
    (para anexar/conferir na MAIS sem abrir o sistema)."""
    from fastapi.responses import Response
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    dados = await notas_listar(mes=mes, posto=posto, pendentes=0, payload=payload)
    mes = dados["mes"]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"Notas {mes}"
    azul = PatternFill("solid", fgColor="1A2A5E"); branco = Font(color="FFFFFF", bold=True)
    cab = ["Data", "Posto", "CNPJ", "Cupom", "Combustível", "Rubrica", "Equipamentos (rateio)",
           "Litros", "R$/L", "Valor (R$)", "Quem registrou", "Lançada na MAIS", "Foto da nota (7 dias)"]
    ws.append(cab)
    for c in ws[1]:
        c.fill = azul; c.font = branco; c.alignment = Alignment(horizontal="center")
    # links de 7 dias: reassina a partir do caminho bruto
    brutos = await ajard_query(
        "SELECT id, foto_nota FROM operacional.abastecimento_notas WHERE to_char(data,'YYYY-MM')=%s AND ativo=true", (mes,)) or []
    caminho = {str(r["id"]): r["foto_nota"] for r in brutos}
    for n in dados["notas"]:
        raw = caminho.get(n["id"])
        foto = storage_url(raw, 7 * 24 * 3600) if raw else ""
        ws.append([(n["data"] or "")[:16].replace("T", " "), n.get("posto") or "", n.get("cnpj") or "",
                   n.get("numero_cupom") or "", n.get("combustivel") or "", n.get("rubrica") or "",
                   n.get("equipamentos") or "", n.get("litros_total"), n.get("preco_litro"), n.get("valor_total"),
                   n.get("usuario_nome") or "", "SIM" if n.get("lancada_mais") else "não", foto])
    ws.append([])
    ws.append(["TOTAIS DO MÊS POR POSTO"]); ws[ws.max_row][0].font = Font(bold=True)
    ws.append(["Posto", "Notas", "Litros", "Valor (R$)", "Pendentes de lançar"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for t in dados["totais"]:
        ws.append([t["posto"], t["notas"], t["litros"], t["valor"], t["pendentes"]])
    for col, w in zip("ABCDEFGHIJKLM", (15, 26, 16, 10, 12, 8, 44, 9, 8, 11, 16, 13, 60)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf)
    nome = f"notas-abastecimento-{mes}" + (f"-{posto[:8]}" if posto else "") + ".xlsx"
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{nome}"'})


@router.post("/operacional/api/abastecimentos/postos")
async def posto_criar(request: Request, payload=Depends(verificar_abastecimento)):
    """(01/09/2026) Cadastrar/marcar posto direto da tela de abastecimento."""
    d = await request.json()
    fid = await _posto_resolver(d.get("nome"), d.get("cnpj"))
    return {"ok": True, "id": fid, "postos": await _postos()}


@router.patch("/operacional/api/abastecimentos/postos/{fid}")
async def posto_editar(fid: str, request: Request, payload=Depends(verificar_abastecimento)):
    """Renomear / corrigir CNPJ do posto (cadastro único public.fornecedores)."""
    d = await request.json()
    nome = (d.get("nome") or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Informe o nome")
    dup = await ajard_query("SELECT id FROM public.fornecedores WHERE lower(nome)=lower(%s) AND id<>%s AND ativo=true LIMIT 1", (nome, fid))
    if dup:
        raise HTTPException(status_code=409, detail="Já existe fornecedor com este nome")
    cnpj = _norm_cnpj(d.get("cnpj")) or (d.get("cnpj") or None)
    await ajard_query("UPDATE public.fornecedores SET nome=%s, cnpj=COALESCE(%s, cnpj), eh_posto=true WHERE id=%s",
                      (nome, cnpj, fid), fetch="none")
    return {"ok": True, "postos": await _postos()}


async def _postos():
    """Só quem é posto: flag eh_posto (papel, não tipo — o tipo é do Compras),
    tipo legado 'posto' OU já tem nota. Ordem = uso mais recente primeiro."""
    await _ddl()
    r = await ajard_query(
        """SELECT f.id, f.nome, f.cnpj, MAX(n.data) AS ultimo_uso
             FROM public.fornecedores f
             LEFT JOIN operacional.abastecimento_notas n ON n.fornecedor_id=f.id AND n.ativo=true
            WHERE f.ativo=true AND (COALESCE(f.eh_posto,false) OR lower(COALESCE(f.tipo,'')) LIKE 'posto%%' OR n.id IS NOT NULL)
            GROUP BY f.id, f.nome, f.cnpj
            ORDER BY MAX(n.data) DESC NULLS LAST, f.nome""")
    return [{"id": x["id"], "nome": x["nome"], "cnpj": x["cnpj"]} for x in (r or [])]


_PALAVRAS_VAZIAS = {"POSTO", "AUTO", "LTDA", "ME", "EPP", "EIRELI", "COMBUSTIVEIS", "COMBUSTÍVEIS", "DE", "DO", "DA", "E", "SA", "S/A"}

def _tokens_nome(n):
    t = re.sub(r"[^A-Z0-9 ]", " ", (n or "").upper())
    return {w for w in t.split() if len(w) >= 2 and w not in _PALAVRAS_VAZIAS}


async def _combustiveis():
    r = await ajard_query(
        "SELECT codigo, nome, rubrica, ativo FROM manutencao.combustiveis ORDER BY nome")
    return [dict(x) for x in (r or [])]


def _match_combustivel(texto, lista):
    """'Diesel B S500', 'S10', 'óleo diesel', 'gasolina comum' → código do domínio."""
    t = (texto or "").upper()
    if not t:
        return None
    if "S10" in t or "S-10" in t:
        return next((c["codigo"] for c in lista if "S10" in c["codigo"]), None)
    if "S500" in t or "S-500" in t:
        return next((c["codigo"] for c in lista if "S500" in c["codigo"]), None)
    if "DIESEL" in t:
        dieseis = [c["codigo"] for c in lista if c["codigo"].startswith("DIESEL")
                   and c.get("ativo") is not False]
        return dieseis[0] if len(dieseis) == 1 else None  # ambíguo = humano escolhe
    if "ETANOL" in t or "ALCOOL" in t or "ÁLCOOL" in t:
        return next((c["codigo"] for c in lista if "ETANOL" in c["codigo"]), None)
    if "GASOL" in t:
        return next((c["codigo"] for c in lista if "GASOL" in c["codigo"]), None)
    return None


async def _saldo_galao(galao_id):
    r = await ajard_query(
        """SELECT COALESCE(SUM(CASE WHEN equipamento_id=%s AND destino_tipo='galao' THEN litros ELSE 0 END),0)
                - COALESCE(SUM(CASE WHEN galao_id=%s AND origem='galao' THEN litros ELSE 0 END),0) AS saldo
           FROM operacional.abastecimentos WHERE ativo=true""", (galao_id, galao_id))
    return float(r[0]["saldo"] or 0) if r else 0.0


async def _resumo_equip(eq, dias=None):
    """Rodapé ManWinWin. FMD-R = (última−primeira leitura)/dias entre elas.
    CMD-R = litros dos abastecimentos DEPOIS do primeiro ÷ (última−primeira)
    (×100 para km). Distância = 1/CMD (km/L ou h/L)."""
    dias = dias or JANELA_DIAS
    r = await ajard_query(
        """SELECT data, leitura, litros FROM operacional.abastecimentos
            WHERE equipamento_id=%s AND ativo=true AND leitura IS NOT NULL
              AND COALESCE(divergencia_leitura,false)=false
              AND destino_tipo='equipamento'
              AND data >= now() - (%s || ' days')::interval
            ORDER BY leitura ASC, data ASC""", (eq["id"], str(int(dias))))
    rows = [dict(x) for x in (r or [])]
    km = eq.get("medicao") == "km"
    out = {"janela_dias": dias, "n": len(rows), "medicao": "km" if km else "h",
           "unidade_fmd": "km/d" if km else "h/d",
           "unidade_cmd": "L/100km" if km else "L/h",
           "unidade_dist": "km/L" if km else "h/L",
           "ultimo_registo": None, "fmd_r": None, "cmd_r": None, "distancia": None}
    ult = await ajard_query(
        """SELECT data, leitura FROM operacional.abastecimentos
            WHERE equipamento_id=%s AND ativo=true AND leitura IS NOT NULL
              AND COALESCE(divergencia_leitura,false)=false
            ORDER BY data DESC LIMIT 1""", (eq["id"],))
    if ult:
        out["ultimo_registo"] = {"data": ult[0]["data"].isoformat(), "leitura": _f(ult[0]["leitura"])}
    if len(rows) >= 2:
        delta = float(rows[-1]["leitura"]) - float(rows[0]["leitura"])
        dias_real = max((rows[-1]["data"] - rows[0]["data"]).total_seconds() / 86400.0, 0.0)
        litros = sum(float(x["litros"] or 0) for x in rows[1:])
        if delta > 0:
            out["fmd_r"] = round(delta / dias_real, 2) if dias_real > 0 else None
            cmd = litros / delta * (100.0 if km else 1.0)
            out["cmd_r"] = round(cmd, 2)
            out["distancia"] = round((100.0 / cmd) if km else (1.0 / cmd), 2) if cmd > 0 else None
    return out


# ── Claude API (visão) ──────────────────────────────────────────────
_PROMPT_NOTA = (
    "Você lê documentos de abastecimento de uma empresa de terraplenagem no Brasil. "
    "Pode receber: (a) foto da NOTINHA/cupom do posto — impressa, às vezes com anotações "
    "à caneta (KM ou horímetro anotado pelo frentista, identificador do equipamento "
    "escrito pelo colaborador, ex.: 'EH-50', 'CB 06', 'PC-01'); (b) foto do painel/"
    "horímetro/odômetro; (c) texto livre do colaborador (ex.: 'EH-50 4520h, resto no "
    "galão'). Extraia o máximo possível. Um cupom pode abastecer MAIS DE UM equipamento "
    "(ex.: galão de 200 L dividido) — liste cada um em 'itens' com seus litros quando "
    "indicado; se só há um identificador, um item com litros null (o sistema assume o "
    "total). REGRAS CRÍTICAS: (1) 'leitura' é o número do ODÔMETRO/HORÍMETRO — na nota "
    "impressa costuma aparecer como 'KM: 441545' no rodapé, e à caneta como 'KM 441545' "
    "ou 'HR 4520'; tem tipicamente 3 a 7 dígitos. NUNCA use valores acompanhados de "
    "KM/L ou L/KM ou a palavra MEDIA — isso é consumo, não leitura. (2) 'combustivel' é "
    "a descrição LITERAL impressa do produto (ex.: 'OLEO DIESEL B S-500'), nunca resuma "
    "para 'Diesel'. (3) O rodapé impresso pode trazer PLACA e identificador do veículo "
    "(ex.: 'PLACA: GYC-9741 CB-05') — use-os como item se não houver outro. (4) 'itens' "
    "NUNCA fica vazio quando existe identificador em qualquer lugar (caneta, impresso ou "
    "texto): o que você escreveria em 'anotacoes' como identificador vai em 'itens'. "
    "(5) Identificador de equipamento é SEMPRE letras+números (CPO36, EH-50, CB 05) ou placa; "
    "nome do cliente (GARRA), do motorista ou do posto NUNCA é identificador. Números no "
    "padrão brasileiro (1.234,56). Responda SOMENTE JSON válido, sem "
    "markdown, exatamente com este formato: "
    '{"posto":{"nome":str|null,"cnpj":str|null},"cupom":str|null,'
    '"data_hora":"dd/mm/aaaa hh:mm"|null,"combustivel":str|null,'
    '"litros":num|null,"preco_litro":num|null,"valor_total":num|null,'
    '"itens":[{"identificador":str,"litros":num|null,"leitura":num|null,'
    '"tipo_leitura":"km"|"horimetro"|null}],'
    '"leitura_painel":{"valor":num|null,"tipo":"km"|"horimetro"|null,"placa":str|null},'
    '"anotacoes":str|null,"confianca":"alta"|"media"|"baixa"}')


def _json_tolerante(texto):
    """O modelo às vezes devolve JSON com vírgula sobrando ou chave sem aspas.
    Conserta o que dá antes de desistir — a leitura da nota não pode morrer
    por uma vírgula."""
    t = (texto or "").replace("```json", "").replace("```", "").strip()
    a, b = t.find("{"), t.rfind("}")
    if a >= 0 and b > a:
        t = t[a:b + 1]
    try:
        return json.loads(t)
    except json.JSONDecodeError:
        pass
    t2 = re.sub(r"//[^\n]*", "", t)                                # comentários // …
    t2 = re.sub(r"/\*.*?\*/", "", t2, flags=re.S)                  # comentários /* … */
    t2 = re.sub(r",\s*([}\]])", r"\1", t2)                        # vírgula antes de } ou ]
    t2 = re.sub(r"([{,]\s*)'([^']*)'\s*:", r'\1"\2":', t2)        # chave com aspas simples
    t2 = re.sub(r':\s*\'([^\'\n]*)\'', lambda m: ': "' + m.group(1).replace('"', '\\"') + '"', t2)  # valor com aspas simples
    t2 = re.sub(r'([{,]\s*)([A-Za-z_][A-Za-z0-9_]*)\s*:', r'\1"\2":', t2)  # chave sem aspas
    t2 = t2.replace("\u201c", '"').replace("\u201d", '"')       # aspas tipográficas
    t2 = re.sub(r"\bNone\b", "null", t2).replace("True", "true").replace("False", "false")
    return json.loads(t2)


def _claude_reparar_json(texto_quebrado):
    """Segunda chance: o modelo conserta o JSON que ele mesmo quebrou.
    Chamada só de texto — barata e rápida."""
    import requests as req_lib
    r = req_lib.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01",
                 "content-type": "application/json"},
        json={"model": MODELO_VISAO, "max_tokens": 900,
              "messages": [{"role": "user", "content":
                  "O texto abaixo deveria ser JSON válido mas está com erro de sintaxe. "
                  "Devolva SOMENTE o mesmo conteúdo como JSON estrito (aspas duplas em todas as "
                  "chaves e strings, sem vírgula sobrando, sem comentários, sem markdown):\n\n" + texto_quebrado}]},
        timeout=40)
    r.raise_for_status()
    t = "".join(b.get("text", "") for b in r.json().get("content", []) if b.get("type") == "text")
    return _json_tolerante(t)


def _claude_ler(imagens, texto_livre):
    import requests as req_lib
    if not ANTHROPIC_API_KEY:
        return {"_erro": "ANTHROPIC_API_KEY ausente no ambiente"}
    conteudo = []
    for rotulo, dados, mime in imagens:
        conteudo.append({"type": "text", "text": f"[{rotulo}]"})
        conteudo.append({"type": "image", "source": {"type": "base64", "media_type": mime,
                                                     "data": base64.b64encode(dados).decode()}})
    if texto_livre:
        conteudo.append({"type": "text", "text": "[texto do colaborador]\n" + texto_livre})
    conteudo.append({"type": "text", "text": _PROMPT_NOTA})
    try:
        r = req_lib.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": ANTHROPIC_API_KEY,
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": MODELO_VISAO, "max_tokens": 900,
                  "messages": [{"role": "user", "content": conteudo}]},
            timeout=60)
        r.raise_for_status()
        texto = "".join(b.get("text", "") for b in r.json().get("content", [])
                        if b.get("type") == "text")
        try:
            return _json_tolerante(texto)
        except json.JSONDecodeError:
            return _claude_reparar_json(texto)
    except Exception as e:  # extração nunca derruba o registro manual
        return {"_erro": str(e)[:200]}


async def _resolver(extr):
    """Casa a leitura bruta com o banco: equipamentos por identificador/placa,
    posto por CNPJ, combustível por texto. Devolve sugestão pronta pra tela."""
    frota = await _frota()
    combs = await _combustiveis()
    por_chave, por_placa = {}, {}
    for e in frota:
        k = _chave_cod(e["codigo"])
        if k:
            por_chave.setdefault(k, e)
        # descrição pode carregar apelido ('... - CB-06')
        for tok in re.findall(r"[A-Za-z]{1,4}[-\s]?0*\d{1,5}", e.get("descricao") or ""):
            kk = _chave_cod(tok)
            if kk and kk not in por_chave:
                por_chave[kk] = e
        p = _norm_placa(e.get("placa"))
        if p:
            por_placa[p] = e

    itens, nao_casados = [], []
    for it in (extr.get("itens") or []):
        ident = str((it or {}).get("identificador") or "").strip()
        if not re.search(r"\d", ident):
            continue  # sem número não é código nem placa (GARRA, JAIR, '.') — nem linha, nem aviso
        e = por_chave.get(_chave_cod(ident)) or por_placa.get(_norm_placa(ident) or "")
        row = {"identificador_lido": ident, "litros": _num((it or {}).get("litros")),
               "leitura": _num((it or {}).get("leitura")),
               "tipo_leitura": (it or {}).get("tipo_leitura")}
        if e:
            row.update({"equipamento_id": e["id"], "codigo": e["codigo"], "descricao": e.get("descricao"),
                        "medicao": e.get("medicao"), "galao": _is_galao(e)})
        else:
            nao_casados.append(ident)
        itens.append(row)

    # o modelo às vezes acha o equipamento e o escreve só na anotação
    # ("Identificador na placa: CB-05") — pesca tokens que CASEM com a frota
    # (o que não casar, como um rabisco "CQ.J3", é ignorado em silêncio)
    if not any(i.get("equipamento_id") for i in itens):
        fonte_txt = " ".join(str(extr.get(k) or "") for k in ("anotacoes",))
        vistos = set()
        for tok in re.findall(r"\b[A-Za-z]{1,4}[-. ]?0*\d{1,5}\b", fonte_txt):
            k = _chave_cod(tok)
            e = k and por_chave.get(k)
            if e and e["id"] not in vistos:
                vistos.add(e["id"])
                itens.append({"identificador_lido": tok, "litros": None, "leitura": None, "tipo_leitura": None,
                              "equipamento_id": e["id"], "codigo": e["codigo"], "descricao": e.get("descricao"),
                              "medicao": e.get("medicao"), "galao": _is_galao(e), "via_anotacao": True})

    painel = extr.get("leitura_painel") or {}
    placa = _norm_placa(painel.get("placa"))
    if placa and placa in por_placa and not any(i.get("equipamento_id") == por_placa[placa]["id"] for i in itens):
        e = por_placa[placa]
        itens.append({"identificador_lido": placa, "litros": None,
                      "leitura": _num(painel.get("valor")), "tipo_leitura": painel.get("tipo"),
                      "equipamento_id": e["id"], "codigo": e["codigo"], "descricao": e.get("descricao"),
                      "medicao": e.get("medicao"), "galao": False, "via_placa": True})
    # nota de UM equipamento (fora galões): completa o item com o que a nota
    # inteira já disse — litros do total; leitura da anotação ("KM 441545")
    # quando o modelo a deixou fora do item. Regra do único: sem ambiguidade.
    unicos = [i for i in itens if i.get("equipamento_id") and not i.get("galao")]
    if len(unicos) == 1:
        alvo = unicos[0]
        if alvo.get("litros") is None:
            alvo["litros"] = _num(extr.get("litros"))
        if alvo.get("leitura") is None:
            fonte_txt = " ".join(str(extr.get(k) or "") for k in ("anotacoes",))
            m = re.search(r"(?:KM|HR|HORIMETRO|HORÍMETRO)\D{0,3}([\d\.\,]{3,12})", fonte_txt, re.I)
            if m:
                alvo["leitura"] = _num(m.group(1))
                alvo["leitura_via"] = "anotacao"

    # trava de plausibilidade: caminhão (km) com "leitura" < 1000 é quase sempre
    # consumo (11,9 KM/L) lido por engano — não preenche, o humano decide
    for i in itens:
        if i.get("leitura") is not None and (i.get("medicao") == "km") and i["leitura"] < 1000:
            i["leitura_suspeita"] = i["leitura"]
            i["leitura"] = None

    # leitura do painel cai no item único sem leitura
    if _num(painel.get("valor")) is not None and len([i for i in itens if not i.get("galao")]) == 1:
        alvo = next(i for i in itens if not i.get("galao"))
        if alvo.get("leitura") is None:
            alvo["leitura"] = _num(painel.get("valor")); alvo["leitura_via"] = "painel"

    posto = extr.get("posto") or {}
    cnpj = _norm_cnpj(posto.get("cnpj"))
    fornecedor = None
    if cnpj:
        r = await ajard_query(
            """SELECT id, nome, cnpj FROM public.fornecedores
                WHERE regexp_replace(COALESCE(cnpj,''),'\\D','','g')=%s AND ativo=true LIMIT 1""", (cnpj,))
        if r:
            fornecedor = dict(r[0])
    # (31/08) Sem CNPJ casado, o nome é comparado SÓ entre postos (nunca contra
    # a oficina do Luciano): precisa de todos os tokens fortes do nome mais
    # curto (MR, IPIRANGA…) presentes no outro. Vem como sugestão laranja.
    if not fornecedor and (posto.get("nome") or "").strip():
        meus = _tokens_nome(posto["nome"])
        if meus:
            cands = []
            for pz in await _postos():
                seus = _tokens_nome(pz["nome"])
                menor, maior = (meus, seus) if len(meus) <= len(seus) else (seus, meus)
                if menor and menor <= maior:
                    cands.append(pz)
            if len(cands) == 1:
                fornecedor = cands[0]

    return {
        "posto": {"nome_lido": posto.get("nome"), "cnpj_lido": cnpj or posto.get("cnpj"),
                  "fornecedor_id": fornecedor and fornecedor["id"],
                  "fornecedor_nome": fornecedor and fornecedor["nome"]},
        "cupom": (extr.get("cupom") or None),
        "data_hora": extr.get("data_hora"),
        "combustivel": _match_combustivel(extr.get("combustivel"), combs),
        "combustivel_lido": extr.get("combustivel"),
        "litros": _num(extr.get("litros")),
        "preco_litro": _num(extr.get("preco_litro")),
        "valor_total": _num(extr.get("valor_total")),
        "itens": itens,
        "nao_casados": nao_casados,
        "anotacoes": extr.get("anotacoes"),
        "confianca": extr.get("confianca"),
        "erro": extr.get("_erro"),
    }


# ── PÁGINA MOBILE ───────────────────────────────────────────────────
@router.get("/abastecimento/notas", response_class=HTMLResponse)
async def pagina_notas():
    """(01/09/2026) Balcão financeiro das notinhas — feito para a Luana lançar
    contas a pagar na MAIS Soluções sem entrar na Manutenção: mês, posto,
    foto, valores, carimbo 'lançada' e Excel da competência."""
    return HTMLResponse("""<!doctype html><html lang="pt-BR"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Garra — Notas de Abastecimento</title>
<style>
:root{--azul:#1A2A5E;--lar:#F97316}
*{box-sizing:border-box}body{font-family:system-ui,Segoe UI,Roboto,sans-serif;margin:0;background:#F1F5F9;color:#0F172A}
header{background:var(--azul);color:#fff;padding:10px 16px;display:flex;justify-content:space-between;align-items:center}
header b{font-size:15px}header .l{color:var(--lar);font-weight:900}
main{max-width:1200px;margin:14px auto;padding:0 12px}
.filtros{display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;background:#fff;border:1px solid #E2E8F0;border-radius:10px;padding:10px}
.f{display:flex;flex-direction:column;font-size:11px;color:#475569;font-weight:700;gap:3px}
input,select,button{font:inherit;padding:6px 8px;border:1px solid #CBD5E1;border-radius:7px;background:#fff}
button{cursor:pointer}.p{background:var(--azul);color:#fff;border-color:var(--azul);font-weight:700}
.cards{display:flex;gap:10px;flex-wrap:wrap;margin:12px 0}
.card{background:#fff;border:1px solid #E2E8F0;border-radius:10px;padding:10px 14px;min-width:200px}
.card b{color:var(--azul)}.card .v{font-size:17px;font-weight:900;color:var(--azul)}
.card .pend{color:#B91C1C;font-weight:800;font-size:11px}
table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #E2E8F0;border-radius:10px;overflow:hidden;font-size:12px}
th{background:#F8FAFC;text-align:left;padding:7px 8px;font-size:10px;text-transform:uppercase;color:#475569;border-bottom:2px solid #E2E8F0}
td{padding:6px 8px;border-bottom:1px solid #F1F5F9;vertical-align:middle}
tr.ok td{background:#F0FDF4}
.r{text-align:right;font-variant-numeric:tabular-nums}
.chip{font-size:10px;padding:1px 6px;border-radius:8px;background:#FEF3C7;color:#92400E;font-weight:700}
a{color:var(--azul);font-weight:700;text-decoration:none}
.muted{color:#64748B}
#toast{position:fixed;bottom:16px;left:50%;transform:translateX(-50%);background:var(--azul);color:#fff;padding:10px 16px;border-radius:9px;display:none;font-weight:700;z-index:9}
</style></head><body>
<header><b><span class="l">GARRA</span> · Notas de Abastecimento</b><span id="quem" style="font-size:12px"></span></header>
<main>
 <div class="filtros">
  <div class="f">Competência<input type="month" id="f-mes"></div>
  <div class="f">Posto<select id="f-posto"><option value="">— todos —</option></select></div>
  <div class="f">&nbsp;<label style="font-weight:400"><input type="checkbox" id="f-pend"> só pendentes de lançar</label></div>
  <button class="p" onclick="carregar()">🔍 Buscar</button>
  <button onclick="excel()">📊 Excel da competência</button>
  <span class="muted" style="font-size:11px;flex:1;text-align:right">Fatura do posto fecha dia 15 · marque ✔ ao lançar na MAIS · foto abre em nova aba (link 1 h)</span>
 </div>
 <div class="cards" id="cards"></div>
 <div style="overflow:auto"><table><thead><tr>
  <th>Data</th><th>Posto</th><th>Cupom</th><th>Combustível</th><th>Equipamentos (rateio)</th>
  <th class="r">Litros</th><th class="r">R$/L</th><th class="r">Valor</th><th>Fotos</th><th>Quem</th><th style="width:130px">Lançada na MAIS</th>
 </tr></thead><tbody id="corpo"><tr><td colspan="11" class="muted" style="padding:14px">Carregando…</td></tr></tbody></table></div>
</main>
<div id="toast"></div>
<script>
const $=i=>document.getElementById(i);
(function(){const u=new URL(location.href);const sso=u.searchParams.get('sso')||u.searchParams.get('token');
 if(sso&&sso.length>10){sessionStorage.setItem('garra_notas_token',sso);history.replaceState(null,'',u.pathname);}})();
function tok(){return sessionStorage.getItem('garra_notas_token')||localStorage.getItem('garra_adm_token')||'';}
async function api(u,o){const r=await fetch(u,Object.assign({},o,{headers:Object.assign({'Content-Type':'application/json','Authorization':'Bearer '+tok()},(o&&o.headers)||{})}));
 if(r.status===401){document.body.innerHTML='<div style="padding:40px;text-align:center">🔒 Sessão expirada — abra o <a href="/admin">Admin</a> e clique em Notas de Abastecimento</div>';throw new Error('401');}
 if(r.status===403){document.body.innerHTML='<div style="padding:40px;text-align:center">Sem permissão para Notas de Abastecimento — peça ao administrador (Admin ▸ Permissões ▸ Notas de Abastecimento)</div>';throw new Error('403');}
 const j=await r.json().catch(()=>({})); if(!r.ok)throw new Error(j.detail||('Erro '+r.status)); return j;}
function toast(m){const t=$('toast');t.textContent=m;t.style.display='block';clearTimeout(t._h);t._h=setTimeout(()=>t.style.display='none',3500);}
const esc=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
const fN=(v,d)=>v==null?'—':Number(v).toLocaleString('pt-BR',{minimumFractionDigits:d??2,maximumFractionDigits:d??2});
let DADOS=null;
async function carregar(){
 const mes=$('f-mes').value||new Date().toISOString().slice(0,7);
 const q='mes='+mes+($('f-posto').value?'&posto='+$('f-posto').value:'')+($('f-pend').checked?'&pendentes=1':'');
 try{DADOS=await api('/operacional/api/abastecimentos/notas?'+q);}catch(e){return;}
 const sel=$('f-posto'),atual=sel.value; const postos={};
 DADOS.notas.forEach(n=>{if(n.fornecedor_id)postos[n.fornecedor_id]=n.posto;});
 if(!atual){sel.innerHTML='<option value="">— todos —</option>'+Object.entries(postos).map(([id,n])=>'<option value="'+id+'">'+esc(n)+'</option>').join('');}
 $('cards').innerHTML=(DADOS.totais||[]).map(t=>'<div class="card"><b>'+esc(t.posto)+'</b><div class="v">R$ '+fN(t.valor)+'</div><div class="muted" style="font-size:11px">'+t.notas+' nota(s) · '+fN(t.litros)+' L'+(t.pendentes?' · <span class="pend">'+t.pendentes+' pendente(s)</span>':' · ✔ tudo lançado')+'</div></div>').join('')||'<div class="muted">Nenhuma nota na competência.</div>';
 $('corpo').innerHTML=(DADOS.notas||[]).map(n=>'<tr class="'+(n.lancada_mais?'ok':'')+'">'
  +'<td style="white-space:nowrap">'+(n.data||'').slice(0,16).replace('T',' ')+'</td>'
  +'<td><b>'+esc(n.posto||'—')+'</b></td><td>'+esc(n.numero_cupom||'—')+'</td>'
  +'<td>'+esc(n.combustivel||'')+' <span class="muted" style="font-size:10px">'+esc(n.rubrica||'')+'</span></td>'
  +'<td style="font-size:11px">'+esc(n.equipamentos||'—')+'</td>'
  +'<td class="r">'+fN(n.litros_total)+'</td><td class="r">'+fN(n.preco_litro,3)+'</td><td class="r"><b>'+fN(n.valor_total)+'</b></td>'
  +'<td>'+(n.foto_nota_url?'<a href="'+esc(n.foto_nota_url)+'" target="_blank" title="Foto da nota">🧾</a> ':'<span class="chip">sem foto</span> ')
        +(n.foto_leitura_url?'<a href="'+esc(n.foto_leitura_url)+'" target="_blank" title="Foto do horímetro/KM">🕐</a>':'')+'</td>'
  +'<td style="font-size:11px">'+esc(n.usuario_nome||'')+'</td>'
  +'<td><label style="cursor:pointer;font-size:11px;font-weight:700;color:'+(n.lancada_mais?'#15803D':'#B91C1C')+'"><input type="checkbox" '+(n.lancada_mais?'checked':'')+' onchange="marcar(\''+n.id+'\',this)"> '
    +(n.lancada_mais?('✔ '+esc((n.lancada_por_nome||'').split(' ')[0])+' '+(n.lancada_em?new Date(n.lancada_em).toLocaleDateString('pt-BR'):'')):'lançar')+'</label></td></tr>').join('')
  ||'<tr><td colspan="11" class="muted" style="padding:14px">Nenhuma nota'+($('f-pend').checked?' pendente':'')+' na competência.</td></tr>';
}
async function marcar(id,ck){
 try{await api('/operacional/api/abastecimentos/notas/'+id+'/lancada',{method:'PATCH',body:JSON.stringify({lancada:ck.checked})});
  toast(ck.checked?'✔ Marcada como lançada na MAIS':'Desmarcada');carregar();}
 catch(e){ck.checked=!ck.checked;toast('❌ '+e.message);}
}
function excel(){
 const mes=$('f-mes').value||new Date().toISOString().slice(0,7);
 const q='mes='+mes+($('f-posto').value?'&posto='+$('f-posto').value:'');
 fetch('/operacional/api/abastecimentos/notas/excel?'+q,{headers:{'Authorization':'Bearer '+tok()}})
  .then(r=>{if(!r.ok)throw new Error('Erro '+r.status);return r.blob();})
  .then(b=>{const a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='notas-abastecimento-'+mes+'.xlsx';a.click();})
  .catch(e=>toast('❌ '+e.message));
}
(function(){ $('f-mes').value=new Date().toISOString().slice(0,7);
 try{const p=JSON.parse(atob(tok().split('.')[1].replace(/-/g,'+').replace(/_/g,'/')));$('quem').textContent=p.nome||p.login||'';}catch(e){}
 carregar(); })();
</script></body></html>""")


@router.get("/abastecimento", response_class=HTMLResponse)
async def pagina_abastecimento():
    return _PAGINA


# ── API ─────────────────────────────────────────────────────────────
@router.get("/operacional/api/abastecimentos/contexto")
async def contexto(payload=Depends(verificar_token)):
    """Tudo que a tela precisa num GET: frota (sem componentes), galões com
    saldo, combustíveis ativos, postos (fornecedores)."""
    await _ddl()
    frota = await _frota()
    galoes = []
    for e in frota:
        if _is_galao(e):
            galoes.append({"id": e["id"], "codigo": e["codigo"], "descricao": e.get("descricao"),
                           "capacidade_l": _f(e.get("capacidade_l")), "saldo": await _saldo_galao(e["id"])})
    combs = [c for c in await _combustiveis() if c.get("ativo") is not False]
    postos = await _postos()
    return {"frota": [{k: (_f(v) if k == "capacidade_l" else v) for k, v in e.items()} for e in frota],
            "galoes": galoes, "combustiveis": combs,
            "postos": [dict(p) for p in (postos or [])]}


@router.post("/operacional/api/abastecimentos/extrair")
async def extrair(foto_nota: UploadFile = File(None),
                  foto_leitura: UploadFile = File(None),
                  texto_livre: str = Form(None),
                  payload=Depends(verificar_abastecimento)):
    """Sobe fotos ao Storage e lê nota + painel + texto numa chamada só."""
    await _ddl()
    imagens, paths, erros_foto = [], {}, []
    pasta = "abastecimentos/" + datetime.now().strftime("%Y-%m")
    for chave, up in (("nota", foto_nota), ("leitura", foto_leitura)):
        if not up:
            continue
        dados = await up.read()
        if not dados:
            continue
        mime = up.content_type or "image/jpeg"
        path = f"{pasta}/{chave}-{os.urandom(6).hex()}.jpg"
        try:
            # (01/09/2026) guardar o caminho DEVOLVIDO — vem com o prefixo do bucket
            # ("garra-fotos:…"); sem ele, storage_url procurava no bucket legado e o
            # link da foto nunca nascia (bug que escondia o 🧾 no desktop)
            paths[chave] = storage_upload(dados, path, mime)
        except Exception as e:
            paths[chave] = None
            erros_foto.append(f"{chave}: {e}")
            print(f"[abastecimento] upload da foto {chave} falhou: {e!r}")
        imagens.append(("foto da nota" if chave == "nota" else "foto do painel/horímetro", dados, mime))
    if not imagens and not (texto_livre or "").strip():
        raise HTTPException(status_code=400, detail="Envie a foto da nota ou um texto")
    bruto = await asyncio.to_thread(_claude_ler, imagens, (texto_livre or "").strip())
    sug = await _resolver(bruto if isinstance(bruto, dict) else {})
    sug["paths"] = paths
    sug["bruto"] = bruto
    if erros_foto:
        sug["erro_foto"] = "; ".join(erros_foto)
    return sug


async def _validar_itens(itens_in, litros_total):
    if not itens_in:
        raise HTTPException(status_code=400, detail="Informe ao menos um equipamento ou galão")
    soma = 0.0
    itens = []
    for it in itens_in:
        eq_id = (it or {}).get("equipamento_id")
        if not eq_id:
            raise HTTPException(status_code=400, detail="Item sem equipamento")
        eq = await _equip(eq_id)
        litros = _num(it.get("litros"))
        if litros is None or litros <= 0:
            raise HTTPException(status_code=400, detail=f"Litros inválidos em {eq['codigo']}")
        soma += litros
        itens.append((eq, litros, it))
    if litros_total is not None and abs(soma - litros_total) > TOLERANCIA_RATEIO_L:
        raise HTTPException(
            status_code=400,
            detail=f"Rateio não fecha: itens somam {soma:.2f} L e a nota tem {litros_total:.2f} L")
    return itens, soma


async def _gravar_item(nota_id, eq, litros, it, comb, rubrica, preco_litro, data, payload,
                       origem="nota", galao_id=None, fotos=None):
    """Grava UM item com cascata, divergências, cadastro e inteligência.
    Retorna dict de resultado pronto pra tela."""
    fotos = fotos or {}
    galao = _is_galao(eq)
    if galao:
        r = await ajard_query(
            """INSERT INTO operacional.abastecimentos
                 (nota_id, equipamento_id, data, litros, valor_total, medicao, destino_tipo, origem,
                  galao_id, combustivel, rubrica, centro_custo, preco_litro, usuario_id, usuario_nome,
                  observacao, foto_painel, foto_bomba)
               VALUES (%s,%s,%s,%s,%s,%s,'galao',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (nota_id, eq["id"], data, litros, (round(litros * preco_litro, 2) if preco_litro else None),
             None, origem, galao_id, comb, rubrica, eq.get("centro_custo"), preco_litro,
             _uid(payload), payload.get("nome") or payload.get("login"),
             (it.get("observacao") or "").strip() or None, fotos.get("leitura"), fotos.get("nota")))
        return {"id": str(r[0]["id"]), "equipamento": eq["codigo"], "galao": True, "litros": litros,
                "saldo_galao": await _saldo_galao(eq["id"])}

    digitada = _num(it.get("leitura_digitada"))
    if digitada is None:
        digitada = _num(it.get("leitura"))  # leitura confirmada na tela = digitada
    foto = _num(it.get("leitura_foto"))
    ultima = await _ultima_leitura(eq)
    cadastro = eq.get("km_atual") if eq.get("medicao") == "km" else eq.get("horimetro_atual")
    leitura, fonte = _cascata(digitada, foto, ultima, cadastro)
    if it.get("leitura_via") == "painel" and fonte == "digitada":
        fonte = "foto"

    div_leitura = bool(leitura is not None and ultima and leitura < ultima)
    if digitada is not None and foto:
        if abs(digitada - foto) > max(0.05 * foto, 1):
            div_leitura = True
    placa_foto = _norm_placa(it.get("placa_foto"))
    placa_eq = _norm_placa(eq.get("placa"))
    div_placa = bool(placa_foto and placa_eq and placa_foto != placa_eq)

    # ── inteligência: consumo desde o abastecimento anterior ──
    consumo, alerta, sem_func, cmd_r = None, False, False, None
    prev = await ajard_query(
        """SELECT data, leitura FROM operacional.abastecimentos
            WHERE equipamento_id=%s AND ativo=true AND leitura IS NOT NULL
              AND COALESCE(divergencia_leitura,false)=false AND destino_tipo='equipamento'
              AND data < %s ORDER BY data DESC LIMIT 1""", (eq["id"], data))
    if prev and leitura is not None and not div_leitura and fonte in ("digitada", "foto"):
        delta = leitura - float(prev[0]["leitura"])
        if delta > 0:
            km = eq.get("medicao") == "km"
            consumo = round(litros / delta * (100.0 if km else 1.0), 3)
            res = await _resumo_equip(eq)
            cmd_r = res.get("cmd_r")
            if res.get("n", 0) >= MIN_ABAST_PARA_ALERTA and cmd_r and consumo > cmd_r * FATOR_ALERTA_CONSUMO:
                alerta = True
        try:
            from routers.manutencao import _garantir_leituras
            await _garantir_leituras()
            f = await ajard_query(
                """SELECT COUNT(*) AS n FROM operacional.v_leituras
                    WHERE equipamento_id=%s AND fonte IN ('parte','checklist')
                      AND data > %s AND data <= %s""", (eq["id"], prev[0]["data"], data))
            sem_func = bool(f and int(f[0]["n"]) == 0)
        except Exception:
            sem_func = False

    r = await ajard_query(
        """INSERT INTO operacional.abastecimentos
             (nota_id, equipamento_id, data, litros, valor_total, leitura, leitura_fonte, medicao,
              leitura_digitada, leitura_foto, placa_foto, divergencia_leitura, divergencia_placa,
              foto_painel, foto_bomba, usuario_id, usuario_nome, observacao,
              origem, galao_id, destino_tipo, combustivel, rubrica, centro_custo, preco_litro,
              consumo_medio, alerta_consumo, sem_funcionamento)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'equipamento',%s,%s,%s,%s,%s,%s,%s)
           RETURNING id""",
        (nota_id, eq["id"], data, litros, (round(litros * preco_litro, 2) if preco_litro else None),
         leitura, fonte, eq.get("medicao"), digitada, foto, placa_foto, div_leitura, div_placa,
         fotos.get("leitura"), fotos.get("nota"),
         _uid(payload), payload.get("nome") or payload.get("login"),
         (it.get("observacao") or "").strip() or None,
         origem, galao_id, comb, rubrica, eq.get("centro_custo"), preco_litro,
         consumo, alerta, sem_func))

    if leitura is not None and not div_leitura and fonte in ("digitada", "foto"):
        campo = "km_atual" if eq.get("medicao") == "km" else "horimetro_atual"
        await ajard_query(
            f"""UPDATE operacional.equipamentos SET {campo}=%s, atualizado_em=now()
                WHERE id=%s AND COALESCE({campo},0) < %s""", (leitura, eq["id"], leitura), fetch="none")

    return {"id": str(r[0]["id"]), "equipamento": eq["codigo"], "galao": False, "litros": litros,
            "leitura": leitura, "leitura_fonte": fonte, "medicao": eq.get("medicao"),
            "ultima_conhecida": ultima, "divergencia_leitura": div_leitura, "divergencia_placa": div_placa,
            "consumo_medio": consumo, "cmd_r": cmd_r, "alerta_consumo": alerta,
            "sem_funcionamento": sem_func}


@router.post("/operacional/api/abastecimentos/nota")
async def registrar_nota(request: Request, payload=Depends(verificar_abastecimento)):
    try:
        return await _registrar_nota(request, payload)
    except HTTPException:
        raise
    except Exception as e:  # nunca "Erro interno" mudo: causa na tela e traceback no log
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Falha ao registrar: {type(e).__name__}: {str(e)[:180]}")


async def _registrar_nota(request, payload):
    """Nota mãe + itens. Body:
    {nota:{numero_cupom, fornecedor_id|posto_nome+cnpj, data_hora, combustivel,
           litros_total, preco_litro, valor_total, foto_nota, foto_leitura, texto_livre, bruto},
     itens:[{equipamento_id, litros, leitura, leitura_foto, placa_foto, observacao}]}"""
    await _ddl()
    d = await request.json()
    n = d.get("nota") or {}
    itens_in = d.get("itens") or []

    combs = await _combustiveis()
    comb = n.get("combustivel") or None
    if not comb:
        raise HTTPException(status_code=400, detail="Informe o combustível")
    cd = next((c for c in combs if c["codigo"] == comb), None)
    if not cd or cd.get("ativo") is False:
        raise HTTPException(status_code=400, detail="Combustível não cadastrado")
    rubrica = cd.get("rubrica")

    litros_total = _num(n.get("litros_total"))
    valor_total = _num(n.get("valor_total"))
    preco_litro = _num(n.get("preco_litro"))
    if litros_total is None:
        litros_total = sum(_num(i.get("litros")) or 0 for i in itens_in) or None
    if litros_total is None or litros_total <= 0:
        raise HTTPException(status_code=400, detail="Litros da nota inválidos")
    if preco_litro is None and valor_total and litros_total:
        preco_litro = round(valor_total / litros_total, 3)
    if valor_total is None and preco_litro:
        valor_total = round(preco_litro * litros_total, 2)

    fornecedor_id = n.get("fornecedor_id") or None
    cnpj = _norm_cnpj(n.get("cnpj")) or (n.get("cnpj") or None)
    if not fornecedor_id and (n.get("posto_nome") or "").strip():
        # posto lido da nota → reaproveita por CNPJ/nome ou nasce no cadastro único
        fornecedor_id = await _posto_resolver(n["posto_nome"], cnpj)

    cupom = (str(n.get("numero_cupom") or "").strip() or None)
    if cupom and fornecedor_id:
        dup = await ajard_query(
            """SELECT id FROM operacional.abastecimento_notas
                WHERE ativo=true AND fornecedor_id=%s AND numero_cupom=%s LIMIT 1""",
            (fornecedor_id, cupom))
        if dup:
            raise HTTPException(status_code=409, detail=f"Cupom {cupom} deste posto já foi lançado")

    itens, soma = await _validar_itens(itens_in, litros_total)
    data = _dt(n.get("data_hora"))

    r = await ajard_query(
        """INSERT INTO operacional.abastecimento_notas
             (numero_cupom, fornecedor_id, fornecedor_nome_lido, cnpj_lido, data, combustivel, rubrica,
              litros_total, preco_litro, valor_total, foto_nota, foto_leitura, texto_livre, extracao,
              usuario_id, usuario_nome)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (cupom, fornecedor_id, n.get("posto_nome"), cnpj, data, comb, rubrica,
         litros_total, preco_litro, valor_total, n.get("foto_nota"), n.get("foto_leitura"),
         (n.get("texto_livre") or "").strip() or None,
         json.dumps(n.get("bruto")) if n.get("bruto") is not None else None,
         _uid(payload), payload.get("nome") or payload.get("login")))
    nota_id = str(r[0]["id"])

    fotos = {"nota": n.get("foto_nota"), "leitura": n.get("foto_leitura")}
    saida = []
    for eq, litros, it in itens:
        saida.append(await _gravar_item(nota_id, eq, litros, it, comb, rubrica, preco_litro, data,
                                        payload, fotos=fotos))
    return {"nota_id": nota_id, "cupom": cupom, "combustivel": comb, "rubrica": rubrica,
            "litros_total": litros_total, "valor_total": valor_total, "preco_litro": preco_litro,
            "itens": saida}


@router.post("/operacional/api/abastecimentos/do-galao")
async def registrar_do_galao(request: Request, payload=Depends(verificar_abastecimento)):
    """Saída do galão para um equipamento (sem nota). Debita o saldo do galão;
    combustível/rubrica herdados da última entrada do galão."""
    await _ddl()
    d = await request.json()
    galao = await _equip(d.get("galao_id") or "")
    if not _is_galao(galao):
        raise HTTPException(status_code=400, detail=f"{galao['codigo']} não é um galão")
    eq = await _equip(d.get("equipamento_id") or "")
    if _is_galao(eq):
        raise HTTPException(status_code=400, detail="Destino deve ser um equipamento")
    litros = _num(d.get("litros"))
    if not litros or litros <= 0:
        raise HTTPException(status_code=400, detail="Litros inválidos")
    saldo = await _saldo_galao(galao["id"])
    if litros > saldo + TOLERANCIA_RATEIO_L:
        raise HTTPException(status_code=400,
                            detail=f"Galão {galao['codigo']} tem {saldo:.1f} L — não cobre {litros:.1f} L")
    ult = await ajard_query(
        """SELECT combustivel, rubrica, preco_litro FROM operacional.abastecimentos
            WHERE equipamento_id=%s AND destino_tipo='galao' AND ativo=true
            ORDER BY data DESC LIMIT 1""", (galao["id"],))
    comb = (ult and ult[0]["combustivel"]) or d.get("combustivel")
    rubrica = (ult and ult[0]["rubrica"]) or None
    preco = _f(ult and ult[0]["preco_litro"])
    res = await _gravar_item(None, eq, litros, d, comb, rubrica, preco, _dt(d.get("data_hora")),
                             payload, origem="galao", galao_id=galao["id"])
    res["saldo_galao"] = await _saldo_galao(galao["id"])
    res["galao_codigo"] = galao["codigo"]
    return res


@router.get("/operacional/api/abastecimentos")
async def listar(equipamento_id: str = None, limite: int = 30, meus: int = 0, payload=Depends(verificar_token)):
    """Lista no formato da tela ManWinWin (Doc. Custo, produto, data, leitura,
    litros, custo, objeto, entidade, rubrica) + fonte/flags."""
    await _ddl()
    limite = max(1, min(int(limite or 30), 300))
    filtro, params = "", []
    if equipamento_id:
        filtro += " AND a.equipamento_id=%s"
        params.append(equipamento_id)
    if meus:
        uid = _uid(payload)
        if uid:
            filtro += " AND a.usuario_id=%s"; params.append(uid)
        else:
            filtro += " AND a.usuario_nome=%s"; params.append(payload.get("nome") or payload.get("login"))
    r = await ajard_query(
        f"""SELECT a.id, a.nota_id, a.data, a.litros, a.valor_total, a.leitura, a.leitura_fonte,
                   a.medicao, a.divergencia_leitura, a.divergencia_placa, a.usuario_nome,
                   a.observacao, a.origem, a.destino_tipo, a.combustivel, a.rubrica,
                   a.centro_custo, a.preco_litro, a.consumo_medio, a.alerta_consumo,
                   a.sem_funcionamento,
                   e.codigo AS equipamento, e.descricao AS equipamento_desc,
                   n.numero_cupom, n.foto_nota, n.foto_leitura, fo.nome AS posto, n.fornecedor_id,
                   (SELECT COUNT(*) FROM operacional.abastecimentos b WHERE b.nota_id=a.nota_id AND b.ativo=true) AS n_itens,
                   g.codigo AS galao_codigo, c.nome AS combustivel_nome
            FROM operacional.abastecimentos a
            LEFT JOIN operacional.equipamentos e ON e.id=a.equipamento_id
            LEFT JOIN operacional.equipamentos g ON g.id=a.galao_id
            LEFT JOIN operacional.abastecimento_notas n ON n.id=a.nota_id
            LEFT JOIN public.fornecedores fo ON fo.id=n.fornecedor_id
            LEFT JOIN manutencao.combustiveis c ON c.codigo=a.combustivel
            WHERE a.ativo=true {filtro}
            ORDER BY a.data DESC LIMIT {limite}""", tuple(params) or None)
    out = []
    for x in (r or []):
        x = dict(x)
        for k in ("litros", "valor_total", "leitura", "preco_litro", "consumo_medio"):
            x[k] = _f(x.get(k))
        x["data"] = x["data"].isoformat() if x.get("data") else None
        for k in ("id", "nota_id", "fornecedor_id"):
            x[k] = str(x[k]) if x.get(k) else None
        x["n_itens"] = int(x.get("n_itens") or 1)
        # links assinados (1 h) para conferência da foto no desktop
        x["foto_nota_url"] = storage_url(x["foto_nota"]) if x.get("foto_nota") else None
        x["foto_leitura_url"] = storage_url(x["foto_leitura"]) if x.get("foto_leitura") else None
        out.append(x)
    return out


@router.patch("/operacional/api/abastecimentos/{item_id}")
async def editar_abastecimento(item_id: str, request: Request, payload=Depends(verificar_abastecimento)):
    """(01/09/2026) Edição de um registro já efetivado (duplo clique no desktop).
    Item: data, litros, preco_litro, valor_total, leitura, combustivel, observacao.
    Nota-mãe (se a nota tem UM item): posto, cupom, data, combustível, litros,
    preço, valor. Com mais de um item, o cabeçalho da nota não muda por aqui
    (rateio é decisão humana no app). Regras mantidas: rubrica herdada do
    combustível, cupom único por posto, leitura nunca recua no cadastro,
    divergência recalculada contra as outras leituras. Trilha em observacao."""
    await _ddl()
    d = await request.json()
    it = await ajard_query(
        """SELECT a.*, (SELECT COUNT(*) FROM operacional.abastecimentos b WHERE b.nota_id=a.nota_id AND b.ativo=true) AS n_itens
           FROM operacional.abastecimentos a WHERE a.id=%s AND a.ativo=true""", (item_id,))
    if not it:
        raise HTTPException(status_code=404, detail="Abastecimento não encontrado")
    it = it[0]
    if it.get("destino_tipo") == "galao" or it.get("origem") == "galao":
        raise HTTPException(status_code=400, detail="Registro de galão — edite pelo app de campo")
    eq = await _equip(str(it["equipamento_id"]))
    combs = await _combustiveis()
    comb = (d.get("combustivel") or it.get("combustivel") or None)
    cd = next((c for c in combs if c["codigo"] == comb), None) if comb else None
    if comb and (not cd or cd.get("ativo") is False):
        raise HTTPException(status_code=400, detail="Combustível não cadastrado")
    rubrica = cd.get("rubrica") if cd else it.get("rubrica")
    litros = _num(d.get("litros")) if d.get("litros") not in (None, "") else float(it["litros"] or 0)
    if not litros or litros <= 0:
        raise HTTPException(status_code=400, detail="Litros deve ser positivo")
    preco = _num(d.get("preco_litro")) if d.get("preco_litro") not in (None, "") else (float(it["preco_litro"]) if it.get("preco_litro") is not None else None)
    valor = _num(d.get("valor_total")) if d.get("valor_total") not in (None, "") else None
    if valor is None and preco:
        valor = round(litros * preco, 2)
    if preco is None and valor:
        preco = round(valor / litros, 4)
    data = _dt(d.get("data_hora")) if d.get("data_hora") else it["data"]
    # leitura: digitada → recalcula divergência contra as OUTRAS leituras
    leitura = it.get("leitura"); fonte = it.get("leitura_fonte"); div = bool(it.get("divergencia_leitura"))
    if "leitura" in d:
        nova = _num(d.get("leitura"))
        if nova is not None:
            outras = await ajard_query(
                """SELECT COALESCE(MAX(leitura),0) AS m FROM operacional.abastecimentos
                   WHERE equipamento_id=%s AND ativo=true AND id<>%s AND data < %s
                     AND COALESCE(divergencia_leitura,false)=false AND destino_tipo='equipamento'""",
                (eq["id"], item_id, data))
            ant = float(outras[0]["m"] or 0) if outras else 0.0
            leitura, fonte, div = nova, "digitada", bool(ant and nova < ant)
    quem = payload.get("nome") or payload.get("login") or ""
    trilha = f"editado por {quem} em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    obs = (d.get("observacao") if "observacao" in d else it.get("observacao")) or ""
    obs = (obs + (" · " if obs else "") + trilha)[:500]
    await ajard_query(
        """UPDATE operacional.abastecimentos SET data=%s, litros=%s, preco_litro=%s, valor_total=%s,
             leitura=%s, leitura_fonte=%s, divergencia_leitura=%s, combustivel=%s, rubrica=%s, observacao=%s
           WHERE id=%s""",
        (data, litros, preco, valor, leitura, fonte, div, comb, rubrica, obs, item_id), fetch="none")
    if leitura is not None and not div:
        campo = "km_atual" if eq.get("medicao") == "km" else "horimetro_atual"
        await ajard_query(
            f"""UPDATE operacional.equipamentos SET {campo}=%s, atualizado_em=now()
                WHERE id=%s AND COALESCE({campo},0) < %s""", (leitura, eq["id"], leitura), fetch="none")
    nota_alterada = False
    if it.get("nota_id") and int(it["n_itens"] or 1) == 1:
        forn = (d.get("fornecedor_id") or "").strip() or None
        if not forn and (d.get("posto_nome") or "").strip():
            forn = await _posto_resolver(d.get("posto_nome"), d.get("cnpj"))
        cupom = (str(d.get("numero_cupom") or "").strip() or None) if "numero_cupom" in d else None
        n = await ajard_query("SELECT fornecedor_id, numero_cupom FROM operacional.abastecimento_notas WHERE id=%s", (it["nota_id"],))
        n = n[0] if n else {}
        forn_final = forn or n.get("fornecedor_id")
        cupom_final = cupom if "numero_cupom" in d else n.get("numero_cupom")
        if cupom_final and forn_final:
            dup = await ajard_query(
                """SELECT id FROM operacional.abastecimento_notas
                   WHERE ativo=true AND fornecedor_id=%s AND numero_cupom=%s AND id<>%s LIMIT 1""",
                (forn_final, cupom_final, it["nota_id"]))
            if dup:
                raise HTTPException(status_code=409, detail=f"Cupom {cupom_final} deste posto já existe em outra nota")
        await ajard_query(
            """UPDATE operacional.abastecimento_notas SET fornecedor_id=%s, numero_cupom=%s, data=%s, combustivel=%s,
                 rubrica=%s, litros_total=%s, preco_litro=%s, valor_total=%s WHERE id=%s""",
            (forn_final, cupom_final, data, comb, rubrica, litros, preco, valor, it["nota_id"]), fetch="none")
        nota_alterada = True
    return {"ok": True, "id": item_id, "leitura": leitura, "divergencia_leitura": div, "nota_alterada": nota_alterada,
            "valor_total": valor, "preco_litro": preco}


@router.post("/operacional/api/abastecimentos/{item_id}/foto")
async def anexar_foto(item_id: str, foto: UploadFile = File(...), tipo: str = Form("nota"),
                      payload=Depends(verificar_abastecimento)):
    """(01/09/2026) Operador esqueceu a foto? Anexa depois, sem recriar a nota.
    tipo = nota | leitura. Vai para a nota-mãe (vale para todos os itens dela)."""
    await _ddl()
    if tipo not in ("nota", "leitura"):
        raise HTTPException(status_code=400, detail="tipo deve ser nota ou leitura")
    it = await ajard_query("SELECT id, nota_id, usuario_id FROM operacional.abastecimentos WHERE id=%s AND ativo=true", (item_id,), fetch="one")
    if not it:
        raise HTTPException(status_code=404, detail="Abastecimento não encontrado")
    if not it.get("nota_id"):
        raise HTTPException(status_code=400, detail="Registro sem nota-mãe (galão) — foto não se aplica")
    dados = await foto.read()
    if not dados:
        raise HTTPException(status_code=400, detail="Arquivo vazio")
    pasta = "abastecimentos/" + datetime.now().strftime("%Y-%m")
    caminho = storage_upload(dados, f"{pasta}/{tipo}-{os.urandom(6).hex()}.jpg", foto.content_type or "image/jpeg")
    col = "foto_nota" if tipo == "nota" else "foto_leitura"
    await ajard_query(f"UPDATE operacional.abastecimento_notas SET {col}=%s WHERE id=%s", (caminho, it["nota_id"]), fetch="none")
    return {"ok": True, "url": storage_url(caminho)}


@router.get("/operacional/api/abastecimentos/resumo/{eq_id}")
async def resumo(eq_id: str, dias: int = None, payload=Depends(verificar_token)):
    await _ddl()
    eq = await _equip(eq_id)
    res = await _resumo_equip(eq, dias)
    res["equipamento"] = eq["codigo"]
    return res


@router.get("/operacional/api/abastecimentos/galoes")
async def galoes(payload=Depends(verificar_token)):
    await _ddl()
    out = []
    for e in await _frota():
        if _is_galao(e):
            out.append({"id": e["id"], "codigo": e["codigo"], "descricao": e.get("descricao"),
                        "capacidade_l": _f(e.get("capacidade_l")), "saldo": await _saldo_galao(e["id"])})
    return out


@router.get("/operacional/api/abastecimentos/ultima/{eq_id}")
async def ultima_leitura(eq_id: str, payload=Depends(verificar_token)):
    await _ddl()
    eq = await _equip(eq_id)
    ultima = await _ultima_leitura(eq)
    cadastro = eq.get("km_atual") if eq.get("medicao") == "km" else eq.get("horimetro_atual")
    return {"equipamento": eq["codigo"], "medicao": eq.get("medicao"),
            "ultima_conhecida": ultima, "cadastro": _f(cadastro), "placa": eq.get("placa"),
            "galao": _is_galao(eq)}


@router.delete("/operacional/api/abastecimentos/{ab_id}")
async def excluir(ab_id: str, payload=Depends(verificar_token)):
    """Soft delete do item. Se era o último item vivo da nota, a nota também
    sai de cena — cupom volta a poder ser lançado."""
    await _ddl()
    if (payload.get("perfil") or "").lower() not in ("admin", "gestor"):
        raise HTTPException(status_code=403, detail="Só gestão exclui abastecimento")
    r = await ajard_query(
        "UPDATE operacional.abastecimentos SET ativo=false WHERE id=%s RETURNING nota_id", (ab_id,))
    nota_id = r and r[0]["nota_id"]
    if nota_id:
        vivos = await ajard_query(
            "SELECT COUNT(*) AS n FROM operacional.abastecimentos WHERE nota_id=%s AND ativo=true", (nota_id,))
        if vivos and int(vivos[0]["n"]) == 0:
            await ajard_query(
                "UPDATE operacional.abastecimento_notas SET ativo=false WHERE id=%s", (nota_id,), fetch="none")
    return {"ok": True}


# ── PÁGINA (inline — autocontida, padrão pedido-mobile) ──────────────
_PAGINA = r"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>Garra — Abastecimento</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:#F1F5F9;color:#0F172A;font-size:15px;padding:12px 12px 96px;max-width:560px;margin:0 auto}
h1{font-size:16px;color:#1A2A5E;margin:6px 0 12px}h1 span{color:#E8820C}
.card{background:#fff;border:1px solid #E2E8F0;border-radius:12px;padding:14px;margin-bottom:10px}
label{display:block;font-size:12px;color:#475569;font-weight:600;margin:10px 0 4px}
input,select,textarea{width:100%;padding:12px;border:1px solid #CBD5E1;border-radius:9px;font-size:16px;background:#fff}
input.lido,select.lido{border-color:#E8820C;background:#FFF7ED}
input.ruim{border-color:#DC2626;background:#FEF2F2}
.btn{width:100%;padding:14px;border:none;border-radius:10px;font-size:16px;font-weight:700;cursor:pointer;margin-top:10px}
.btn-p{background:#1A2A5E;color:#fff}.btn-o{background:#E8820C;color:#fff}.btn-c{background:#E2E8F0;color:#0F172A}
.btn:disabled{opacity:.5}
.row{display:flex;gap:8px}.row>*{flex:1}
.muted{font-size:12px;color:#64748B}
.flag{font-size:13px;padding:9px 11px;border-radius:9px;margin-top:8px;line-height:1.35}
.flag-warn{background:#FEF3C7;color:#92400E}.flag-ok{background:#DCFCE7;color:#166534}.flag-err{background:#FEE2E2;color:#991B1B}.flag-info{background:#DBEAFE;color:#1E3A8A}
.hist{font-size:13px;border-top:1px solid #E2E8F0;padding:8px 0}.hist b{color:#1A2A5E}
#toast{position:fixed;bottom:100px;left:50%;transform:translateX(-50%);background:#0F172A;color:#fff;padding:10px 16px;border-radius:9px;font-size:14px;display:none;z-index:99;max-width:92%}
.hide{display:none}
.foto{display:flex;gap:10px;align-items:center;border:2px dashed #CBD5E1;border-radius:12px;padding:10px;background:#F8FAFC;cursor:pointer;min-height:76px}
.foto.ok{border-style:solid;border-color:#16A34A;background:#F0FDF4}
.foto img{width:64px;height:64px;object-fit:cover;border-radius:8px;background:#E2E8F0}
.foto .ic{width:64px;height:64px;border-radius:8px;background:#E2E8F0;display:flex;align-items:center;justify-content:center;font-size:28px}
.foto .tx{flex:1;font-size:14px}.foto .tx b{display:block;color:#1A2A5E}
.foto{cursor:default}
.btn-ditar{display:inline-block;margin-top:6px;padding:9px 12px;border:1px solid #CBD5E1;border-radius:8px;background:#F8FAFC;color:#1A2A5E;font-size:13px;font-weight:700;cursor:pointer}.btn-ditar.gravando{background:#FEE2E2;border-color:#DC2626;color:#991B1B}
.fbtn{margin-top:6px}
.fbtn .btn{margin-top:0;padding:11px;font-size:14px}
.prog{height:6px;background:#E2E8F0;border-radius:3px;overflow:hidden;margin-top:8px}.prog i{display:block;height:100%;width:0;background:#E8820C;transition:width .4s}
.item{border:1px solid #E2E8F0;border-radius:10px;padding:10px;margin-top:8px;background:#FAFAFA}
.item .top{display:flex;gap:8px;align-items:center}.item .top select{flex:1}
.item .x{width:40px;height:40px;border:none;border-radius:8px;background:#FEE2E2;color:#991B1B;font-size:18px;font-weight:700}
.item .sub{font-size:12px;color:#64748B;margin-top:4px}
.soma{display:flex;justify-content:space-between;font-size:13px;margin-top:8px;padding:8px 10px;border-radius:8px;background:#F1F5F9}
.soma.ok{background:#DCFCE7;color:#166534}.soma.ruim{background:#FEE2E2;color:#991B1B}
#rodape{position:fixed;left:0;right:0;bottom:0;background:#fff;border-top:1px solid #E2E8F0;padding:10px 12px;z-index:50}
#rodape .btn{margin:0}
.chip{display:inline-block;font-size:11px;padding:2px 8px;border-radius:10px;background:#E2E8F0;color:#334155;margin-right:4px}
.chip.g{background:#DCFCE7;color:#166534}.chip.w{background:#FEF3C7;color:#92400E}.chip.r{background:#FEE2E2;color:#991B1B}
</style></head><body>
<script>(function(){ try {
  const q = new URLSearchParams(location.search);
  const sso = q.get('sso');
  if (sso && sso.length > 10) { localStorage.setItem('garra_tok_abast', sso); }
  if (q.get('embedded') === '1') document.documentElement.classList.add('embedded');
  if (sso && sso.length > 10) document.documentElement.classList.add('sso');
  if (sso || q.get('embedded')) history.replaceState(null, '', location.pathname);
} catch(e) {} })();</script>
<style>.embedded h1{display:none !important}.embedded body{padding-top:6px}.sso #tela-login{display:none !important}</style>
<h1>⛽ Abastecimento <span>Garra</span></h1>

<div class="card" id="tela-login">
  <b>Entrar</b><div class="muted">Use o mesmo login dos aplicativos Garra.</div>
  <label>Login</label><input id="lg-user" autocomplete="username">
  <label>Senha</label><input id="lg-senha" type="password" autocomplete="current-password">
  <button class="btn btn-p" onclick="entrar()">Entrar</button>
</div>

<div class="hide" id="tela-form">
  <div class="card">
    <div class="foto" id="box-nota">
      <span class="ic" id="ic-nota">🧾</span>
      <span class="tx"><b>Foto da notinha do posto</b><span class="muted" id="tx-nota">Obrigatória</span></span>
    </div>
    <div class="row fbtn">
      <button type="button" class="btn btn-o" onclick="$('f-nota').click()">📷 Tirar foto</button>
      <button type="button" class="btn btn-c" onclick="$('g-f-nota').click()">🖼 Galeria</button>
    </div>
    <input type="file" id="f-nota" accept="image/*" capture="environment" style="display:none" onchange="fotoEscolhida('nota',this)">
    <input type="file" id="g-f-nota" accept="image/*" style="display:none" onchange="fotoEscolhida('nota',this)">
    <div class="foto" id="box-leitura" style="margin-top:12px">
      <span class="ic" id="ic-leitura">🕐</span>
      <span class="tx"><b>Foto do horímetro / KM</b><span class="muted" id="tx-leitura">Opcional — se não está anotado na nota</span></span>
    </div>
    <div class="row fbtn">
      <button type="button" class="btn btn-o" onclick="$('f-leitura').click()">📷 Tirar foto</button>
      <button type="button" class="btn btn-c" onclick="$('g-f-leitura').click()">🖼 Galeria</button>
    </div>
    <input type="file" id="f-leitura" accept="image/*" capture="environment" style="display:none" onchange="fotoEscolhida('leitura',this)">
    <input type="file" id="g-f-leitura" accept="image/*" style="display:none" onchange="fotoEscolhida('leitura',this)">
    <label>Escreva como mandaria no grupo <span class="muted">(opcional)</span></label>
    <input id="f-texto" placeholder="ex.: CB-05 441545 km · EH-50 120 L" onblur="talvezLer()">
    <button type="button" class="btn-ditar" data-ditar="f-texto">🎤 Ditar</button>
    <div class="prog hide" id="prog"><i id="prog-i"></i></div>
    <div class="muted" id="prog-tx" style="margin-top:6px"></div>
    <div id="leitura-flags"></div>
  </div>

  <div class="muted" style="text-align:center;margin:2px 0 8px" id="lk-manual"><a href="#" onclick="mostrarDados();return false">✍️ Preencher sem foto →</a></div>

  <div class="hide" id="sec-dados">
  <div class="card" id="card-nota">
    <b style="color:#1A2A5E">Nota</b> <span class="muted" id="confianca"></span>
    <div class="row">
      <div><label>Posto</label><select id="n-posto" onchange="postoMudou()"></select></div>
    </div>
    <div class="row hide" id="posto-novo">
      <div><label>Nome do posto (novo)</label><input id="n-posto-nome" placeholder="como está na nota"></div>
      <div style="flex:.8"><label>CNPJ</label><input id="n-cnpj" inputmode="numeric" placeholder="só números"></div>
    </div>
    <div class="row">
      <div><label>Nº do cupom</label><input id="n-cupom" placeholder="Doc."></div>
      <div><label>Data / hora</label><input id="n-data" placeholder="dd/mm/aaaa hh:mm"></div>
    </div>
    <div class="row">
      <div><label>Combustível *</label><select id="n-comb"></select></div>
    </div>
    <div class="row">
      <div><label>Litros da nota *</label><input id="n-litros" inputmode="decimal" placeholder="0,00" oninput="herdarLitrosUnico();recalcSoma()"></div>
      <div><label>Preço / L</label><input id="n-preco" inputmode="decimal" placeholder="0,000" oninput="recalcValor('preco')"></div>
      <div><label>Valor total</label><input id="n-valor" inputmode="decimal" placeholder="0,00" oninput="recalcValor('valor')"></div>
    </div>
  </div>

  <div class="card">
    <b style="color:#1A2A5E">Para onde foi o combustível</b>
    <div class="muted">Um item por máquina, caminhão ou galão. Os litros dos itens têm que fechar com a nota.</div>
    <div id="itens"></div>
    <button class="btn btn-c" onclick="addItem()">➕ Adicionar equipamento ou galão</button>
    <div class="soma" id="soma"><span>Soma dos itens</span><b id="soma-v">0,00 L</b></div>
  </div>

  </div><!-- /sec-dados -->

  <div id="f-resultado"></div>

  <div class="card" id="tela-hist"><b>Meus últimos registros</b><div class="muted">Confirmação do que você já enviou — evita mandar a mesma nota duas vezes.</div><div id="hist"></div></div>
</div>

<div id="rodape" class="hide"><button class="btn btn-p" id="btn-salvar" onclick="salvar()">💾 Registrar abastecimento</button></div>
<div id="toast"></div>
<script>
const API = location.origin;
let TOK = localStorage.getItem('garra_tok_abast') || '';
let CTX = {frota:[], galoes:[], combustiveis:[], postos:[]};
let FOTOS = {nota:null, leitura:null};   // File comprimido
let PATHS = {};                          // caminhos no storage
let BRUTO = null;
let LENDO = false, LIDO_HASH = '';
let ULTIMAS = {};                        // eq_id -> {ultima_conhecida, medicao}

function mostrarDados(){ mostrar('sec-dados',true); mostrar('rodape',true); const l=$('lk-manual'); if(l) l.style.display='none'; }
function toast(m){ const t=document.getElementById('toast'); t.textContent=m; t.style.display='block'; setTimeout(()=>t.style.display='none',3400); }
function mostrar(id,on){ document.getElementById(id).classList[on?'remove':'add']('hide'); }
function $(id){ return document.getElementById(id); }
function num(v){ if(v==null||v==='') return null; v=String(v).replace(/[^0-9,\.\-]/g,''); if(v.indexOf(',')>=0) v=v.replace(/\./g,'').replace(',','.'); else if(/^\d{1,3}(\.\d{3})+$/.test(v)) v=v.replace(/\./g,''); const n=parseFloat(v); return isNaN(n)?null:n; }
function fmt(n,d){ return n==null?'—':n.toLocaleString('pt-BR',{minimumFractionDigits:d==null?2:d,maximumFractionDigits:d==null?2:d}); }
function esc(s){ return String(s==null?'':s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c])); }

async function api(path, opts){
  opts = opts || {};
  opts.headers = Object.assign({'Authorization':'Bearer '+TOK}, opts.headers||{});
  const r = await fetch(API+path, opts);
  if (r.status===401){ TOK=''; localStorage.removeItem('garra_tok_abast'); mostrar('tela-login',true); mostrar('tela-form',false); mostrar('rodape',false); throw new Error('Sessão expirada — entre de novo'); }
  if (!r.ok){ let d; try{ d=await r.json(); }catch(e){} throw new Error((d&&d.detail)||('Erro '+r.status)); }
  return r.json();
}

async function entrar(){
  const login=$('lg-user').value.trim(), senha=$('lg-senha').value;
  if(!login||!senha){ toast('Preencha login e senha'); return; }
  try{
    const r=await fetch(API+'/auth/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({login:login,senha:senha})});
    if(!r.ok) throw new Error('Login inválido');
    const d=await r.json(); TOK=d.token||d.access_token;
    localStorage.setItem('garra_tok_abast',TOK);
    iniciar();
  }catch(e){ toast('❌ '+e.message); }
}

async function iniciar(){
  mostrar('tela-login',false); mostrar('tela-form',true); mostrar('rodape',false);
  try{
    CTX = await api('/operacional/api/abastecimentos/contexto');
    $('n-comb').innerHTML = '<option value="">— escolha —</option>'+CTX.combustiveis.map(c=>'<option value="'+c.codigo+'">'+esc(c.nome)+'</option>').join('');
    $('n-posto').innerHTML = '<option value="">— escolha o posto —</option>'+CTX.postos.map(p=>'<option value="'+p.id+'">'+esc(p.nome)+'</option>').join('')+'<option value="__novo">＋ Posto novo (não está na lista)</option>';
    if(!$('itens').children.length){ addItem(localStorage.getItem('garra_abast_ultimo_eq')||''); }
    carregarHist();
  }catch(e){ toast('❌ '+e.message); }
}

function ehGalao(e){ return ((e.categoria||'').toLowerCase().indexOf('gal')===0) || e.galao===true; }
function opcoesEquip(sel){
  const m=CTX.frota.filter(e=>!ehGalao(e));
  let h='<option value="">— escolha —</option>';
  h+=m.map(e=>'<option value="'+e.id+'"'+(e.id===sel?' selected':'')+'>'+esc(e.codigo)+' — '+esc(e.descricao||'')+'</option>').join('');
  return h;
}

// ── fotos: comprime no aparelho antes de subir ──
function comprimir(file, max){
  return new Promise(res=>{
    const img=new Image(); const url=URL.createObjectURL(file);
    img.onload=()=>{
      let w=img.width,h=img.height; const k=Math.min(1,(max||1600)/Math.max(w,h)); w=Math.round(w*k); h=Math.round(h*k);
      const c=document.createElement('canvas'); c.width=w; c.height=h; c.getContext('2d').drawImage(img,0,0,w,h);
      c.toBlob(b=>{ URL.revokeObjectURL(url); res(b?new File([b],file.name.replace(/\.\w+$/,'')+'.jpg',{type:'image/jpeg'}):file); },'image/jpeg',0.82);
    };
    img.onerror=()=>{ URL.revokeObjectURL(url); res(file); };
    img.src=url;
  });
}
async function fotoEscolhida(qual,origem){
  const inp=origem||$('f-'+qual); const f=inp.files&&inp.files[0]; if(!f) return;
  $('tx-'+qual).textContent='Comprimindo…';
  const c=await comprimir(f,1600); FOTOS[qual]=c;
  const box=$('box-'+qual); box.classList.add('ok');
  const ic=$('ic-'+qual); const img=document.createElement('img'); img.src=URL.createObjectURL(c); img.id='ic-'+qual; ic.replaceWith(img);
  $('tx-'+qual).textContent=Math.round(c.size/1024)+' KB · toque para trocar';
  talvezLer();
}
function hashEntrada(){ return (FOTOS.nota?FOTOS.nota.size+'-'+FOTOS.nota.lastModified:'')+'|'+(FOTOS.leitura?FOTOS.leitura.size:'')+'|'+$('f-texto').value.trim(); }
function talvezLer(){ if(!FOTOS.nota && !$('f-texto').value.trim()) return; const h=hashEntrada(); if(h===LIDO_HASH||LENDO) return; ler(); }

let progTimer=null;
function progresso(on,txt){
  const p=$('prog'), i=$('prog-i'); $('prog-tx').textContent=txt||'';
  if(on){ p.classList.remove('hide'); let v=8; i.style.width=v+'%'; clearInterval(progTimer); progTimer=setInterval(()=>{ v=Math.min(v+4,90); i.style.width=v+'%'; },500); }
  else { clearInterval(progTimer); i.style.width='100%'; setTimeout(()=>{ p.classList.add('hide'); i.style.width='0'; },500); }
}

async function ler(){
  LENDO=true; LIDO_HASH=hashEntrada(); $('leitura-flags').innerHTML='';
  progresso(true,'Enviando '+(FOTOS.nota?'nota':'')+(FOTOS.leitura?' + horímetro':'')+' e lendo…');
  try{
    const fd=new FormData();
    if(FOTOS.nota) fd.append('foto_nota',FOTOS.nota,'nota.jpg');
    if(FOTOS.leitura) fd.append('foto_leitura',FOTOS.leitura,'leitura.jpg');
    if($('f-texto').value.trim()) fd.append('texto_livre',$('f-texto').value.trim());
    const r=await api('/operacional/api/abastecimentos/extrair',{method:'POST',body:fd});
    PATHS=r.paths||{}; BRUTO=r.bruto||null;
    if(r.erro_foto) toast('⚠ A foto NÃO ficou guardada ('+r.erro_foto+'). O registro segue, mas anexe a foto depois em "Meus últimos registros".');
    aplicarSugestao(r);
    progresso(false,r.erro?'Leitura automática indisponível — preencha à mão.':'Lido. Confira os campos em laranja.');
    if(r.erro) $('leitura-flags').innerHTML='<div class="flag flag-warn">Leitura automática indisponível ('+esc(r.erro)+'). Preencha à mão — a foto ficou guardada.</div>';
  }catch(e){ progresso(false,''); toast('❌ '+e.message); }
  LENDO=false;
}
function setLido(id,v){ const el=$(id); if(v==null||v==='') return; el.value=(typeof v==='number')?String(v).replace('.',','):v; el.classList.add('lido'); }
function aplicarSugestao(s){
  mostrarDados();
  $('confianca').textContent = s.confianca?('· leitura '+s.confianca):'';
  if(s.posto){
    if(s.posto.fornecedor_id){ $('n-posto').value=s.posto.fornecedor_id; $('n-posto').classList.add('lido'); mostrar('posto-novo',false); }
    else if(s.posto.nome_lido){ $('n-posto').value='__novo'; mostrar('posto-novo',true); setLido('n-posto-nome',s.posto.nome_lido); setLido('n-cnpj',s.posto.cnpj_lido||''); }
  }
  setLido('n-cupom',s.cupom); setLido('n-data',s.data_hora);
  if(s.combustivel){ $('n-comb').value=s.combustivel; $('n-comb').classList.add('lido'); }
  setLido('n-litros',s.litros); setLido('n-preco',s.preco_litro); setLido('n-valor',s.valor_total);
  const its=(s.itens||[]).filter(i=>i.equipamento_id);
  its.forEach(i=>{ if(i.leitura!=null && i.medicao==='km' && i.leitura<1000){ i.leitura=null; } });
  if(its.length){
    $('itens').innerHTML='';
    its.forEach(i=>addItem(i.equipamento_id,i.litros,i.leitura,true));
    if(its.length===1 && num($('n-litros').value)!=null && !its[0].litros){ const li=$('itens').querySelector('.it-litros'); li.value=$('n-litros').value; li.classList.add('lido'); }
  }
  const flags=[];
  if(s.nao_casados&&s.nao_casados.length) flags.push('<div class="flag flag-warn">Não achei na frota: <b>'+s.nao_casados.map(esc).join(', ')+'</b>. Escolha o equipamento na lista.</div>');
  const susp=(s.itens||[]).find(i=>i.leitura_suspeita!=null);
  if(susp) flags.push('<div class="flag flag-warn">O número '+susp.leitura_suspeita+' parece consumo (KM/L), não leitura — digite o KM do odômetro.</div>');
  if(s.anotacoes) flags.push('<div class="flag flag-info">Anotação lida: '+esc(s.anotacoes)+'</div>');
  $('leitura-flags').innerHTML=flags.join('');
  recalcSoma(); its.forEach(i=>checarLeitura(i.equipamento_id));
}
function postoMudou(){ mostrar('posto-novo',$('n-posto').value==='__novo'); }
function recalcValor(origem){
  const l=num($('n-litros').value), p=num($('n-preco').value), v=num($('n-valor').value);
  if(origem==='preco' && l && p) $('n-valor').value=fmt(l*p,2);
  else if(origem==='valor' && l && v) $('n-preco').value=fmt(v/l,3);
}

// ── itens ──
let SEQ=0;
function addItem(eqId,litros,leitura,lido){
  const id='it'+(++SEQ);
  const d=document.createElement('div'); d.className='item'; d.id=id;
  d.innerHTML='<div class="top"><select class="it-eq'+(lido&&eqId?' lido':'')+'" onchange="itemMudou(\''+id+'\')">'+opcoesEquip(eqId)+'</select><button class="x" onclick="rmItem(\''+id+'\')">✕</button></div>'
    +'<div class="row"><div><label>Litros neste</label><input class="it-litros'+(lido&&litros!=null?' lido':'')+'" inputmode="decimal" placeholder="0,00" oninput="recalcSoma()" value="'+(litros!=null?fmt(litros,2):'')+'"></div>'
    +'<div><label class="it-lb-leitura">Horímetro / KM</label><input class="it-leitura'+(lido&&leitura!=null?' lido':'')+'" inputmode="decimal" placeholder="anotado na nota" oninput="checarLeituraEl(this)" value="'+(leitura!=null?String(leitura).replace('.',','):'')+'"></div></div>'
    +'<div class="sub"></div>';
  $('itens').appendChild(d);
  if(eqId) itemMudou(id,true);
}
function rmItem(id){ const el=$(id); if(el) el.remove(); recalcSoma(); }
function herdarLitrosUnico(){
  const rows=Array.from(document.querySelectorAll('.item')).filter(d=>d.querySelector('.it-eq').value);
  if(rows.length!==1) return;
  const li=rows[0].querySelector('.it-litros'), tot=$('n-litros').value;
  if(!li.value && num(tot)!=null){ li.value=tot; li.classList.add('lido'); }
}
function itemMudou(id,semSoma){
  const d=$(id), sel=d.querySelector('.it-eq'), eqId=sel.value;
  herdarLitrosUnico();
  const e=CTX.frota.find(x=>x.id===eqId); const lei=d.querySelector('.it-leitura'), lb=d.querySelector('.it-lb-leitura'), sub=d.querySelector('.sub');
  if(e && ehGalao(e)){ lei.value=''; lei.disabled=true; lei.placeholder='galão não tem leitura'; lb.textContent='Leitura'; const g=CTX.galoes.find(x=>x.id===eqId); sub.textContent='Galão · saldo atual '+fmt(g?g.saldo:0,1)+' L'+(g&&g.capacidade_l?' de '+fmt(g.capacidade_l,0):''); }
  else { lei.disabled=false; lei.placeholder='anotado na nota'; lb.textContent=(e&&e.medicao==='km')?'KM':'Horímetro (h)'; sub.textContent=''; if(eqId) checarLeitura(eqId); }
  if(!semSoma) recalcSoma();
}
async function checarLeitura(eqId){
  if(!eqId) return;
  if(!ULTIMAS[eqId]){ try{ ULTIMAS[eqId]=await api('/operacional/api/abastecimentos/ultima/'+eqId); }catch(e){ return; } }
  document.querySelectorAll('.item').forEach(d=>{ if(d.querySelector('.it-eq').value===eqId) checarLeituraEl(d.querySelector('.it-leitura')); });
}
function checarLeituraEl(inp){
  const d=inp.closest('.item'), eqId=d.querySelector('.it-eq').value, u=ULTIMAS[eqId], sub=d.querySelector('.sub'), v=num(inp.value);
  if(!u||u.galao) return;
  const un=u.medicao==='km'?' km':' h';
  inp.classList.remove('ruim');
  if(u.ultima_conhecida==null){ sub.textContent='Sem leitura anterior conhecida.'; return; }
  if(v!=null && v<u.ultima_conhecida){ inp.classList.add('ruim'); sub.innerHTML='<span style="color:#991B1B">Menor que a última conhecida ('+fmt(u.ultima_conhecida,0)+un+'). Confira antes de salvar — vai entrar como divergente.</span>'; }
  else sub.textContent='Última conhecida: '+fmt(u.ultima_conhecida,0)+un+(v!=null?' · +'+fmt(v-u.ultima_conhecida,0)+un:'');
}
function itensColeta(){
  return Array.from(document.querySelectorAll('.item')).map(d=>({equipamento_id:d.querySelector('.it-eq').value, litros:num(d.querySelector('.it-litros').value), leitura:num(d.querySelector('.it-leitura').value)})).filter(i=>i.equipamento_id||i.litros);
}
function recalcSoma(){
  const its=itensColeta(), soma=its.reduce((a,i)=>a+(i.litros||0),0), tot=num($('n-litros').value);
  const el=$('soma'); $('soma-v').textContent=fmt(soma,2)+' L'+(tot!=null?' de '+fmt(tot,2):'');
  el.className='soma'+(tot==null?'':(Math.abs(soma-tot)<=0.05?' ok':' ruim'));
}

// ── salvar ──
async function salvar(){
  const its=itensColeta();
  if(!$('n-comb').value){ toast('Escolha o combustível'); return; }
  const litros=num($('n-litros').value); if(!litros){ toast('Informe os litros da nota'); return; }
  if(!its.length||its.some(i=>!i.equipamento_id)){ toast('Escolha o equipamento de cada item'); return; }
  if(its.some(i=>!i.litros)){ toast('Informe os litros de cada item'); return; }
  const soma=its.reduce((a,i)=>a+i.litros,0); if(Math.abs(soma-litros)>0.05){ toast('Os itens somam '+fmt(soma,2)+' L e a nota tem '+fmt(litros,2)+' L'); return; }
  const postoSel=$('n-posto').value;
  const nota={numero_cupom:$('n-cupom').value.trim()||null, data_hora:$('n-data').value.trim()||null, combustivel:$('n-comb').value,
    litros_total:litros, preco_litro:num($('n-preco').value), valor_total:num($('n-valor').value),
    foto_nota:PATHS.nota||null, foto_leitura:PATHS.leitura||null, texto_livre:$('f-texto').value.trim()||null, bruto:BRUTO};
  if(postoSel==='__novo'){ nota.posto_nome=$('n-posto-nome').value.trim(); nota.cnpj=$('n-cnpj').value.trim()||null; if(!nota.posto_nome){ toast('Informe o nome do posto novo'); return; } }
  else if(postoSel) nota.fornecedor_id=postoSel;
  const b=$('btn-salvar'); b.disabled=true;
  try{
    const r=await api('/operacional/api/abastecimentos/nota',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nota:nota,itens:its})});
    mostrarResultado(r);
    const eqs=its.filter(i=>{ const e=CTX.frota.find(x=>x.id===i.equipamento_id); return e&&!ehGalao(e); });
    if(eqs.length===1) localStorage.setItem('garra_abast_ultimo_eq',eqs[0].equipamento_id);
    limpar(); carregarHist(); ULTIMAS={};
    window.scrollTo({top:document.body.scrollHeight,behavior:'smooth'});
  }catch(e){ toast('❌ '+e.message); }
  b.disabled=false;
}
function mostrarResultado(r){
  const linhas=(r.itens||[]).map(i=>{
    if(i.galao) return '<div class="flag flag-ok">🛢 <b>'+esc(i.equipamento)+'</b> recebeu '+fmt(i.litros,1)+' L · saldo '+fmt(i.saldo_galao,1)+' L</div>';
    const un=i.medicao==='km'?' km':' h';
    let h='<div class="flag '+(i.divergencia_leitura?'flag-warn':'flag-ok')+'">✅ <b>'+esc(i.equipamento)+'</b> · '+fmt(i.litros,1)+' L'+(i.leitura!=null?' · '+fmt(i.leitura,0)+un+' <span class="muted">('+esc(i.leitura_fonte)+')</span>':'');
    if(i.consumo_medio!=null) h+='<br>Consumo neste trecho: <b>'+fmt(i.consumo_medio,i.medicao==='km'?1:2)+(i.medicao==='km'?' L/100 km':' L/h')+'</b>'+(i.cmd_r!=null?' · média do equipamento '+fmt(i.cmd_r,i.medicao==='km'?1:2):'');
    if(i.divergencia_leitura) h+='<br>⚠️ Leitura menor que a última conhecida — registrada como divergente, não moveu o cadastro.';
    if(i.alerta_consumo) h+='<br>🔥 Consumo acima da própria média — avisar manutenção (filtro, bico, vazamento).';
    if(i.sem_funcionamento) h+='<br>⚠️ Nenhuma parte diária ou checklist desde o abastecimento anterior.';
    if(i.divergencia_placa) h+='<br>⚠️ Placa da foto não é deste equipamento.';
    return h+'</div>';
  }).join('');
  $('f-resultado').innerHTML='<div class="card"><b>Registrado</b> '+(r.cupom?'<span class="chip">cupom '+esc(r.cupom)+'</span>':'')+'<span class="chip">'+fmt(r.litros_total,2)+' L</span>'+(r.valor_total!=null?'<span class="chip">R$ '+fmt(r.valor_total,2)+'</span>':'')+linhas+'</div>';
}
function limpar(){
  FOTOS={nota:null,leitura:null}; PATHS={}; BRUTO=null; LIDO_HASH='';
  ['f-nota','f-leitura','g-f-nota','g-f-leitura'].forEach(id=>{ $(id).value=''; });
  ['nota','leitura'].forEach(q=>{ const b=$('box-'+q); b.classList.remove('ok'); const cur=$('ic-'+q); const s=document.createElement('span'); s.className='ic'; s.id='ic-'+q; s.textContent=q==='nota'?'🧾':'🕐'; cur.replaceWith(s); $('tx-'+q).textContent=q==='nota'?'Toque para fotografar · obrigatória':'Opcional — se não está anotado na nota'; });
  ['f-texto','n-cupom','n-data','n-litros','n-preco','n-valor','n-posto-nome','n-cnpj'].forEach(id=>{ $(id).value=''; $(id).classList.remove('lido'); });
  ['n-posto','n-comb'].forEach(id=>{ $(id).value=''; $(id).classList.remove('lido'); }); mostrar('posto-novo',false);
  $('confianca').textContent=''; $('leitura-flags').innerHTML=''; $('prog-tx').textContent='';
  $('itens').innerHTML=''; addItem(localStorage.getItem('garra_abast_ultimo_eq')||''); recalcSoma();
  mostrar('sec-dados',false); mostrar('rodape',false); const lk=$('lk-manual'); if(lk) lk.style.display='';
}

async function anexarFoto(id, inp){
  const f=inp.files&&inp.files[0]; if(!f) return;
  try{
    const b=await comprimir(f); const fd=new FormData(); fd.append('foto',b,'nota.jpg'); fd.append('tipo','nota');
    await api('/operacional/api/abastecimentos/'+id+'/foto',{method:'POST',body:fd});
    toast('✅ Foto anexada à nota'); carregarHist();
  }catch(e){ toast('❌ '+e.message); }
}
async function carregarHist(){
  try{
    const l=await api('/operacional/api/abastecimentos?limite=5&meus=1');
    $('hist').innerHTML=l.length?l.map(x=>{
      const d=new Date(x.data); const un=x.medicao==='km'?' km':' h';
      return '<div class="hist"><b>'+esc(x.equipamento)+'</b> · '+d.toLocaleDateString('pt-BR')+' · '+fmt(x.litros,1)+' L'
        +(x.destino_tipo==='galao'?' <span class="chip">galão</span>':(x.leitura!=null?' · '+fmt(x.leitura,0)+un:''))
        +(x.origem==='galao'?' <span class="chip">do '+esc(x.galao_codigo||'galão')+'</span>':'')
        +(x.posto?' · '+esc(x.posto):'')+(x.numero_cupom?' <span class="chip">'+esc(x.numero_cupom)+'</span>':'')
        +(x.divergencia_leitura?' <span class="chip w">divergente</span>':'')+(x.alerta_consumo?' <span class="chip r">consumo</span>':'')
        +'<div class="muted">'+esc(x.usuario_nome||'')+(x.combustivel_nome?' · '+esc(x.combustivel_nome):'')
        +(x.nota_id?(x.foto_nota_url?' · 🧾 foto ok':' · <span class="chip w">sem foto</span> <label style="color:#1A2A5E;font-weight:700;cursor:pointer">📎 anexar<input type="file" accept="image/*" capture="environment" style="display:none" onchange="anexarFoto(\''+x.id+'\',this)"></label>'):'')
        +'</div></div>';
    }).join(''):'<div class="muted">Nenhum registro ainda.</div>';
  }catch(e){}
}


// ── Ditado por voz (Web Speech, reconhecimento do próprio Android/Chrome) ──
// Toca no 🎤, fala, o texto SOMA no campo (falar de novo completa, não apaga).
// Sem suporte no navegador → botão nem aparece. Sem custo, sem chave.
(function(){
  const SR = window.SpeechRecognition || window.webkitSpeechRecognition;
  if (!SR) {
    document.querySelectorAll('[data-ditar]').forEach(function(btn){
      const d = document.createElement('div'); d.className = 'muted'; d.style.cssText = 'font-size:11px;margin-top:4px';
      d.textContent = 'Para ditar, use o 🎤 do teclado do celular.'; btn.replaceWith(d);
    });
    return;
  }
  document.querySelectorAll('[data-ditar]').forEach(function(btn){
    let rec = null, ouvindo = false;
    btn.addEventListener('click', function(ev){
      ev.preventDefault();
      const alvo = document.getElementById(btn.dataset.ditar);
      if (ouvindo) { try { rec.stop(); } catch(e){} return; }
      rec = new SR(); rec.lang = 'pt-BR'; rec.interimResults = false; rec.continuous = true;
      let base = alvo.value ? alvo.value.replace(/\s+$/, '') + ' ' : '';
      rec.onresult = function(e){
        let txt = '';
        for (let i = e.resultIndex; i < e.results.length; i++)
          if (e.results[i].isFinal) txt += e.results[i][0].transcript;
        if (txt) { base = (base + txt).replace(/\s+/g, ' ') + ' '; alvo.value = base; alvo.dispatchEvent(new Event('input')); }
      };
      rec.onstart = function(){ ouvindo = true; btn.textContent = '🔴 Falando… tocar para parar'; btn.classList.add('gravando'); };
      const parar = function(){ ouvindo = false; btn.textContent = '🎤 Ditar'; btn.classList.remove('gravando'); if (alvo.id === 'f-texto') talvezLer(); };
      rec.onend = parar; rec.onerror = parar;
      try { rec.start(); } catch(e){ parar(); }
    });
  });
})();

if (TOK) iniciar();
</script></body></html>
"""
