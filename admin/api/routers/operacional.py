"""routers.operacional — as 40 rotas do módulo Operacional (OS, partes
diárias, cadastros, controle mensal).

Refatoração Fase 2 · Etapa 2 (04/07/2026). Corpos IDÊNTICOS aos do main.py —
mudou apenas @app→@router e a origem dos imports (core/).
"""
import os, io, json, time, uuid, calendar, secrets
from datetime import datetime, timedelta, date
from typing import Optional
from fastapi import APIRouter, Request, HTTPException, Depends, Header, UploadFile, File, Form, Body
from fastapi.responses import RedirectResponse, FileResponse, HTMLResponse, RedirectResponse, JSONResponse, Response, StreamingResponse

from core.config import OPERACIONAL_DIR
from core.db import ajard_query, ajard_query_id, get_db
from core.auth import verificar_token, verificar_gestor, verificar_admin

router = APIRouter()

import re as _re_val

def _validar_medicao_parte(m):
    """(08/07/2026) Sanidade de medições — Regra 62 (paridade front/backend).
    A API não pode aceitar o que o front bloqueia. Levanta 400 com motivo claro.
    Fonte única: usada no POST (registro) e no PATCH (correção do operador)."""
    def _num(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None
    # Horas de relógio: formato HH:MM (00-23 / 00-59) — input type=time só protege o front
    for campo, rotulo in (("hora_inicio", "hora início"), ("hora_fim", "hora fim")):
        v = m.get(campo)
        if v not in (None, "") and not _re_val.match(r"^([01]\d|2[0-3]):[0-5]\d(:[0-5]\d)?$", str(v)):
            raise HTTPException(status_code=400, detail=f"Hora inválida em {rotulo}: {v}")
    # Nada de valores negativos em medição
    for campo, rotulo in (("horimetro_inicial", "horímetro inicial"), ("horimetro_final", "horímetro final"),
                          ("km_inicial", "KM inicial"), ("km_final", "KM final"),
                          ("qtd_metros", "metros"), ("qtd_viagens", "viagens"),
                          ("quantidade_diarias", "diárias")):
        v = _num(m.get(campo))
        if v is not None and v < 0:
            raise HTTPException(status_code=400, detail=f"Valor negativo em {rotulo}")
    # KM coerente
    k_ini, k_fin = _num(m.get("km_inicial")), _num(m.get("km_final"))
    if k_ini is not None and k_fin is not None and k_fin < k_ini:
        raise HTTPException(status_code=400, detail="KM final menor que o KM inicial")

def _validar_data_parte(data_str, permitir_futuro_dias=1):
    """Data ISO válida e não-futura (margem de 1 dia p/ fuso). Levanta 400."""
    try:
        dt = datetime.strptime(str(data_str), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"Data inválida: {data_str}")
    if dt > date.today() + timedelta(days=permitir_futuro_dias):
        raise HTTPException(status_code=400, detail="Data no futuro — confira o dia do registro")
    return dt

async def _validar_sobreposicao_horimetro(equipamento_id, h_ini, h_fin, ignorar_parte_id=None):
    """(13/07/2026) Horímetro só anda para frente: dois registros do MESMO
    equipamento com faixas que se cruzam é fisicamente impossível e dobra
    horas no CM. Bloqueia com o registro conflitante identificado.
    Faixas encostadas (final de um = inicial do outro) são o fluxo normal.
    (29/07/2026) Comparação com ROUND(,1) — precisão real das leituras:
    poeira de ponto flutuante gravada no banco (3612.6000000004) fazia
    faixa ENCOSTADA parecer sobreposta e barrava registro legítimo do
    campo (caso EH-39, Ítalo). Arredondar os DOIS lados mata a poeira
    de qualquer origem sem afrouxar a proteção real."""
    if not equipamento_id:
        return
    try:
        hi, hf = float(h_ini), float(h_fin)
    except (TypeError, ValueError):
        return
    if hf <= hi:
        return  # já barrado pela validação anterior
    conflito = await ajard_query(
        """SELECT p.data, p.horimetro_inicial, p.horimetro_final, os.numero
             FROM operacional.partes_diarias p
             JOIN operacional.ordens_servico os ON os.id = p.os_id
            WHERE p.equipamento_id = %s AND p.ativo = true
              AND p.horimetro_inicial IS NOT NULL AND p.horimetro_final IS NOT NULL
              AND p.id::text <> %s
              AND ROUND(p.horimetro_inicial::numeric, 1) < ROUND(%s::numeric, 1)
              AND ROUND(p.horimetro_final::numeric, 1)   > ROUND(%s::numeric, 1)
            ORDER BY p.data DESC LIMIT 1""",
        (equipamento_id, str(ignorar_parte_id or ""), hf, hi), fetch="one")
    if conflito:
        d = conflito["data"]
        d_fmt = d.strftime("%d/%m") if hasattr(d, "strftime") else str(d)
        # (31/07/2026) Duplicata EXATA (mesma faixa, arredondada a 1 casa):
        # quase sempre é o operador reenviando um dia que a fila offline já
        # subiu, ou que o escritório já lançou (caso Eliton 30/07). Mensagem
        # técnica de "sobreposição" confundia — dizer a verdade simples.
        try:
            _mesma_faixa = (round(float(conflito["horimetro_inicial"]), 1) == round(float(hi), 1)
                            and round(float(conflito["horimetro_final"]), 1) == round(float(hf), 1))
        except (TypeError, ValueError):
            _mesma_faixa = False
        if _mesma_faixa:
            raise HTTPException(status_code=400,
                detail=(f"Este dia já está registrado para esta máquina "
                        f"({conflito['numero']} em {d_fmt}, {conflito['horimetro_inicial']}→{conflito['horimetro_final']}) — "
                        f"por você ou pelo escritório. Confira no Histórico; não é preciso reenviar."))
        raise HTTPException(status_code=400,
            detail=(f"Horímetro {hi}→{hf} sobrepõe registro já existente desta máquina "
                    f"({conflito['numero']} em {d_fmt}: {conflito['horimetro_inicial']}→{conflito['horimetro_final']}). "
                    f"Confira as leituras — o horímetro só anda para frente."))

def _validar_horas_plausiveis(horas):
    """Teto de plausibilidade: um registro não pode ter mais de 24h.
    Mata o typo de horímetro (10125 → 101250 = 91.131h no Controle Mensal)."""
    if horas is not None and horas > 24:
        raise HTTPException(
            status_code=400,
            detail=f"{horas}h num único registro — confira o horímetro (mais de 24h não é permitido)")

def _cat_frota_checklist(codigo, categoria_op):
    """(09/07/2026) Fronteira entre taxonomias — LISTA BRANCA: só entra no
    espelho do checklist o que TEM checklist. CA-% → carro · caminhões →
    caminhao · máquinas motorizadas → maquinas. Caçambas estacionárias,
    gerador, componentes, moto, apoio e 'outro' ficam FORA (retorna None)."""
    cod = (codigo or "").upper()
    cat = (categoria_op or "").lower()
    if cod.startswith("CA-"):
        return "carro"
    if "caminh" in cat:
        return "caminhao"
    if cat in ("escavadeira", "retroescavadeira", "patrol", "carregadeira", "compactador"):
        return "maquinas"
    return None

def _calc_horas_parte(d):
    """(23/07/2026) REGRA OFICIAL DE HORAS — confirmada pela gestão:
    1º RELÓGIO explícito (início+fim): nasce VAZIO no formulário, então
       preenchido = intenção deliberada do operador (botão Hora). Carrega o
       julgamento humano (ex.: descontar parada por quebra) — PREVALECE.
       Com desconto de almoço quando a janela cruza o meio-dia.
    2º HORÍMETRO completo (fim-ini): assume as horas apenas sem relógio.
       É SEMPRE registrado como controle de máquina/manutenção.
    Divergência relógio × máquina → ⚠ para a GESTÃO decidir na Cobradas.
    O ÍCONE na tela segue a fonte real que determinou as horas.
    Na CORREÇÃO, a fonte editada por último vence (override no PATCH).
    Validação do horímetro roda mesmo quando o relógio prevalece."""
    # Horímetro: valida sempre (final < inicial é erro em qualquer cenário)
    h_maq = None
    h_ini = d.get("horimetro_inicial")
    h_fin = d.get("horimetro_final")
    if h_ini is not None and h_fin is not None:
        try:
            h_maq = round(float(h_fin) - float(h_ini), 2)
            if h_maq < 0:
                raise HTTPException(status_code=400, detail="Horímetro final menor que inicial")
        except (TypeError, ValueError):
            h_maq = None
    # 1º Relógio explícito
    hi = d.get("hora_inicio"); hf = d.get("hora_fim")
    if hi and hf:
        try:
            ph = lambda s: int(str(s)[:2]) * 60 + int(str(s)[3:5])
            ini_m = ph(hi); fim_m = ph(hf)
            diff = fim_m - ini_m
            if diff < 0: diff += 24 * 60
            bruto = diff / 60
            cruza_almoco = (ini_m < 12*60) and (fim_m > 12*60 or fim_m < ini_m)
            sem_almoco = bool(d.get("sem_almoco"))
            almoco = 1 if (not sem_almoco and cruza_almoco and bruto > 6) else 0
            return round(max(0, bruto - almoco), 2)
        except (TypeError, ValueError):
            pass
    # 2º Horímetro (sem relógio)
    return h_maq

def _horas_relogio(d):
    """Janela do relógio (com almoço) — usada no override de edição."""
    hi = d.get("hora_inicio"); hf = d.get("hora_fim")
    if not (hi and hf):
        return None
    try:
        ph = lambda s: int(str(s)[:2]) * 60 + int(str(s)[3:5])
        ini_m = ph(hi); fim_m = ph(hf)
        diff = fim_m - ini_m
        if diff < 0: diff += 24 * 60
        bruto = diff / 60
        cruza = (ini_m < 12*60) and (fim_m > 12*60 or fim_m < ini_m)
        almoco = 1 if (not bool(d.get("sem_almoco")) and cruza and bruto > 6) else 0
        return round(max(0, bruto - almoco), 2)
    except (TypeError, ValueError):
        return None

# (23/07/2026) AQUI EXISTIA um bloco duplicado dos helpers de parte
# (sobreposição/plausíveis/calc_horas) — resíduo de refatoração. Como em
# Python a última definição vence, a cópia ANTIGA de _calc_horas_parte
# (17/07: relógio manda) sobrescrevia silenciosamente a REGRA FINAL de
# 18/07 (horímetro completo manda; relógio fallback; divergência → gestão
# decide na Cobradas). Bloco removido; a regra final voltou a valer.

@router.get("/operacional/manifest.json")
async def operacional_manifest():
    """PWA manifest para o mobile operacional."""
    return {
        "name": "Garra Operacional",
        "short_name": "Garra Op",
        "start_url": "/operacional/mobile",
        "display": "standalone",
        "background_color": "#F0F4FF",
        "theme_color": "#1A2A5E",
        "icons": [
            {"src": "/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    }

@router.get("/operacional/mobile")
async def operacional_mobile():
    # (07/07/2026) Protótipo antigo aposentado (operacional-mobile.html
    # removido do repo) — link legado redireciona ao app oficial.
    return RedirectResponse(url="/mobile", status_code=307)

@router.get("/operacional/api/tipos-servico")
async def op_listar_tipos_servico(_auth=Depends(verificar_token)):
    """Lista tipos de serviço ativos para popular select."""
    rows = await ajard_query(
        "SELECT id, nome, descricao, medicao FROM operacional.tipos_servico WHERE ativo=true ORDER BY nome"
    )
    return [dict(r) for r in (rows or [])]

@router.post("/operacional/api/tipos-servico")
async def op_criar_tipo_servico(request: Request, payload=Depends(verificar_admin)):
    """Cria novo tipo de serviço (somente admin/gestor)."""
    d = await request.json()
    nome = (d.get("nome") or "").strip()
    descricao = (d.get("descricao") or "").strip()
    medicao = (d.get("medicao") or "horimetro").strip().lower()
    if medicao not in ("horimetro","hora","diaria","viagem","metros","km"):
        medicao = "horimetro"
    if not nome:
        raise HTTPException(status_code=400, detail="Nome é obrigatório")
    try:
        row = await ajard_query(
            """INSERT INTO operacional.tipos_servico (nome, descricao, medicao, ativo)
               VALUES (%s, %s, %s, true)
               RETURNING id, nome, descricao, medicao""",
            (nome, descricao or None, medicao),
            fetch="one"
        )
        return dict(row) if row else {"nome": nome, "medicao": medicao}
    except Exception as e:
        if "duplicate" in str(e).lower():
            raise HTTPException(status_code=409, detail="Tipo de serviço já existe")
        raise HTTPException(status_code=500, detail=str(e))

@router.patch("/operacional/api/tipos-servico/{tipo_id}")
async def op_editar_tipo_servico(tipo_id: str, request: Request, payload=Depends(verificar_admin)):
    """Edita tipo de serviço (somente admin/gestor)."""
    d = await request.json()
    updates = []
    valores = []
    if "nome" in d:
        nome = (d.get("nome") or "").strip()
        if not nome:
            raise HTTPException(status_code=400, detail="Nome não pode ser vazio")
        updates.append("nome=%s"); valores.append(nome)
    if "descricao" in d:
        updates.append("descricao=%s"); valores.append((d.get("descricao") or "").strip() or None)
    if "medicao" in d:
        med = (d.get("medicao") or "horimetro").strip().lower()
        if med not in ("horimetro","hora","diaria","viagem","metros","km"):
            med = "horimetro"
        updates.append("medicao=%s"); valores.append(med)
    if not updates:
        raise HTTPException(status_code=400, detail="Nada a atualizar")
    valores.append(tipo_id)
    try:
        row = await ajard_query(
            f"UPDATE operacional.tipos_servico SET {', '.join(updates)} WHERE id=%s RETURNING id, nome, descricao, medicao",
            tuple(valores),
            fetch="one"
        )
        if not row:
            raise HTTPException(status_code=404, detail="Tipo de serviço não encontrado")
        return dict(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/operacional/api/tipos-servico/{tipo_id}")
async def op_remover_tipo_servico(tipo_id: str, payload=Depends(verificar_admin)):
    """Soft delete (ativo=false) — preserva histórico."""
    try:
        await ajard_query(
            "UPDATE operacional.tipos_servico SET ativo=false WHERE id=%s",
            (tipo_id,), fetch="none"
        )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/operacional/api/regimes-cobranca")
async def op_listar_regimes(_auth=Depends(verificar_token)):
    rows = await ajard_query(
        "SELECT id, nome, descricao, valor_padrao FROM operacional.regimes_cobranca WHERE ativo=true ORDER BY nome"
    )
    return [dict(r) for r in (rows or [])]

@router.post("/operacional/api/regimes-cobranca")
async def op_criar_regime(request: Request, payload=Depends(verificar_gestor)):
    d = await request.json()
    nome = (d.get("nome") or "").strip().lower()
    descricao = (d.get("descricao") or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome é obrigatório")
    try:
        vp = d.get("valor_padrao")
        vp = float(vp) if vp not in (None, "") else None
        row = await ajard_query(
            """INSERT INTO operacional.regimes_cobranca (nome, descricao, valor_padrao, ativo)
               VALUES (%s, %s, %s, true) RETURNING id, nome, descricao, valor_padrao""",
            (nome, descricao or None, vp), fetch="one"
        )
        return dict(row) if row else {"nome": nome}
    except Exception as e:
        if "duplicate" in str(e).lower():
            raise HTTPException(status_code=409, detail="Regime já existe")
        raise HTTPException(status_code=500, detail=str(e))

@router.patch("/operacional/api/regimes-cobranca/{reg_id}")
async def op_editar_regime(reg_id: str, request: Request, payload=Depends(verificar_gestor)):
    d = await request.json()
    updates = []
    valores = []
    if "nome" in d:
        nome = (d.get("nome") or "").strip().lower()
        if not nome:
            raise HTTPException(status_code=400, detail="Nome não pode ser vazio")
        updates.append("nome=%s"); valores.append(nome)
    if "descricao" in d:
        updates.append("descricao=%s"); valores.append((d.get("descricao") or "").strip() or None)
    if "valor_padrao" in d:
        vp = d.get("valor_padrao")
        updates.append("valor_padrao=%s"); valores.append(float(vp) if vp not in (None, "") else None)
    if not updates:
        raise HTTPException(status_code=400, detail="Nada a atualizar")
    valores.append(reg_id)
    try:
        row = await ajard_query(
            f"UPDATE operacional.regimes_cobranca SET {', '.join(updates)} WHERE id=%s RETURNING id, nome, descricao, valor_padrao",
            tuple(valores), fetch="one"
        )
        if not row:
            raise HTTPException(status_code=404, detail="Regime não encontrado")
        return dict(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/operacional/api/regimes-cobranca/{reg_id}")
async def op_remover_regime(reg_id: str, payload=Depends(verificar_gestor)):
    try:
        await ajard_query("UPDATE operacional.regimes_cobranca SET ativo=false WHERE id=%s", (reg_id,), fetch="none")
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/operacional/api/equipamentos")
async def op_listar_equipamentos(uso: str = None, _auth=Depends(verificar_token)):
    """Lista equipamentos ativos para popular select e tela de cadastro.
    (07/07/2026) uso=os → apenas MOTORIZADOS (máquinas, caminhões e o
    Apoio/Combinado): a Onda 2 do ManWinWin trouxe caçambas estacionárias,
    pneus e componentes que não entram em OS — mas seguem visíveis para a
    Manutenção (sem o parâmetro)."""
    filtro_uso = ""
    if (uso or "").lower() == "os":
        # (09/07/2026) LISTA BRANCA — determinística: OS aceita apenas
        # caminhões, máquinas motorizadas e Apoio/Combinado. Moto, carro de
        # apoio, componentes, implementos e 'outro' ficam fora por definição.
        # Máquina nova só entra em OS com a categoria certa no Admin.
        filtro_uso = ("AND lower(coalesce(eq.categoria,'')) IN "
                      "('caminhao','escavadeira','retroescavadeira','patrol',"
                      "'carregadeira','compactador','apoio')")
    rows = await ajard_query(
        f"""SELECT eq.id, eq.codigo, eq.descricao, eq.categoria, eq.medicao, eq.agenda_ics_url,
                  eq.marca, eq.modelo, eq.ano, eq.placa,
                  -- (13/07/2026) Horímetro atual VIVO: maior leitura registrada
                  -- nas partes (a coluna estática ninguém atualizava — por isso
                  -- a sugestão não aparecia para o operador conferir)
                  COALESCE((SELECT MAX(p.horimetro_final)
                              FROM operacional.partes_diarias p
                             WHERE p.equipamento_id = eq.id AND p.ativo = true),
                           eq.horimetro_atual) AS horimetro_atual,
                  eq.km_atual, eq.ativo,
                  eq.operador_responsavel_id,
                  resp.nome AS operador_responsavel_nome
           FROM operacional.equipamentos eq
           LEFT JOIN public.usuarios_garra resp ON resp.id = eq.operador_responsavel_id
           WHERE eq.ativo=true {filtro_uso}
           ORDER BY eq.categoria, eq.codigo"""
    )
    return [dict(r) for r in (rows or [])]

@router.post("/operacional/api/equipamentos")
async def op_criar_equipamento(request: Request, payload=Depends(verificar_gestor)):
    """Cria equipamento — sincroniza em operacional.equipamentos + checklist.frota."""
    d = await request.json()
    codigo    = (d.get("codigo") or "").strip()
    descricao = (d.get("descricao") or "").strip()
    categoria = (d.get("categoria") or "").strip().lower()
    medicao   = (d.get("medicao") or "horimetro").strip().lower()
    if medicao not in ("horimetro","hora","diaria","viagem","metros","km"):
        medicao = "horimetro"
    if not codigo or not descricao or not categoria:
        raise HTTPException(status_code=400, detail="Código, descrição e categoria são obrigatórios")
    marca  = (d.get("marca")  or "").strip() or None
    modelo = (d.get("modelo") or "").strip() or None
    ano    = d.get("ano")
    placa  = (d.get("placa")  or "").strip() or None
    operador_resp = (d.get("operador_responsavel_id") or "").strip() or None
    agenda_ics_url = (d.get("agenda_ics_url") or "").strip() or None
    try:
        # (20/07/2026) BUG CPO-36: o soft delete mantém o registro no banco com
        # ativo=false — o equipamento some da relação, mas o UNIQUE de codigo
        # barra o recadastro com "Código já existe". Beco sem saída para o
        # admin. Correção: se o código pertence a um equipamento INATIVO,
        # reativa o MESMO registro (preserva id + todo o histórico de OS e
        # partes vinculado) e atualiza os dados informados no formulário.
        # Duplicado ATIVO segue retornando 409 normalmente.
        inativo = await ajard_query(
            "SELECT id FROM operacional.equipamentos WHERE codigo=%s AND ativo=false",
            (codigo,), fetch="one"
        )
        if inativo:
            row = await ajard_query(
                """UPDATE operacional.equipamentos
                   SET descricao=%s, categoria=%s, medicao=%s, marca=%s,
                       modelo=%s, ano=%s, placa=%s, operador_responsavel_id=%s,
                       agenda_ics_url=%s, ativo=true, atualizado_em=now()
                   WHERE id=%s
                   RETURNING id, codigo, descricao, categoria, medicao,
                             marca, modelo, ano, placa, operador_responsavel_id""",
                (descricao, categoria, medicao, marca, modelo, ano, placa,
                 operador_resp, agenda_ics_url, inativo["id"]),
                fetch="one"
            )
            out = dict(row) if row else {"codigo": codigo}
            out["reativado"] = True
            return out
        row = await ajard_query(
            """INSERT INTO operacional.equipamentos
               (codigo, descricao, categoria, medicao, marca, modelo, ano, placa, operador_responsavel_id, agenda_ics_url, ativo)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, true)
               RETURNING id, codigo, descricao, categoria, medicao,
                         marca, modelo, ano, placa, operador_responsavel_id""",
            (codigo, descricao, categoria, medicao, marca, modelo, ano, placa, operador_resp, agenda_ics_url),
            fetch="one"
        )
        # (09/07/2026) Sync com checklist.frota REMOVIDO — o checklist lê
        # direto do cadastro único via /frota-checklist. Espelho aposentado.
        return dict(row) if row else {"codigo": codigo}
    except Exception as e:
        if "duplicate" in str(e).lower():
            raise HTTPException(status_code=409, detail="Código já existe")
        raise HTTPException(status_code=500, detail=str(e))

@router.patch("/operacional/api/equipamentos/{eq_id}")
async def op_editar_equipamento(eq_id: str, request: Request, payload=Depends(verificar_gestor)):
    """Edita equipamento — sincroniza com checklist.frota."""
    d = await request.json()
    updates = []
    valores = []
    for campo, key in [("codigo","codigo"),("descricao","descricao"),
                       ("categoria","categoria"),("medicao","medicao"),
                       ("marca","marca"),("modelo","modelo"),
                       ("ano","ano"),("placa","placa"),
                       ("operador_responsavel_id","operador_responsavel_id"),
                       ("agenda_ics_url","agenda_ics_url")]:
        if key in d:
            valor = d.get(key)
            if isinstance(valor, str):
                valor = valor.strip() or None
                if campo in ("categoria","medicao") and valor:
                    valor = valor.lower()
            updates.append(f"{campo}=%s"); valores.append(valor)
    if not updates:
        raise HTTPException(status_code=400, detail="Nada a atualizar")
    valores.append(eq_id)
    try:
        row = await ajard_query(
            f"""UPDATE operacional.equipamentos SET {', '.join(updates)}
                WHERE id=%s
                RETURNING id, codigo, descricao, categoria, medicao,
                          marca, modelo, ano, placa""",
            tuple(valores),
            fetch="one"
        )
        if not row:
            raise HTTPException(status_code=404, detail="Equipamento não encontrado")
        # (09/07/2026) Sync com espelho removido — leitura direta no checklist.
        return dict(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/operacional/api/equipamentos/{eq_id}")
async def op_remover_equipamento(eq_id: str, payload=Depends(verificar_gestor)):
    """Soft delete — desativa em ambas as tabelas."""
    try:
        row = await ajard_query(
            "UPDATE operacional.equipamentos SET ativo=false WHERE id=%s RETURNING codigo",
            (eq_id,), fetch="one"
        )
        # (09/07/2026) Espelho aposentado — nada a propagar.
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/operacional/api/clientes")
async def op_listar_clientes(_auth=Depends(verificar_token)):
    """Lista clientes ativos para popular select da OS e tela de cadastro."""
    rows = await ajard_query(
        """SELECT id, nome, cnpj_cpf, telefone, email, contato, ativo
           FROM public.clientes_garra
           WHERE ativo=true OR ativo IS NULL
           ORDER BY nome"""
    )
    return [dict(r) for r in (rows or [])]

@router.post("/operacional/api/clientes")
async def op_criar_cliente(request: Request, payload=Depends(verificar_gestor)):
    """Cria novo cliente."""
    d = await request.json()
    nome = (d.get("nome") or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome é obrigatório")
    try:
        row = await ajard_query(
            """INSERT INTO public.clientes_garra
               (nome, cnpj_cpf, telefone, email, contato, ativo)
               VALUES (%s,%s,%s,%s,%s, true)
               RETURNING id, nome, cnpj_cpf, telefone, email, contato""",
            (nome,
             (d.get("cnpj_cpf") or "").strip() or None,
             (d.get("telefone") or "").strip() or None,
             (d.get("email") or "").strip() or None,
             (d.get("contato") or "").strip() or None),
            fetch="one"
        )
        return dict(row) if row else {"nome": nome}
    except Exception as e:
        if "duplicate" in str(e).lower():
            raise HTTPException(status_code=409, detail="Cliente já existe")
        raise HTTPException(status_code=500, detail=str(e))

@router.patch("/operacional/api/clientes/{cli_id}")
async def op_editar_cliente(cli_id: str, request: Request, payload=Depends(verificar_gestor)):
    """Edita cliente."""
    d = await request.json()
    updates = []
    valores = []
    for campo in ("nome","cnpj_cpf","telefone","email","contato"):
        if campo in d:
            val = d.get(campo)
            if isinstance(val, str):
                val = val.strip() or None
            updates.append(f"{campo}=%s"); valores.append(val)
    if not updates:
        raise HTTPException(status_code=400, detail="Nada a atualizar")
    valores.append(cli_id)
    try:
        row = await ajard_query(
            f"""UPDATE public.clientes_garra SET {', '.join(updates)}
                WHERE id=%s
                RETURNING id, nome, cnpj_cpf, telefone, email, contato""",
            tuple(valores), fetch="one"
        )
        if not row:
            raise HTTPException(status_code=404, detail="Cliente não encontrado")
        return dict(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/operacional/api/clientes/{cli_id}")
async def op_remover_cliente(cli_id: str, payload=Depends(verificar_gestor)):
    """Soft delete cliente."""
    try:
        await ajard_query(
            "UPDATE public.clientes_garra SET ativo=false WHERE id=%s",
            (cli_id,), fetch="none"
        )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/operacional/api/operadores")
async def op_listar_operadores(_auth=Depends(verificar_token)):
    """Lista usuários elegíveis a operar equipamentos (operadores, motoristas, campo)."""
    rows = await ajard_query(
        """SELECT id, nome, login, perfil
           FROM public.usuarios_garra
           WHERE ativo=true AND perfil IN ('operador','motorista','campo')
           ORDER BY nome"""
    )
    return [dict(r) for r in (rows or [])]

@router.get("/operacional/api/proximo-numero")
async def op_proximo_numero(_auth=Depends(verificar_gestor)):
    """Retorna o próximo número de OS disponível para o ano atual."""
    ano = datetime.utcnow().year
    row = await ajard_query(
        "SELECT operacional.proximo_numero_os(%s) AS numero",
        (ano,), fetch="one"
    )
    return {"numero": row["numero"], "ano": ano}

@router.post("/operacional/api/os")
async def op_criar_os(request: Request, payload=Depends(verificar_gestor)):
    """Cria nova OS. Somente admin, gestor ou luana."""
    d = await request.json()

    # Campos obrigatórios
    obra = (d.get("obra") or "").strip()
    if not obra:
        raise HTTPException(status_code=400, detail="Obra é obrigatória")

    # Cliente: ou cliente_id (cadastrado) OU cliente_nome_avulso
    cliente_id = d.get("cliente_id")
    cliente_nome_avulso = (d.get("cliente_nome_avulso") or "").strip() or None
    if not cliente_id and not cliente_nome_avulso:
        raise HTTPException(status_code=400, detail="Informe cliente cadastrado ou nome avulso")

    tipo_servico_id     = d.get("tipo_servico_id")
    equipamento_id      = d.get("equipamento_id") or None
    operador_id         = d.get("operador_id") or None
    endereco            = (d.get("endereco") or "").strip() or None
    descricao           = (d.get("descricao") or "").strip() or None
    data_inicio         = d.get("data_inicio") or datetime.utcnow().date()
    data_fim_prevista   = d.get("data_fim_prevista") or None
    codigo_erp          = (d.get("codigo_erp") or "").strip() or None
    origem              = d.get("origem") or "escritorio"
    # (21/07/2026) Observação da gestão — visível SÓ no admin (pode conter
    # condições financeiras vindas da Agenda Google). Nunca vai ao mobile.
    observacao_gestao   = (d.get("observacao_gestao") or "").strip() or None

    # Gerar número de OS
    ano = datetime.utcnow().year
    row = await ajard_query("SELECT operacional.proximo_numero_os(%s) AS numero", (ano,), fetch="one")
    numero = row["numero"]
    sequencia = int(numero.split("-")[-1])

    # Status inicial baseado em ter ou não codigo_erp
    status = "aberta_completa" if codigo_erp else "aberta_sem_erp"

    # Snapshot do criador
    criado_por = payload.get("sub")  # login
    user_row = await ajard_query("SELECT id FROM public.usuarios_garra WHERE login=%s",
                         (criado_por,), fetch="one")
    criado_por_id = user_row["id"] if user_row else None

    codigo_erp_em  = datetime.utcnow() if codigo_erp else None
    codigo_erp_por = criado_por_id if codigo_erp else None

    try:
        os_row = await ajard_query_id(
            """INSERT INTO operacional.ordens_servico
               (numero, ano, sequencia, codigo_erp, codigo_erp_em, codigo_erp_por,
                cliente_id, cliente_nome_avulso, tipo_servico_id,
                equipamento_id, operador_id,
                obra, endereco, descricao, observacao_gestao,
                data_inicio, data_fim_prevista,
                status, origem, criado_por, horas_padrao_dia)
               VALUES (%s,%s,%s,%s,%s,%s, %s,%s,%s, %s,%s, %s,%s,%s,%s, %s,%s, %s,%s,%s, %s)""",
            (numero, ano, sequencia, codigo_erp, codigo_erp_em, codigo_erp_por,
             cliente_id, cliente_nome_avulso, tipo_servico_id,
             equipamento_id, operador_id,
             obra, endereco, descricao, observacao_gestao,
             data_inicio, data_fim_prevista,
             status, origem, criado_por_id,
             float(d.get("horas_padrao_dia")) if d.get("horas_padrao_dia") else None)
        )
        return dict(os_row)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao criar OS: {str(e)}")

@router.get("/operacional/api/os")
async def op_listar_os(
    status: Optional[str] = None,
    cliente_id: Optional[str] = None,
    ano: Optional[int] = None,
    busca: Optional[str] = None,
    limit: int = 100,
    _auth=Depends(verificar_token)
):
    """Lista OS com filtros opcionais."""
    sql = """
        SELECT
            os.id, os.numero, os.ano, os.sequencia,
            os.codigo_erp, os.obra, os.endereco, os.descricao, os.observacao_gestao,
            os.data_inicio, os.data_fim_prevista, os.data_fim_real,
            os.status, os.origem, os.criado_em,
            os.regime_cobranca, os.valor_combinado, os.horas_padrao_dia,
            os.valor_hora, os.valor_metro, os.valor_diaria, os.valor_km, os.valor_viagem,
            os.cliente_id, COALESCE(c.nome, os.cliente_nome_avulso) AS cliente_nome,
            os.cliente_nome_avulso,
            os.tipo_servico_id, ts.nome AS tipo_servico_nome,
            os.equipamento_id, eq.codigo AS equipamento_codigo, eq.descricao AS equipamento_descricao,
            eq.medicao AS equipamento_medicao,
            eq.operador_responsavel_id AS equipamento_responsavel_id,
            resp.nome AS equipamento_responsavel_nome,
            os.operador_id, op.nome AS operador_nome,
            u.nome AS criado_por_nome
        FROM operacional.ordens_servico os
        LEFT JOIN public.clientes_garra c       ON c.id = os.cliente_id
        LEFT JOIN operacional.tipos_servico ts  ON ts.id = os.tipo_servico_id
        LEFT JOIN operacional.equipamentos eq   ON eq.id = os.equipamento_id
        LEFT JOIN public.usuarios_garra op      ON op.id = os.operador_id
        LEFT JOIN public.usuarios_garra resp    ON resp.id = eq.operador_responsavel_id
        LEFT JOIN public.usuarios_garra u       ON u.id = os.criado_por
        WHERE os.ativo = true
    """
    params = []
    if status:
        sql += " AND os.status = %s"; params.append(status)
    if cliente_id:
        sql += " AND os.cliente_id = %s"; params.append(cliente_id)
    if ano:
        sql += " AND os.ano = %s"; params.append(ano)
    if busca:
        sql += " AND (os.numero ILIKE %s OR os.obra ILIKE %s OR os.codigo_erp ILIKE %s)"
        like = f"%{busca}%"
        params.extend([like, like, like])
    sql += " ORDER BY os.ano DESC, os.sequencia DESC LIMIT %s"
    params.append(limit)

    rows = await ajard_query(sql, tuple(params))
    saida = [dict(r) for r in (rows or [])]
    # (21/07/2026) observacao_gestao pode conter condições financeiras —
    # menor privilégio: só admin/gestor/luana recebem o campo.
    if (_auth.get("perfil") or "").lower() not in ("admin", "gestor", "luana"):
        for r in saida:
            r.pop("observacao_gestao", None)
    return saida

@router.get("/operacional/api/os/{os_id}")
async def op_detalhe_os(os_id: str, _auth=Depends(verificar_token)):
    """Retorna detalhe completo da OS, com partes diárias."""
    row = await ajard_query(
        """SELECT os.*,
                  COALESCE(c.nome, os.cliente_nome_avulso) AS cliente_nome,
                  ts.nome AS tipo_servico_nome,
                  COALESCE((SELECT MAX(p2.horimetro_final)
                              FROM operacional.partes_diarias p2
                             WHERE p2.equipamento_id = eq.id AND p2.ativo = true),
                           eq.horimetro_atual) AS equipamento_horimetro_atual,
                  u.nome AS criado_por_nome
           FROM operacional.ordens_servico os
           LEFT JOIN public.clientes_garra c       ON c.id = os.cliente_id
           LEFT JOIN operacional.tipos_servico ts  ON ts.id = os.tipo_servico_id
           LEFT JOIN operacional.equipamentos eq   ON eq.id = os.equipamento_id
           LEFT JOIN public.usuarios_garra u       ON u.id = os.criado_por
           LEFT JOIN public.usuarios_garra op      ON op.id = os.operador_id
           WHERE os.id = %s AND os.ativo = true""",
        (os_id,), fetch="one"
    )
    if not row:
        raise HTTPException(status_code=404, detail="OS não encontrada")

    # Buscar partes diárias da OS
    partes = await ajard_query(
        """SELECT pd.*,
                  e.codigo AS equipamento_codigo, e.descricao AS equipamento_descricao,
                  e.categoria AS equipamento_categoria,
                  COALESCE(u.nome, pd.operador_nome_avulso) AS operador_nome
           FROM operacional.partes_diarias pd
           LEFT JOIN operacional.equipamentos e  ON e.id = pd.equipamento_id
           LEFT JOIN public.usuarios_garra u    ON u.id = pd.operador_id
           WHERE pd.os_id = %s AND pd.ativo = true
           ORDER BY pd.data DESC, pd.criado_em DESC""",
        (os_id,)
    )

    os_dict = dict(row)
    if (_auth.get("perfil") or "").lower() not in ("admin", "gestor", "luana"):
        os_dict.pop("observacao_gestao", None)
    os_dict["partes_diarias"] = [dict(p) for p in (partes or [])]
    return os_dict

@router.patch("/operacional/api/os/{os_id}")
async def op_atualizar_os(os_id: str, request: Request, payload=Depends(verificar_gestor)):
    """Atualiza OS — útil para inserir codigo_erp retroativo, mudar status, etc."""
    d = await request.json()

    # Verificar se OS existe
    existente = await ajard_query(
        "SELECT * FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one"
    )
    if not existente:
        raise HTTPException(status_code=404, detail="OS não encontrada")

    # Campos editáveis
    campos_editaveis = ["codigo_erp", "obra", "endereco", "descricao", "observacao_gestao",
                        "data_fim_prevista", "data_fim_real", "status",
                        "tipo_servico_id", "cliente_id", "cliente_nome_avulso",
                        "equipamento_id", "operador_id",
                        "regime_cobranca", "valor_combinado", "data_inicio", "horas_padrao_dia",
                        "valor_hora", "valor_metro", "valor_diaria", "valor_km", "valor_viagem"]
    updates = []
    params = []
    for campo in campos_editaveis:
        if campo in d:
            val = d[campo]
            # valor_combinado: aceitar número, vazio vira NULL
            if campo in ("valor_combinado", "horas_padrao_dia",
                         "valor_hora", "valor_metro", "valor_diaria", "valor_km", "valor_viagem"):
                val = float(val) if (val not in (None, "")) else None
            elif val == "":
                val = None
            updates.append(f"{campo} = %s")
            params.append(val)

    if not updates:
        return dict(existente)

    # Snapshot quem inseriu codigo_erp
    if "codigo_erp" in d and d["codigo_erp"]:
        login = payload.get("sub")
        user = await ajard_query("SELECT id FROM public.usuarios_garra WHERE login=%s",
                         (login,), fetch="one")
        if user:
            updates.append("codigo_erp_em = %s")
            params.append(datetime.utcnow())
            updates.append("codigo_erp_por = %s")
            params.append(user["id"])
        # Se tinha status aberta_sem_erp e agora tem erp, vira aberta_completa
        if existente["status"] == "aberta_sem_erp":
            updates.append("status = %s")
            params.append("aberta_completa")

    updates.append("atualizado_em = %s")
    params.append(datetime.utcnow())
    params.append(os_id)

    sql = f"UPDATE operacional.ordens_servico SET {', '.join(updates)} WHERE id = %s"
    await ajard_query(sql, tuple(params), fetch="none")

    # Retornar atualizado
    return await op_detalhe_os(os_id, _auth=payload)

@router.delete("/operacional/api/os/{os_id}")
async def op_remover_os(os_id: str, payload=Depends(verificar_token)):
    """(13/07/2026) Soft delete da OS, em camadas:
    - QUALQUER perfil: bloqueado se houver partes ativas (integridade do CM —
      trate os registros primeiro ou use Concluir).
    - admin/gestor: podem excluir qualquer OS sem partes.
    - operador/motorista: só a própria OS avulsa (origem=campo,
      status=aberta_sem_erp, criada por ele ou dele) — caso do toque duplo."""
    login = payload.get("sub", "")
    perfil = (payload.get("perfil") or "").lower()
    os_row = await ajard_query(
        "SELECT * FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one")
    if not os_row:
        raise HTTPException(status_code=404, detail="OS não encontrada")
    qtd = await ajard_query(
        "SELECT count(*) AS n FROM operacional.partes_diarias WHERE os_id=%s AND ativo=true",
        (os_id,), fetch="one")
    if (qtd or {}).get("n", 0) > 0:
        raise HTTPException(status_code=400,
            detail=f"Esta OS tem {qtd['n']} registro(s) de trabalho — exclua os registros primeiro ou use Concluir.")
    if perfil not in ("admin", "gestor"):
        user = await ajard_query(
            "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one")
        uid = str((user or {}).get("id") or "")
        eh_dono = uid and (str(os_row.get("criado_por") or "") == uid or str(os_row.get("operador_id") or "") == uid)
        if not (os_row.get("origem") == "campo"
                and os_row.get("status") == "aberta_sem_erp"
                and eh_dono):
            raise HTTPException(status_code=403,
                detail="Você só pode excluir OS avulsa aberta criada por você — para outras, fale com a gestão.")
    await ajard_query(
        "UPDATE operacional.ordens_servico SET ativo=false, atualizado_em=now() WHERE id=%s",
        (os_id,), fetch="none")
    return {"ok": True, "id": os_id}

@router.post("/operacional/api/os/{os_id}/partes")
async def op_criar_parte(os_id: str, request: Request, payload=Depends(verificar_token)):
    """Registra parte diária. Qualquer operador logado pode registrar numa OS ativa."""
    d = await request.json()
    login = payload.get("sub","")

    # Verificar se OS existe e está ativa
    os_row = await ajard_query(
        "SELECT * FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one"
    )
    if not os_row:
        raise HTTPException(status_code=404, detail="OS não encontrada")
    # (08/07/2026) OS concluída/cancelada não aceita novos registros
    if os_row.get("status") in ("concluida_completa", "concluida_sem_erp", "cancelada"):
        raise HTTPException(status_code=400, detail="OS concluída/cancelada — não aceita novos registros")

    # Campos obrigatórios
    data = d.get("data")
    vinculo = (d.get("vinculo_operador") or "proprio").lower()
    eh_terceiro = (vinculo == "terceiro")
    equipamento_terceiro = (d.get("equipamento_terceiro") or "").strip() or None
    # Terceiro: não usa equipamento da Garra; exige o nome livre
    if eh_terceiro:
        equipamento_id = None
        if not data or not equipamento_terceiro:
            raise HTTPException(status_code=400, detail="Data e equipamento de terceiro são obrigatórios")
    else:
        equipamento_id = d.get("equipamento_id") or os_row.get("equipamento_id")
        if not data or not equipamento_id:
            raise HTTPException(status_code=400, detail="Data e equipamento são obrigatórios")

    _validar_data_parte(data)
    _validar_medicao_parte(d)
    horas = _calc_horas_parte(d)
    _validar_horas_plausiveis(horas)
    await _validar_sobreposicao_horimetro(
        equipamento_id, d.get("horimetro_inicial"), d.get("horimetro_final"))
    h_ini = d.get("horimetro_inicial")
    h_fin = d.get("horimetro_final")

    # Buscar ID do operador pelo login (quem está logado = criado_por)
    user_row = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE (login=%s OR email=%s) AND ativo=true",
        (login, login), fetch="one"
    )
    criado_por_id = user_row["id"] if user_row else None

    # operador_id: pode ser informado no payload (Gilson registrando pelo Emilson)
    # ou default para quem está logado.
    # (08/07/2026) Terceiro/diarista/frete: quem opera NÃO é colaborador interno
    # (identificado em fornecedor/nome) — operador_id fica NULL para as horas
    # nunca somarem no resumo de nenhum colaborador.
    if vinculo in ("terceiro", "diarista", "frete"):
        operador_id = None
    else:
        operador_id = d.get("operador_id") or criado_por_id

    # Calcular KM percorrido
    km_ini = d.get("km_inicial")
    km_fin = d.get("km_final")
    km_perc = None
    if km_ini is not None and km_fin is not None:
        try: km_perc = round(float(km_fin) - float(km_ini), 1)
        except: pass

    # Quantidade de metros (serviços medidos por metragem)
    qtd_metros = d.get("qtd_metros")
    try:
        qtd_metros = float(qtd_metros) if qtd_metros not in (None, "") else None
    except (TypeError, ValueError):
        qtd_metros = None

    # "Considerar X horas" (padrão por OS — planilha da Luana): se a OS tem
    # horas_padrao_dia configurado e há horas reais, a cobrada nasce com o padrão.
    horas_cobradas_padrao = None
    try:
        os_pad = await ajard_query(
            "SELECT horas_padrao_dia FROM operacional.ordens_servico WHERE id=%s",
            (os_id,), fetch="one"
        )
        medicao_registro = (d.get("tipo_medicao") or "horimetro").lower()
        if (os_pad and os_pad.get("horas_padrao_dia") and horas and float(horas) > 0
                and medicao_registro not in ("metros", "km")):
            horas_cobradas_padrao = float(os_pad["horas_padrao_dia"])
    except Exception:
        pass

    # (08/07/2026) Idempotência: client_id gerado no celular. Se a resposta se
    # perder e a fila offline re-enviar, o índice único barra e devolvemos o
    # registro já salvo — fim da parte duplicada por rede instável.
    client_id = (d.get("client_id") or "").strip() or None
    try:
        parte = await ajard_query_id(
            """INSERT INTO operacional.partes_diarias
               (os_id, equipamento_id, operador_id, operador_nome_avulso,
                data, hora_inicio, hora_fim,
                tipo_medicao, horimetro_inicial, horimetro_final, horas_trabalhadas,
                km_inicial, km_final, km_percorrido,
                quantidade_diarias, qtd_viagens, qtd_metros,
                vinculo_operador, fornecedor, equipamento_terceiro, observacao, trajeto, por_conta_de, sem_almoco, criado_por,
                horas_cobradas, client_id)
               VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s,%s,%s,%s,%s,%s, %s,%s)""",
            (os_id, equipamento_id, operador_id, d.get("operador_nome_avulso"),
             data, d.get("hora_inicio"), d.get("hora_fim"),
             d.get("tipo_medicao","horimetro"), h_ini, h_fin, horas,
             km_ini, km_fin, km_perc,
             d.get("quantidade_diarias", 0), d.get("qtd_viagens", 0), qtd_metros,
             d.get("vinculo_operador","proprio"), d.get("fornecedor"), equipamento_terceiro,
             d.get("observacao"),
             d.get("trajeto"), d.get("por_conta_de","cliente"), bool(d.get("sem_almoco")),
             criado_por_id,
             horas_cobradas_padrao, client_id)
        )
        # Atualizar horímetro atual do equipamento (só equipamento próprio da Garra)
        if h_fin is not None and equipamento_id:
            await ajard_query(
                "UPDATE operacional.equipamentos SET horimetro_atual=%s, atualizado_em=now() WHERE id=%s",
                (h_fin, equipamento_id), fetch="none"
            )
        return dict(parte)
    except HTTPException:
        raise
    except Exception as e:
        # Retry da fila offline bateu no índice único do client_id → o registro
        # JÁ está salvo. Devolver como sucesso (idempotente), nunca duplicar.
        if client_id and ("client_id" in str(e).lower() or "uq_partes_client" in str(e).lower()):
            existente = await ajard_query(
                "SELECT * FROM operacional.partes_diarias WHERE client_id=%s",
                (client_id,), fetch="one"
            )
            if existente:
                r = dict(existente)
                r["dedup"] = True
                return r
        raise HTTPException(status_code=500, detail=f"Erro ao registrar parte: {str(e)}")

@router.get("/operacional/api/os/{os_id}/partes")
async def op_listar_partes(os_id: str, _auth=Depends(verificar_token)):
    """Lista todas as partes diárias de uma OS, com totais acumulados."""
    partes = await ajard_query(
        """SELECT pd.*,
                  COALESCE(e.codigo, pd.equipamento_terceiro) AS equipamento_codigo,
                  e.descricao AS equipamento_descricao,
                  e.categoria AS equipamento_categoria,
                  COALESCE(u.nome, pd.operador_nome_avulso) AS operador_nome
           FROM operacional.partes_diarias pd
           LEFT JOIN operacional.equipamentos e ON e.id = pd.equipamento_id
           LEFT JOIN public.usuarios_garra u    ON u.id = pd.operador_id
           WHERE pd.os_id = %s AND pd.ativo = true
           ORDER BY pd.data ASC, pd.criado_em ASC""",
        (os_id,)
    )
    lista = [dict(p) for p in (partes or [])]

    # Perfil do solicitante — operador/motorista NÃO vê horas cobradas (valores)
    perfil = _auth.get("perfil", "")
    eh_gestor = perfil in ("admin", "gestor", "luana", "bruna")

    if not eh_gestor:
        # Menor privilégio (decisão 05/07/2026): operador vê SOMENTE os
        # registros que ELE criou — horas do colega = comissão do colega.
        # A continuidade do horímetro é garantida pelo horimetro_atual do
        # equipamento (pré-preenchido no registro), não pela visão alheia.
        login = _auth.get("sub", "")
        eu = await ajard_query(
            "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
        )
        meu_id = str(eu["id"]) if eu else "?"
        lista = [p for p in lista
                 if str(p.get("criado_por") or "") == meu_id
                 or str(p.get("operador_id") or "") == meu_id]

    if not eh_gestor:
        # Remover campos financeiros/cobrança da resposta para operador
        for p in lista:
            p.pop("horas_cobradas", None)
            p.pop("quantidade_diarias_cobradas", None)
            p.pop("valor_calculado", None)
            p.pop("valor_unitario", None)

    # Totais acumulados — horas APENAS de registros medidos em hora
    # (mesma regra de conteúdo da comissão e do CM)
    def _conta_horas(p):
        med = (p.get("tipo_medicao") or "horimetro").lower()
        if med in ("metros", "viagem", "diaria", "km"):
            return False
        if float(p.get("qtd_metros") or 0) > 0:
            return False
        return True
    total_horas     = sum(float(p.get("horas_trabalhadas") or 0) for p in lista if _conta_horas(p))
    total_metros    = sum(float(p.get("qtd_metros") or 0) for p in lista)
    total_diarias   = sum(float(p.get("quantidade_diarias") or 0) for p in lista)
    total_viagens   = sum(float(p.get("qtd_viagens")       or 0) for p in lista)
    dias_trabalhados= len(set(str(p.get("data",""))[:10] for p in lista if p.get("data")))

    totais = {
        "dias_trabalhados":  dias_trabalhados,
        "total_horas":       round(total_horas, 2),
        "total_metros":      round(total_metros, 2),
        "total_diarias":     total_diarias,
        "total_viagens":     total_viagens,
    }
    # Total cobrado só para gestores
    if eh_gestor:
        total_horas_cob = sum(
            float(p.get("horas_cobradas") or 0) if float(p.get("horas_cobradas") or 0) > 0
            else float(p.get("horas_trabalhadas") or 0)
            for p in lista if _conta_horas(p)
        )
        totais["total_horas_cobradas"] = round(total_horas_cob, 2)

    return {"partes": lista, "totais": totais}

@router.patch("/operacional/api/partes/{parte_id}")
async def op_atualizar_parte(parte_id: str, request: Request, payload=Depends(verificar_gestor)):
    """Luana/Admin atualiza parte diária — ajusta horas cobradas, diárias, observação."""
    d = await request.json()
    existente = await ajard_query(
        "SELECT * FROM operacional.partes_diarias WHERE id=%s AND ativo=true",
        (parte_id,), fetch="one"
    )
    if not existente:
        raise HTTPException(status_code=404, detail="Parte diária não encontrada")
    if existente.get("fechado"):
        raise HTTPException(status_code=400, detail="Parte já fechada — não pode editar")

    campos = ["data","horas_cobradas","valor","valor_unitario","quantidade_diarias","quantidade_diarias_cobradas",
              "qtd_viagens","qtd_metros","observacao","hora_inicio","hora_fim",
              "horimetro_inicial","horimetro_final","km_inicial","km_final",
              "equipamento_id","operador_id","operador_nome_avulso",
              "vinculo_operador","fornecedor","por_conta_de","trajeto","sem_almoco"]
    updates, params = [], []
    for c in campos:
        if c in d:
            updates.append(f"{c} = %s")
            params.append(d[c] if d[c] != "" else None)

    # Recalcular horas se horímetro foi atualizado
    if "horimetro_inicial" in d or "horimetro_final" in d:
        h_ini = d.get("horimetro_inicial", existente.get("horimetro_inicial"))
        h_fin = d.get("horimetro_final",   existente.get("horimetro_final"))
        if h_ini is not None and h_fin is not None:
            horas = round(float(h_fin) - float(h_ini), 2)
            updates.append("horas_trabalhadas = %s")
            params.append(horas)

    # Recalcular km_percorrido se km foi atualizado
    if "km_inicial" in d or "km_final" in d:
        k_ini = d.get("km_inicial", existente.get("km_inicial"))
        k_fin = d.get("km_final",   existente.get("km_final"))
        if k_ini is not None and k_fin is not None:
            updates.append("km_percorrido = %s")
            params.append(round(float(k_fin) - float(k_ini), 1))

    if not updates:
        return dict(existente)

    params.append(parte_id)
    await ajard_query(
        f"UPDATE operacional.partes_diarias SET {', '.join(updates)} WHERE id=%s",
        tuple(params), fetch="none"
    )
    row = await ajard_query(
        "SELECT * FROM operacional.partes_diarias WHERE id=%s", (parte_id,), fetch="one"
    )
    return dict(row)

@router.delete("/operacional/api/partes/{parte_id}")
async def op_remover_parte(parte_id: str, _auth=Depends(verificar_gestor)):
    """Soft delete de parte diária."""
    existente = await ajard_query(
        "SELECT fechado FROM operacional.partes_diarias WHERE id=%s AND ativo=true",
        (parte_id,), fetch="one"
    )
    if not existente:
        raise HTTPException(status_code=404, detail="Parte não encontrada")
    if existente.get("fechado"):
        raise HTTPException(status_code=400, detail="Parte fechada — não pode remover")
    await ajard_query(
        "UPDATE operacional.partes_diarias SET ativo=false WHERE id=%s",
        (parte_id,), fetch="none"
    )
    return {"ok": True, "id": parte_id}

@router.post("/operacional/api/os/{os_id}/fechar")
async def op_fechar_os(os_id: str, request: Request, payload=Depends(verificar_gestor)):
    """Fecha OS após revisão pela Luana. Congela todas as partes diárias."""
    os_row = await ajard_query(
        "SELECT * FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one"
    )
    if not os_row:
        raise HTTPException(status_code=404, detail="OS não encontrada")
    if os_row.get("status") in ("concluida_completa","concluida_sem_erp","cancelada"):
        raise HTTPException(status_code=400, detail="OS já está fechada")

    login = payload.get("sub","")
    user  = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
    )
    fechado_por_id = user["id"] if user else None
    agora = datetime.utcnow()

    # Auto-preencher horas_cobradas = horas_trabalhadas onde não foi editado (cobradas=0 ou null)
    await ajard_query(
        """UPDATE operacional.partes_diarias pd
           SET horas_cobradas = COALESCE(os.horas_padrao_dia, pd.horas_trabalhadas)
           FROM operacional.ordens_servico os
           WHERE os.id = pd.os_id
             AND pd.os_id=%s AND pd.ativo=true AND pd.fechado=false
             AND (pd.horas_cobradas IS NULL OR pd.horas_cobradas = 0)
             AND pd.horas_trabalhadas > 0""",
        (os_id,), fetch="none"
    )

    # (24/07/2026) Preço por linha: congelar valor_unitario onde a gestão
    # não definiu — herda o preço da OS pela medição da parte (snapshot;
    # reajuste na OS nunca retroage em linha fechada).
    await ajard_query(
        """UPDATE operacional.partes_diarias pd
           SET valor_unitario = CASE
                 WHEN pd.tipo_medicao = 'metros' THEN os.valor_metro
                 WHEN pd.tipo_medicao = 'viagem' THEN os.valor_viagem
                 WHEN pd.tipo_medicao = 'diaria' THEN os.valor_diaria
                 WHEN pd.tipo_medicao = 'km' THEN
                   COALESCE(NULLIF(os.valor_viagem,0), os.valor_km)
                 ELSE os.valor_hora
               END
           FROM operacional.ordens_servico os
           WHERE os.id = pd.os_id
             AND pd.os_id=%s AND pd.ativo=true AND pd.fechado=false
             AND pd.valor_unitario IS NULL""",
        (os_id,), fetch="none"
    )

    # Fechar todas as partes abertas
    await ajard_query(
        """UPDATE operacional.partes_diarias
           SET fechado=true, fechado_em=%s, fechado_por=%s
           WHERE os_id=%s AND ativo=true AND fechado=false""",
        (agora, fechado_por_id, os_id), fetch="none"
    )

    # Determinar status final
    novo_status = "concluida_completa" if os_row.get("codigo_erp") else "concluida_sem_erp"

    await ajard_query(
        """UPDATE operacional.ordens_servico
           SET status=%s, data_fim_real=%s, atualizado_em=%s
           WHERE id=%s""",
        (novo_status, agora.date(), agora, os_id), fetch="none"
    )

    return await op_detalhe_os(os_id, _auth=payload)

@router.post("/operacional/api/os/{os_id}/partes/lote")
async def op_lancar_diarias_lote(os_id: str, request: Request,
                                 payload=Depends(verificar_gestor)):
    """(24/07/2026) Diárias em LOTE — locação seca (equipamento sem
    motorista, ninguém no campo para registrar). Cria 1 diária por dia no
    período (máx. 30 dias). Idempotente: dias que já têm diária ATIVA do
    mesmo equipamento nesta OS são pulados — rodar de novo não duplica e
    prorrogação só cria os dias novos. Operador é OPCIONAL: vazio grava
    operador_id NULL + rótulo "— locação" (nunca contamina resumo nem
    comissão de colaborador)."""
    d = await request.json()
    os_row = await ajard_query(
        "SELECT id, status FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one"
    )
    if not os_row:
        raise HTTPException(status_code=404, detail="OS não encontrada")
    if os_row.get("status") in ("concluida_sem_erp", "concluida_completa", "cancelada"):
        raise HTTPException(status_code=400, detail="OS concluída — reabra para lançar diárias")

    equipamento_id = (d.get("equipamento_id") or "").strip() or None
    if not equipamento_id:
        raise HTTPException(status_code=400, detail="Equipamento é obrigatório no lote")
    try:
        dt_ini = datetime.strptime(str(d.get("data_inicio")), "%Y-%m-%d").date()
        dt_fim = datetime.strptime(str(d.get("data_fim")), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Período inválido")
    if dt_fim < dt_ini:
        raise HTTPException(status_code=400, detail="Data final antes da inicial")
    if (dt_fim - dt_ini).days + 1 > 62:
        raise HTTPException(status_code=400, detail="Lote máximo: 62 dias por vez")

    incluir_fds = bool(d.get("incluir_fds", False))
    operador_id = (d.get("operador_id") or "").strip() or None
    vu = d.get("valor_unitario")
    try:
        vu = round(float(vu), 2) if vu not in (None, "",) else None
        if vu is not None and vu < 0: vu = None
    except (TypeError, ValueError):
        vu = None

    login = payload.get("sub")
    criador = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
    )
    criado_por_id = criador["id"] if criador else None

    # Dias já ocupados por diária ativa deste equipamento nesta OS
    existentes = await ajard_query(
        """SELECT data FROM operacional.partes_diarias
           WHERE os_id=%s AND equipamento_id=%s AND ativo=true
             AND tipo_medicao='diaria'""",
        (os_id, equipamento_id)
    )
    ocupados = {r["data"] for r in (existentes or [])}

    criadas, puladas_exist, puladas_fds = 0, 0, 0
    cursor = dt_ini
    while cursor <= dt_fim:
        if not incluir_fds and cursor.weekday() >= 5:
            puladas_fds += 1
        elif cursor in ocupados:
            puladas_exist += 1
        else:
            await ajard_query(
                """INSERT INTO operacional.partes_diarias
                   (os_id, equipamento_id, operador_id, operador_nome_avulso,
                    data, tipo_medicao, quantidade_diarias, qtd_viagens,
                    vinculo_operador, observacao, por_conta_de, sem_almoco,
                    criado_por, valor_unitario)
                   VALUES (%s,%s,%s,%s, %s,'diaria',1,0, 'proprio',%s,'cliente',false, %s,%s)""",
                (os_id, equipamento_id, operador_id,
                 None if operador_id else "— locação",
                 cursor, "Diária de locação (lote)", criado_por_id, vu),
                fetch="none"
            )
            criadas += 1
        cursor += timedelta(days=1)

    return {"criadas": criadas, "puladas_existentes": puladas_exist,
            "puladas_fds": puladas_fds,
            "de": dt_ini.isoformat(), "ate": dt_fim.isoformat()}

@router.post("/operacional/api/os/{os_id}/reabrir")
async def op_reabrir_os(os_id: str, payload=Depends(verificar_gestor)):
    """(20/07/2026) Reabre OS concluída — caminho de volta sancionado.
    O fechamento congela as partes (fechado=true) por integridade do
    faturamento; antes desta rota, um registro errado numa OS fechada era
    beco sem saída (sem editar, sem excluir). Reabrir descongela as partes,
    devolve o status de aberta e limpa data_fim_real. Ao corrigir, a gestão
    fecha de novo pelo fluxo normal (revisão → Fechar OS)."""
    os_row = await ajard_query(
        "SELECT id, status, codigo_erp FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one"
    )
    if not os_row:
        raise HTTPException(status_code=404, detail="OS não encontrada")
    if os_row.get("status") not in ("concluida_sem_erp", "concluida_completa"):
        raise HTTPException(status_code=400, detail="Só OS concluída pode ser reaberta")
    # Descongelar as partes desta OS
    await ajard_query(
        """UPDATE operacional.partes_diarias
           SET fechado=false, fechado_em=NULL, fechado_por=NULL
           WHERE os_id=%s AND ativo=true AND fechado=true""",
        (os_id,), fetch="none"
    )
    # Status de volta conforme presença do código ERP
    novo_status = "aberta_completa" if os_row.get("codigo_erp") else "aberta_sem_erp"
    await ajard_query(
        """UPDATE operacional.ordens_servico
           SET status=%s, data_fim_real=NULL, atualizado_em=%s
           WHERE id=%s""",
        (novo_status, datetime.utcnow(), os_id), fetch="none"
    )
    return await op_detalhe_os(os_id, _auth=payload)

@router.post("/operacional/api/os/{os_id}/concluir")
async def op_concluir_os_operador(os_id: str, request: Request, payload=Depends(verificar_token)):
    """Operador marca OS como concluída do lado dele → aguarda fechamento pela Luana."""
    login = payload.get("sub","") or payload.get("login","")
    user = await ajard_query(
        "SELECT id, perfil FROM public.usuarios_garra WHERE login=%s",
        (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    os_row = await ajard_query(
        "SELECT id, numero, status, operador_id FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one"
    )
    if not os_row:
        raise HTTPException(status_code=404, detail="OS não encontrada")

    # Verificar se o operador é dono da OS ou admin/gestor
    if user["perfil"] not in ("admin","gestor","luana") and str(os_row.get("operador_id")) != str(user["id"]):
        raise HTTPException(status_code=403, detail="Você não é o operador desta OS")

    if os_row.get("status") in ("concluida_completa","concluida_sem_erp","cancelada","aguardando_fechamento"):
        raise HTTPException(status_code=400, detail="OS já está concluída ou aguardando fechamento")

    # (18/07/2026) Caso Italo/OS-15: concluir OS SEM nenhum registro esconde a
    # OS do operador e engole o trabalho do dia. Registra primeiro, conclui depois.
    tem_parte = await ajard_query(
        "SELECT 1 FROM operacional.partes_diarias WHERE os_id=%s AND ativo=true LIMIT 1",
        (os_id,), fetch="one"
    )
    if not tem_parte:
        raise HTTPException(status_code=400,
            detail="Esta OS não tem nenhum registro. Registre o trabalho do dia antes de concluir.")

    agora = datetime.utcnow()
    await ajard_query(
        """UPDATE operacional.ordens_servico
           SET status='aguardando_fechamento', atualizado_em=%s
           WHERE id=%s""",
        (agora, os_id), fetch="none"
    )

    return {"ok": True, "numero": os_row.get("numero"), "status": "aguardando_fechamento"}

@router.post("/operacional/api/os/{os_id}/liberar")
async def op_liberar_os(os_id: str, payload=Depends(verificar_token)):
    """
    Operador LIBERA a OS (foi movido para outra obra/máquina).
    A OS perde o responsável (operador_id = NULL) → vira órfã → cai no alerta
    vermelho do admin para a Luana redesignar.
    As partes já registradas continuam com o operador original (comissão preservada).
    """
    login = payload.get("sub","") or payload.get("login","")
    user = await ajard_query(
        "SELECT id, perfil FROM public.usuarios_garra WHERE login=%s",
        (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    os_row = await ajard_query(
        "SELECT id, numero, status, operador_id FROM operacional.ordens_servico WHERE id=%s AND ativo=true",
        (os_id,), fetch="one"
    )
    if not os_row:
        raise HTTPException(status_code=404, detail="OS não encontrada")

    # Só o operador dono (ou admin/gestor) pode liberar
    if user["perfil"] not in ("admin","gestor","luana") and str(os_row.get("operador_id")) != str(user["id"]):
        raise HTTPException(status_code=403, detail="Você não é o operador desta OS")

    if os_row.get("status") in ("concluida_completa","concluida_sem_erp","cancelada"):
        raise HTTPException(status_code=400, detail="OS já está concluída")

    # Zera o responsável → OS órfã. Partes diárias NÃO são tocadas (mantêm operador_id).
    agora = datetime.utcnow()
    await ajard_query(
        """UPDATE operacional.ordens_servico
           SET operador_id=NULL, atualizado_em=%s
           WHERE id=%s""",
        (agora, os_id), fetch="none"
    )

    return {"ok": True, "numero": os_row.get("numero"), "liberada": True}

@router.get("/operacional/api/os/{os_id}/revisao")
async def op_revisao_os(os_id: str, _auth=Depends(verificar_gestor)):
    """Tela de revisão antes de fechar: OS + partes + totais consolidados."""
    os_detail = await op_detalhe_os(os_id, _auth=_auth)
    partes_data = await op_listar_partes(os_id, _auth=_auth)
    return {
        "os":     os_detail,
        "partes": partes_data["partes"],
        "totais": partes_data["totais"],
    }

@router.get("/operacional/api/minhas-partes")
async def op_minhas_partes(payload=Depends(verificar_token)):
    """Operador vê histórico de partes diárias próprias."""
    login = payload.get("sub","")
    user = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    rows = await ajard_query(
        """SELECT pd.*, os.numero AS numero_os, os.obra,
                  e.codigo AS equipamento_codigo
           FROM operacional.partes_diarias pd
           JOIN operacional.ordens_servico os ON os.id = pd.os_id
           LEFT JOIN operacional.equipamentos e ON e.id = pd.equipamento_id
           WHERE pd.operador_id = %s AND pd.ativo = true
           ORDER BY pd.data DESC, pd.criado_em DESC
           LIMIT 60""",
        (user["id"],)
    )
    return [dict(r) for r in (rows or [])]

@router.get("/operacional/api/minhas-os")
async def op_minhas_os(historico: int = 0, payload=Depends(verificar_token)):
    """Operador/motorista vê suas OS. historico=0: ativas | historico=1: concluídas."""
    login = payload.get("sub","")
    user  = await ajard_query(
        "SELECT id, perfil FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    # Verificar permissão do módulo operacional_mobile
    perm = await ajard_query(
        "SELECT permitido FROM public.permissoes_colaborador WHERE usuario_id=%s AND modulo='operacional_mobile'",
        (user["id"],), fetch="one"
    )
    if perm and perm.get("permitido") == False:
        # Permissão explicitamente negada → retorna lista vazia
        return []
    # Se não tem registro, usa padrão do perfil (operador/motorista/admin/gestor/luana/bruna têm acesso)
    if not perm:
        perfis_com_acesso = ('admin','gestor','luana','bruna','operador','motorista')
        if user.get("perfil") not in perfis_com_acesso:
            return []

    # Filtro de status conforme histórico ou ativas
    if historico == 1:
        # Histórico: só OS concluídas COM registro no MÊS CORRENTE
        # (decisão 05/07/2026 — operador não precisa ver OS antigas)
        status_filter = """os.status IN ('concluida_completa','concluida_sem_erp','aguardando_fechamento')
             AND EXISTS (SELECT 1 FROM operacional.partes_diarias p
                         WHERE p.os_id = os.id AND p.ativo = true
                           AND p.data >= date_trunc('month', CURRENT_DATE)::date)"""
    else:
        status_filter = "os.status NOT IN ('concluida_completa','concluida_sem_erp','cancelada')"

    rows = await ajard_query(
        f"""SELECT os.id, os.numero, os.obra, os.regime_cobranca, os.origem,
                  os.descricao AS observacao,
                  os.data_inicio, os.data_fim_prevista, os.status,
                  os.equipamento_id, os.operador_id, os.tipo_servico_id, os.cliente_id,
                  COALESCE(c.nome, os.cliente_nome_avulso) AS cliente_nome,
                  e.codigo AS equipamento_codigo, e.descricao AS equipamento_descricao,
                  e.medicao AS equipamento_medicao,
                  ts.nome AS tipo_servico_nome, ts.medicao AS tipo_servico_medicao
           FROM operacional.ordens_servico os
           LEFT JOIN public.clientes_garra c      ON c.id = os.cliente_id
           LEFT JOIN operacional.equipamentos e   ON e.id = os.equipamento_id
           LEFT JOIN operacional.tipos_servico ts ON ts.id = os.tipo_servico_id
           WHERE os.operador_id = %s
             AND os.ativo = true
             AND {status_filter}
           ORDER BY os.data_inicio DESC""",
        (user["id"],)
    )
    # Sem campos financeiros — operador não vê valores
    return [dict(r) for r in (rows or [])]

@router.get("/operacional/api/minhas-os/debug")
async def op_minhas_os_debug(payload=Depends(verificar_token)):
    """DEBUG — mostra todos os campos para diagnosticar por que OS não aparece."""
    login = payload.get("sub","") or payload.get("login","")
    user = await ajard_query(
        "SELECT id, login, nome, perfil FROM public.usuarios_garra WHERE login=%s",
        (login,), fetch="one"
    )
    if not user:
        return {"erro": "usuário não encontrado", "login_buscado": login}
    
    todas_os = await ajard_query(
        """SELECT id, numero, status, ativo, operador_id, obra
           FROM operacional.ordens_servico
           WHERE operador_id = %s
           ORDER BY data_inicio DESC""",
        (user["id"],)
    )
    
    os_visiveis = await ajard_query(
        """SELECT id, numero, status
           FROM operacional.ordens_servico
           WHERE operador_id = %s
             AND ativo = true
             AND status NOT IN ('concluida_completa','concluida_sem_erp','cancelada')""",
        (user["id"],)
    )
    
    return {
        "login_jwt": login,
        "usuario": dict(user),
        "total_os_vinculadas": len(todas_os or []),
        "todas_os": [dict(r) for r in (todas_os or [])],
        "os_visiveis_no_mobile": [dict(r) for r in (os_visiveis or [])]
    }

@router.post("/operacional/api/os/avulsa")
async def op_criar_os_avulsa(req: Request, payload=Depends(verificar_token)):
    """Operador cria OS avulsa do campo — sem código ERP, status aberta_sem_erp."""
    login = payload.get("sub","") or payload.get("login","")
    user = await ajard_query(
        "SELECT id, nome, perfil FROM public.usuarios_garra WHERE login=%s",
        (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Perfis autorizados
    if user["perfil"] not in ("operador", "motorista", "admin", "gestor", "luana"):
        raise HTTPException(status_code=403, detail="Perfil sem permissão para criar OS avulsa")
    
    body = await req.json()
    obra            = (body.get("obra") or "").strip()
    cliente_nome    = (body.get("cliente_nome") or "").strip()
    equipamento_id  = body.get("equipamento_id")
    tipo_servico_id = body.get("tipo_servico_id")
    # (08/07/2026) OS avulsa nasce SEM regime — o operador não escolhe regime;
    # a gestão define no complemento. Evita sugestão falsa de "diária" no
    # Registrar Dia (bloco de diárias abrindo em toda OS avulsa).
    regime_cobranca = (body.get("regime_cobranca") or "").strip()
    # (13/07/2026) Idempotência da OS avulsa (mesmo padrão validado das partes)
    client_id       = (body.get("client_id") or "").strip() or None
    observacao      = (body.get("observacao") or "").strip()
    
    if not obra:
        raise HTTPException(status_code=400, detail="Obra é obrigatória")
    
    # Gerar próximo número
    ano = datetime.utcnow().year
    ult = await ajard_query(
        "SELECT numero, sequencia FROM operacional.ordens_servico WHERE numero LIKE %s ORDER BY sequencia DESC LIMIT 1",
        (f"OS-{ano}-%",), fetch="one"
    )
    if ult and ult.get("sequencia"):
        seq = int(ult["sequencia"]) + 1
    elif ult:
        try:
            seq = int(ult["numero"].split("-")[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    numero = f"OS-{ano}-{seq:04d}"
    
    # (15/07/2026) SNAPSHOT da tabela de preços na criação — reajuste de
    # tabela nunca retroage em OS existente
    precos_padrao = {"hora": None, "metro": None, "diaria": None, "km": None, "viagem": None}
    try:
        regs = await ajard_query(
            "SELECT nome, valor_padrao FROM operacional.regimes_cobranca WHERE ativo=true")
        for r in (regs or []):
            n = (r.get("nome") or "").lower()
            if n in precos_padrao and r.get("valor_padrao") is not None:
                precos_padrao[n] = float(r["valor_padrao"])
    except Exception:
        pass
    nova = None
    try:
        nova = await ajard_query(
            """INSERT INTO operacional.ordens_servico
               (numero, ano, sequencia, obra, cliente_nome_avulso,
                equipamento_id, tipo_servico_id, regime_cobranca,
                operador_id, status, origem, descricao,
                data_inicio, ativo, criado_por, criado_em, client_id,
                valor_hora, valor_metro, valor_diaria, valor_km, valor_viagem)
               VALUES (%s, %s, %s, %s, %s,
                       %s, %s, %s,
                       %s, 'aberta_sem_erp', 'campo', %s,
                       CURRENT_DATE, true, %s, NOW(), %s,
                       %s, %s, %s, %s, %s)
               RETURNING id, numero, obra, status""",
            (numero, ano, seq, obra, cliente_nome or None,
             equipamento_id, tipo_servico_id, regime_cobranca,
             user["id"], observacao or None, user["id"], client_id,
             precos_padrao["hora"], precos_padrao["metro"], precos_padrao["diaria"],
             precos_padrao["km"], precos_padrao["viagem"]),
            fetch="one"
        )
    except Exception as e:
        # (13/07/2026) Retry da fila bateu no índice único do client_id →
        # a OS JÁ existe. Devolver como sucesso idempotente, nunca duplicar.
        if client_id and "client_id" in str(e).lower():
            existente = await ajard_query(
                "SELECT id, numero, obra, status FROM operacional.ordens_servico WHERE client_id=%s",
                (client_id,), fetch="one")
            if existente:
                return {"ok": True, "os": dict(existente), "dedup": True}
        raise HTTPException(status_code=500, detail=f"Erro ao criar OS: {str(e)}")
    return {"ok": True, "os": dict(nova) if nova else {"numero": numero}}

@router.get("/operacional/api/controle-mensal/periodos")
async def op_controle_mensal_periodos(db=Depends(get_db), _auth=Depends(verificar_gestor)):
    """Lista os meses que têm partes diárias salvas, agrupados por ano.
    Alimenta o menu suspenso de períodos do Controle Mensal (asyncpg — regra #44)."""
    rows = await db.fetch("""
        SELECT EXTRACT(YEAR FROM data)::int  AS ano,
               EXTRACT(MONTH FROM data)::int AS mes,
               COUNT(*)::int                 AS total
        FROM operacional.partes_diarias
        WHERE ativo = TRUE
        GROUP BY 1, 2
        ORDER BY 1 DESC, 2 DESC
    """)
    return [dict(r) for r in rows]

@router.get("/operacional/api/controle-mensal")
async def op_controle_mensal(
    ano: int,
    mes: int = None,
    equipamento_id: str = None,
    operador_id: str = None,
    _auth=Depends(verificar_gestor)
):
    """Retorna partes diárias do mês (ou do ANO inteiro, se mes ausente)
    para preview do controle mensal / exportação anual."""
    filtros = ["pd.ativo=true"]
    params = []
    if mes:
        filtros.append("EXTRACT(MONTH FROM pd.data)=%s")
        params.append(mes)
    filtros.append("EXTRACT(YEAR FROM pd.data)=%s")
    params.append(ano)

    if equipamento_id:
        filtros.append("pd.equipamento_id=%s")
        params.append(equipamento_id)
    if operador_id:
        filtros.append("pd.operador_id=%s")
        params.append(operador_id)

    where = " AND ".join(filtros)

    rows = await ajard_query(f"""
        SELECT pd.id, pd.data, pd.tipo_medicao,
               pd.horimetro_inicial, pd.horimetro_final, pd.horas_trabalhadas, pd.horas_cobradas,
               pd.km_inicial, pd.km_final, pd.km_percorrido,
               pd.qtd_metros, pd.qtd_viagens,
               pd.hora_inicio, pd.hora_fim, pd.sem_almoco,
               pd.por_conta_de, pd.observacao, pd.fechado,
               pd.equipamento_id, pd.operador_id, pd.os_id,
               pd.equipamento_terceiro, pd.vinculo_operador, pd.fornecedor,
               COALESCE(e.codigo, pd.equipamento_terceiro) AS equipamento_codigo,
               e.descricao AS equipamento_descricao,
               e.categoria AS equipamento_categoria, e.medicao AS equipamento_medicao,
               u.nome AS operador_nome,
               os.numero AS os_numero, os.obra AS os_obra, os.regime_cobranca,
               os.codigo_erp,
               COALESCE(c.nome, os.cliente_nome_avulso) AS cliente_nome
        FROM operacional.partes_diarias pd
        LEFT JOIN operacional.equipamentos e ON e.id = pd.equipamento_id
        LEFT JOIN public.usuarios_garra u    ON u.id = pd.operador_id
        LEFT JOIN operacional.ordens_servico os ON os.id = pd.os_id
        LEFT JOIN public.clientes_garra c    ON c.id = os.cliente_id
        WHERE {where}
        ORDER BY pd.data, e.codigo, u.nome
    """, tuple(params))

    # Listar equipamentos e operadores do mês (para filtros)
    equipamentos = {}
    operadores = {}
    partes = []
    total_horas_trab = 0
    total_horas_cobr = 0
    total_km = 0
    total_metros = 0
    total_viagens = 0
    dias_trabalhados = set()

    for r in (rows or []):
        d = dict(r)
        # Converter date para string ISO
        if d.get("data"):
            d["data"] = str(d["data"])
        if d.get("hora_inicio"):
            d["hora_inicio"] = str(d["hora_inicio"])
        if d.get("hora_fim"):
            d["hora_fim"] = str(d["hora_fim"])

        # Horas trabalhadas: usa o gravado; se vazio, calcula pelo relógio
        # com desconto de almoço (cobre registros antigos sem horas_trabalhadas)
        horas_trab = float(d.get("horas_trabalhadas") or 0)
        if horas_trab <= 0 and d.get("hora_inicio") and d.get("hora_fim"):
            try:
                ph = lambda s: int(str(s)[:2]) * 60 + int(str(s)[3:5])
                im = ph(d["hora_inicio"]); fm = ph(d["hora_fim"])
                dm = fm - im
                if dm < 0: dm += 24 * 60
                bruto = dm / 60
                cruza = (im < 12*60) and (fm > 12*60 or fm < im)
                almoco = 1 if (not d.get("sem_almoco") and cruza and bruto > 6) else 0
                horas_trab = round(max(0, bruto - almoco), 2)
                d["horas_trabalhadas"] = horas_trab
            except (TypeError, ValueError):
                pass

        partes.append(d)
        dias_trabalhados.add(d["data"])
        total_horas_trab += horas_trab
        total_horas_cobr += float(d.get("horas_cobradas") or 0)
        total_km += float(d.get("km_percorrido") or 0)
        total_metros += float(d.get("qtd_metros") or 0)
        total_viagens += float(d.get("qtd_viagens") or 0)

        if d.get("equipamento_id") and d.get("equipamento_codigo"):
            equipamentos[d["equipamento_id"]] = {
                "id": d["equipamento_id"],
                "codigo": d["equipamento_codigo"],
                "descricao": d["equipamento_descricao"]
            }
        if d.get("operador_id") and d.get("operador_nome"):
            operadores[d["operador_id"]] = {
                "id": d["operador_id"],
                "nome": d["operador_nome"]
            }

    import calendar
    if mes:
        dias_no_mes = calendar.monthrange(ano, mes)[1]
    else:
        dias_no_mes = 366 if calendar.isleap(ano) else 365

    return {
        "mes": mes, "ano": ano,
        "dias_no_mes": dias_no_mes,
        "total_registros": len(partes),
        "dias_trabalhados": len(dias_trabalhados),
        "dias_parados": dias_no_mes - len(dias_trabalhados),
        "total_horas_trabalhadas": round(total_horas_trab, 2),
        "total_horas_cobradas": round(total_horas_cobr, 2),
        "total_km": round(total_km, 2),
        "total_metros": round(total_metros, 2),
        "total_viagens": int(total_viagens),
        "equipamentos": list(equipamentos.values()),
        "operadores": list(operadores.values()),
        "partes": partes
    }

@router.get("/operacional/api/controle-mensal/excel")
async def op_controle_mensal_excel(
    ano: int,
    mes: int = None,
    view: str = "equipamento",
    equipamento_id: str = None,
    operador_id: str = None,
    _auth=Depends(verificar_gestor)
):
    """Gera Excel do controle mensal (ou ANUAL, se mes ausente) —
    1 aba por equipamento ou por colaborador."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io, calendar

    # Buscar dados
    dados = await op_controle_mensal(ano=ano, mes=mes, equipamento_id=equipamento_id, operador_id=operador_id, _auth=_auth)
    partes = dados["partes"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    meses_pt = ['','Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    if mes:
        titulo_mes = f"{meses_pt[mes]}/{ano}"
        titulo_doc = f"CONTROLE MENSAL — {titulo_mes}"
    else:
        titulo_mes = f"Ano {ano}"
        titulo_doc = f"CONTROLE ANUAL — {ano}"

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1A2A5E", end_color="1A2A5E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center")
    border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    total_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    total_font = Font(bold=True, size=10)

    if view == "equipamento":
        # Agrupar por equipamento
        grupos = {}
        for p in partes:
            key = p.get("equipamento_id") or "sem_equipamento"
            label = f"{p.get('equipamento_codigo','?')} — {p.get('equipamento_descricao','')}"
            if key not in grupos:
                grupos[key] = {"label": label, "partes": []}
            grupos[key]["partes"].append(p)
    else:
        # Agrupar por operador
        grupos = {}
        for p in partes:
            key = p.get("operador_id") or "sem_operador"
            label = p.get("operador_nome") or "Sem operador"
            if key not in grupos:
                grupos[key] = {"label": label, "partes": []}
            grupos[key]["partes"].append(p)

    if not grupos:
        # Aba vazia
        ws = wb.create_sheet("Sem dados")
        ws["A1"] = f"Nenhum registro encontrado para {titulo_mes}"
    else:
        for key, grupo in grupos.items():
            nome_aba = grupo["label"][:31]  # Excel limita 31 chars
            ws = wb.create_sheet(nome_aba)

            # Título
            ws.merge_cells('A1:M1')
            ws['A1'] = titulo_doc
            ws['A1'].font = Font(bold=True, size=14, color="1A2A5E")
            ws.merge_cells('A2:M2')
            ws['A2'] = grupo["label"]
            ws['A2'].font = Font(bold=True, size=11, color="E8820C")

            # Headers
            # (13/07/2026) ESPELHO da tabela do desktop: + Un., Valor (R$),
            # medição real por linha e dias sem apontamento (NÃO RODOU etc.)
            headers = ['Data','Cód Interno','OS','Cliente','Operador' if view=='equipamento' else 'Equipamento',
                       'Inicial','Final','Trab.','Cobr.','Un.','Valor (R$)','Regime','Por conta']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border

            row = 5
            soma_t = {"h": 0.0, "m": 0.0, "km": 0.0, "v": 0.0}
            soma_c = {"h": 0.0, "m": 0.0, "km": 0.0, "v": 0.0}
            soma_valor = 0.0
            dias_set = set()

            def _fmt_soma(o):
                pares = [(o["h"], "h"), (o["m"], "m"), (o["km"], "km"), (o["v"], "viag.")]
                itens = [f"{n:.1f} {u}" for n, u in pares if n > 0]
                return " · ".join(itens) if itens else 0

            # Partes agrupadas por dia (para intercalar os dias sem apontamento)
            por_dia = {}
            for p in grupo["partes"]:
                por_dia.setdefault(str(p.get("data") or ""), []).append(p)

            from datetime import date as dt_date, timedelta
            gap_font_fds = Font(bold=True, size=10, color="DC2626")
            gap_font = Font(bold=True, size=10, color="475569")
            gap_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            if mes:
                hoje = dt_date.today()
                ultimo = calendar.monthrange(ano, mes)[1]
                if ano == hoje.year and mes == hoje.month:
                    ultimo = min(ultimo, hoje.day)
                dias_iter = [dt_date(ano, mes, d).isoformat() for d in range(1, ultimo + 1)]
            else:
                dias_iter = sorted(por_dia.keys())  # anual: sem lacunas (ficaria imenso)

            for dia_iso in dias_iter:
                lista_dia = sorted(por_dia.get(dia_iso, []), key=lambda x: str(x.get("criado_em") or ""))
                if not lista_dia:
                    # ESPELHO: dia sem apontamento — NÃO RODOU / SÁBADO / DOMINGO
                    try:
                        d_obj = dt_date.fromisoformat(dia_iso)
                        rotulo = "SÁBADO" if d_obj.weekday() == 5 else ("DOMINGO" if d_obj.weekday() == 6 else "NÃO RODOU")
                        c1 = ws.cell(row=row, column=1, value=d_obj.strftime("%d/%m/%Y"))
                        c1.fill = gap_fill; c1.border = border
                        c2 = ws.cell(row=row, column=2, value=rotulo)
                        c2.font = gap_font_fds if rotulo != "NÃO RODOU" else gap_font
                        c2.fill = gap_fill
                        for col in range(2, 14):
                            ws.cell(row=row, column=col).fill = gap_fill
                            ws.cell(row=row, column=col).border = border
                        row += 1
                    except Exception:
                        pass
                    continue
                for p in lista_dia:
                    data_fmt = ""
                    try:
                        data_fmt = dt_date.fromisoformat(str(p.get("data"))).strftime("%d/%m/%Y")
                    except Exception:
                        data_fmt = p.get("data") or ""

                    med = (p.get("tipo_medicao") or p.get("equipamento_medicao") or "horimetro").lower()
                    if med == "metros":
                        h_ini = ""; h_fin = ""
                        h_trab = float(p.get("qtd_metros") or 0); h_cobr = h_trab
                        unidade, chave = "m", "m"
                        rotulo_med = "metros"
                    elif med == "viagem":
                        h_ini = ""; h_fin = ""
                        h_trab = float(p.get("qtd_viagens") or 0) or 1.0; h_cobr = h_trab
                        unidade, chave = "viag.", "v"
                        rotulo_med = "viagem"
                    elif med == "km":
                        h_ini = float(p.get("km_inicial") or 0)
                        h_fin = float(p.get("km_final") or 0)
                        h_trab = float(p.get("km_percorrido") or 0); h_cobr = h_trab
                        unidade, chave = "km", "km"
                        rotulo_med = "km"
                    else:
                        # (17/07/2026) Espelho da tela de partes: Inicial/Final e
                        # Trab. = HORÍMETRO quando informado (máquina); Cobr. =
                        # horas do registro (relógio quando o operador registrou)
                        _hi_h = p.get("horimetro_inicial"); _hf_h = p.get("horimetro_final")
                        _hi_rel = p.get("hora_inicio"); _hf_rel = p.get("hora_fim")
                        if _hi_h is not None and _hf_h is not None:
                            h_ini = float(_hi_h); h_fin = float(_hf_h)
                            h_trab = round(h_fin - h_ini, 2)
                        elif _hi_rel and _hf_rel:
                            h_ini = str(_hi_rel)[:5]; h_fin = str(_hf_rel)[:5]
                            h_trab = float(p.get("horas_trabalhadas") or 0)
                        else:
                            h_ini = ""; h_fin = ""
                            h_trab = float(p.get("horas_trabalhadas") or 0)
                        h_cobr = float(p.get("horas_cobradas") or p.get("horas_trabalhadas") or 0)
                        unidade, chave = "h", "h"
                        reg = str(p.get("regime_cobranca") or "").lower()
                        rotulo_med = p.get("regime_cobranca") if "hora" in reg else "hora"

                    soma_t[chave] += h_trab
                    soma_c[chave] += h_cobr
                    val = p.get("valor")
                    try:
                        soma_valor += float(val) if val not in (None, "") else 0.0
                    except Exception:
                        pass
                    dias_set.add(p.get("data"))

                    col4 = p.get("operador_nome", "") if view == "equipamento" else p.get("equipamento_codigo", "")
                    valores = [
                        data_fmt,
                        p.get("os_numero", ""),
                        p.get("codigo_erp", "") or "",
                        p.get("cliente_nome", ""),
                        col4,
                        h_ini, h_fin,
                        round(h_trab, 2) if h_trab else 0,
                        round(h_cobr, 2) if h_cobr else 0,
                        unidade,
                        (round(float(val), 2) if val not in (None, "") else ""),
                        rotulo_med,
                        p.get("por_conta_de", "")
                    ]
                    for col, v in enumerate(valores, 1):
                        cell = ws.cell(row=row, column=col, value=v)
                        cell.alignment = cell_align
                        cell.border = border
                    row += 1

            # Linha TOTAL — por unidade, como no desktop
            row += 1
            ws.cell(row=row, column=1, value="TOTAL").font = total_font
            ws.cell(row=row, column=1).fill = total_fill
            ws.cell(row=row, column=8, value=_fmt_soma(soma_t)).font = total_font
            ws.cell(row=row, column=8).fill = total_fill
            ws.cell(row=row, column=9, value=_fmt_soma(soma_c)).font = total_font
            ws.cell(row=row, column=9).fill = total_fill
            if soma_valor > 0:
                ws.cell(row=row, column=11, value=round(soma_valor, 2)).font = total_font
                ws.cell(row=row, column=11).fill = total_fill

            ws.cell(row=row+1, column=1, value=f"Dias trabalhados: {len(dias_set)}").font = Font(size=10, color="64748B")
            dias_no_mes = dados["dias_no_mes"]
            ws.cell(row=row+2, column=1, value=f"Dias parados: {dias_no_mes - len(dias_set)}").font = Font(size=10, color="64748B")

            # Larguras (13 colunas — espelho do desktop)
            widths = [12, 14, 10, 20, 16, 10, 10, 11, 11, 7, 12, 10, 12]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    # Aba RESUMO
    ws_res = wb.create_sheet("RESUMO", 0)
    ws_res['A1'] = titulo_doc
    ws_res['A1'].font = Font(bold=True, size=14, color="1A2A5E")
    ws_res['A3'] = "Total de registros:"
    ws_res['B3'] = dados["total_registros"]
    ws_res['A4'] = "Dias trabalhados:"
    ws_res['B4'] = dados["dias_trabalhados"]
    ws_res['A5'] = "Dias parados:"
    ws_res['B5'] = dados["dias_parados"]
    ws_res['A6'] = "Horas trabalhadas:"
    ws_res['B6'] = dados["total_horas_trabalhadas"]
    ws_res['A7'] = "Horas cobradas:"
    ws_res['B7'] = dados["total_horas_cobradas"]
    ws_res['A8'] = "KM total:"
    ws_res['B8'] = dados["total_km"]
    ws_res['A9'] = "Metros total:"
    ws_res['B9'] = dados["total_metros"]
    for r in range(3,10):
        ws_res.cell(row=r, column=1).font = Font(bold=True, size=10)
    ws_res.column_dimensions['A'].width = 22
    ws_res.column_dimensions['B'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    view_label = "equipamento" if view == "equipamento" else "colaborador"
    if mes:
        filename = f"controle-mensal-{view_label}-{meses_pt[mes].lower()}{ano}.xlsx"
    else:
        filename = f"controle-anual-{view_label}-{ano}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ══════════════════════════════════════════════════════════════
# RESUMO MENSAL DO OPERADOR (Item C — horas do mês no mobile)
# Adicionado em 04/07/2026. O operador vê no Histórico quantas
# horas acumulou no mês corrente e quanto falta para fechar.
# ══════════════════════════════════════════════════════════════

@router.get("/operacional/api/resumo-mensal")
async def op_resumo_mensal(payload=Depends(verificar_token)):
    """Retorna total de horas e dias trabalhados do operador no mês corrente."""
    login = payload.get("sub", "")
    user = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    hoje = date.today()
    primeiro_dia = hoje.replace(day=1)

    resumo = await ajard_query(
        """SELECT
             COUNT(DISTINCT p.data) AS dias_trabalhados,
             -- (13/07/2026) COMISSÃO pelo CONTEÚDO do registro:
             --   • registro que CONTÉM metros → comissiona pelos METROS
             --     (as horas dele não somam);
             --   • registro de horímetro/hora SEM metros → soma as HORAS.
             COALESCE(SUM(CASE WHEN COALESCE(p.qtd_metros,0) > 0
                                 OR COALESCE(p.tipo_medicao,'horimetro') = 'metros'
                               THEN 0 ELSE p.horas_trabalhadas END), 0) AS total_horas,
             COALESCE(SUM(CASE WHEN COALESCE(p.qtd_metros,0) > 0
                                 OR COALESCE(p.tipo_medicao,'horimetro') = 'metros'
                               THEN 0 ELSE p.horas_cobradas END), 0) AS total_horas_cobradas,
             COALESCE(SUM(p.qtd_metros), 0) AS total_metros,
             COUNT(p.id) AS total_apontamentos,
             COUNT(DISTINCT p.os_id) AS total_os
           FROM operacional.partes_diarias p
           JOIN operacional.ordens_servico os ON os.id = p.os_id
           WHERE p.operador_id = %s
             AND p.ativo = true
             AND COALESCE(p.vinculo_operador, 'proprio') <> 'terceiro'
             AND p.data >= %s
             AND p.data <= %s""",
        (user["id"], primeiro_dia, hoje), fetch="one"
    )

    _MESES = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
              7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}
    return {
        "mes": hoje.strftime("%m/%Y"),
        "mes_nome": _MESES[hoje.month],
        "dias_trabalhados": int(resumo["dias_trabalhados"] or 0),
        "total_horas": round(float(resumo["total_horas"] or 0), 1),
        "total_horas_cobradas": round(float(resumo["total_horas_cobradas"] or 0), 1),
        "total_metros": round(float(resumo["total_metros"] or 0), 1),
        "total_apontamentos": int(resumo["total_apontamentos"] or 0),
        "total_os": int(resumo["total_os"] or 0),
    }


# ══════════════════════════════════════════════════════════════
# OPERADOR EDITA SUA PRÓPRIA PARTE (04/07/2026)
# Caso real: esqueceu o horímetro → precisa conferir e corrigir.
# Permite editar APENAS medição/observação de partes que ele criou
# (ou de OS dele), enquanto não fechadas. Horas recalculam pela
# mesma regra do registro (_calc_horas_parte). Horas COBRADAS não
# são tocadas aqui — domínio da gestão (Luana/Edvania).
# ══════════════════════════════════════════════════════════════

@router.patch("/operacional/api/minhas-partes/{parte_id}")
async def op_editar_minha_parte(parte_id: str, request: Request, payload=Depends(verificar_token)):
    d = await request.json()
    login = payload.get("sub", "")
    user = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    parte = await ajard_query(
        """SELECT p.*, os.operador_id AS os_operador_id
           FROM operacional.partes_diarias p
           JOIN operacional.ordens_servico os ON os.id = p.os_id
           WHERE p.id=%s AND p.ativo=true""",
        (parte_id,), fetch="one"
    )
    if not parte:
        raise HTTPException(status_code=404, detail="Parte não encontrada")
    if parte.get("fechado"):
        raise HTTPException(status_code=403, detail="Parte já fechada — fale com a gestão para ajustes.")
    dono = str(parte.get("criado_por") or "") == str(user["id"]) or \
           str(parte.get("os_operador_id") or "") == str(user["id"])
    if not dono:
        raise HTTPException(status_code=403, detail="Você só pode editar os seus próprios registros.")

    # Campos que o operador pode corrigir (medição + observação)
    # (13/07/2026) + equipamento e operador (vínculo próprio) — corrigir
    # máquina/pessoa errada é correção legítima do dia, com efeito direto na
    # comissão (a atribuição segue o operador da parte).
    EDITAVEIS = ["data", "horimetro_inicial", "horimetro_final", "hora_inicio", "hora_fim",
                 "sem_almoco", "qtd_metros", "observacao",
                 "km_inicial", "km_final", "qtd_viagens", "quantidade_diarias",
                 "equipamento_id", "operador_id"]
    merged = dict(parte)
    algum = False
    for c in EDITAVEIS:
        if c in d:
            merged[c] = d[c]
            algum = True
    if not algum:
        return {"ok": True, "msg": "Nada a alterar."}

    # (08/07/2026) Sanidade — mesma régua do registro (Regra 62)
    _validar_medicao_parte(merged)
    if "data" in d:
        nova_dt = _validar_data_parte(merged.get("data"))
        antiga = parte.get("data")
        try:
            antiga_dt = antiga if isinstance(antiga, date) else datetime.strptime(str(antiga), "%Y-%m-%d").date()
            # Correção do operador é do dia (typo) — mudar de MÊS mexe com
            # período/fechamento e é domínio da gestão.
            if (nova_dt.year, nova_dt.month) != (antiga_dt.year, antiga_dt.month):
                raise HTTPException(status_code=400,
                    detail="Mudança de mês do registro é feita pela gestão — fale com a Edvania/Luana.")
        except HTTPException:
            raise
        except Exception:
            pass

    horas = _calc_horas_parte(merged)
    # (18/07/2026) FONTE EDITADA POR ÚLTIMO VENCE — resolve o fluxo
    # abre-manhã/fecha-noite: completar o horímetro na correção RECALCULA
    # pelo horímetro (mesmo com relógio gravado); mexer no relógio mantém
    # o relógio. O front envia apenas os campos que o operador alterou.
    _tocou_hor = ("horimetro_inicial" in d) or ("horimetro_final" in d)
    _tocou_rel = ("hora_inicio" in d) or ("hora_fim" in d) or ("sem_almoco" in d)
    if _tocou_rel and not _tocou_hor:
        # Editou o RELÓGIO de propósito → relógio vale (já é a base, explícito
        # por clareza e à prova de mudanças futuras)
        _hr = _horas_relogio(merged)
        if _hr is not None:
            horas = _hr
    elif _tocou_hor and not _tocou_rel:
        # (23/07/2026) Espelho que faltava sob a base relógio-primeiro:
        # editar o HORÍMETRO de propósito (ex.: gestor corrigindo a máquina,
        # fluxo abre-manhã/fecha-noite) → horímetro vale NESTE registro,
        # mesmo com relógio gravado. Sem isso a correção era ignorada.
        _hi = merged.get("horimetro_inicial"); _hf = merged.get("horimetro_final")
        if _hi is not None and _hf is not None:
            try:
                _hm = round(float(_hf) - float(_hi), 2)
                if _hm >= 0:
                    horas = _hm
            except (TypeError, ValueError):
                pass
    # Tocou os dois na mesma edição → base oficial decide (relógio prevalece)
    _validar_horas_plausiveis(horas)
    await _validar_sobreposicao_horimetro(
        merged.get("equipamento_id"),
        merged.get("horimetro_inicial"), merged.get("horimetro_final"),
        ignorar_parte_id=parte_id)

    def _num(v):
        try: return float(v) if v not in (None, "") else None
        except (TypeError, ValueError): return None

    k_ini = _num(merged.get("km_inicial")); k_fin = _num(merged.get("km_final"))
    km_perc = round(k_fin - k_ini, 1) if (k_ini is not None and k_fin is not None) else None
    try:
        viagens = int(merged.get("qtd_viagens")) if merged.get("qtd_viagens") not in (None, "") else None
    except (TypeError, ValueError):
        viagens = None
    await ajard_query(
        """UPDATE operacional.partes_diarias
           SET data=%s, horimetro_inicial=%s, horimetro_final=%s,
               hora_inicio=%s, hora_fim=%s, sem_almoco=%s,
               qtd_metros=%s, observacao=%s, horas_trabalhadas=%s,
               km_inicial=%s, km_final=%s, km_percorrido=%s, qtd_viagens=%s
           WHERE id=%s""",
        (merged.get("data"), _num(merged.get("horimetro_inicial")), _num(merged.get("horimetro_final")),
         merged.get("hora_inicio") or None, merged.get("hora_fim") or None,
         bool(merged.get("sem_almoco")),
         _num(merged.get("qtd_metros")),
         (merged.get("observacao") or "").strip() or None,
         horas,
         k_ini, k_fin, km_perc, viagens,
         parte_id),
        fetch="none"
    )
    return {"ok": True, "horas_trabalhadas": horas}


# ══════════════════════════════════════════════════════════════
# ALERTA DE CONFLITO DE EQUIPAMENTO (05/07/2026)
# Caso real: pausar uma obra e levar a máquina para outra é legítimo —
# mas precisa ser DECISÃO CONSCIENTE. Ao abrir/editar OS com máquina
# que já está em OS ativa, o front consulta aqui e pede confirmação.
# ══════════════════════════════════════════════════════════════

@router.get("/operacional/api/equipamentos/{eq_id}/conflito")
async def op_conflito_equipamento(eq_id: str, ignorar_os: str = None,
                                  payload=Depends(verificar_token)):
    """Retorna a OS ATIVA que já usa este equipamento (se houver),
    ignorando a própria OS em edição."""
    args = [eq_id]
    filtro_ignorar = ""
    if ignorar_os:
        args.append(ignorar_os)
        filtro_ignorar = "AND os.id <> %s"
    row = await ajard_query(
        f"""SELECT os.numero, os.obra,
                  COALESCE(c.nome, os.cliente_nome_avulso) AS cliente_nome,
                  op.nome AS operador_nome
           FROM operacional.ordens_servico os
           LEFT JOIN public.clientes_garra c  ON c.id = os.cliente_id
           LEFT JOIN public.usuarios_garra op ON op.id = os.operador_id
           WHERE os.equipamento_id = %s
             AND os.ativo = true
             AND os.status NOT IN ('concluida_completa','concluida_sem_erp','cancelada')
             {filtro_ignorar}
           ORDER BY os.criado_em DESC
           LIMIT 1""",
        args, fetch="one"
    )
    if not row:
        return {"conflito": False}
    return {"conflito": True, "os": dict(row)}


# ══════════════════════════════════════════════════════════════
# OPERADOR EXCLUI SUA PRÓPRIA PARTE (05/07/2026)
# Caso real: registro duplicado por engano. Soft delete (ativo=false),
# mesmas guardas da edição: só o criador/operador da OS, não fechada.
# ══════════════════════════════════════════════════════════════

@router.delete("/operacional/api/minhas-partes/{parte_id}")
async def op_excluir_minha_parte(parte_id: str, payload=Depends(verificar_token)):
    login = payload.get("sub", "")
    user = await ajard_query(
        "SELECT id FROM public.usuarios_garra WHERE login=%s", (login,), fetch="one"
    )
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    parte = await ajard_query(
        """SELECT p.id, p.criado_por, p.fechado, os.operador_id AS os_operador_id
           FROM operacional.partes_diarias p
           JOIN operacional.ordens_servico os ON os.id = p.os_id
           WHERE p.id=%s AND p.ativo=true""",
        (parte_id,), fetch="one"
    )
    if not parte:
        raise HTTPException(status_code=404, detail="Registro não encontrado")
    if parte.get("fechado"):
        raise HTTPException(status_code=403, detail="Registro fechado — fale com a gestão.")
    dono = str(parte.get("criado_por") or "") == str(user["id"]) or \
           str(parte.get("os_operador_id") or "") == str(user["id"])
    if not dono:
        raise HTTPException(status_code=403, detail="Você só pode excluir os seus próprios registros.")
    await ajard_query(
        "UPDATE operacional.partes_diarias SET ativo=false WHERE id=%s",
        (parte_id,), fetch="none"
    )
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════
# AGENDA GOOGLE → OS A CONFIRMAR  (21/07/2026)
# Cada equipamento tem sua agenda Google; a URL ICS secreta fica no
# cadastro do equipamento (agenda_ics_url — Regra 63, sem hardcode).
# Sincronizar busca os eventos, fatia eventos longos em 1 segmento
# por mês (regra: evento longo = 1 OS por mês) e grava em
# operacional.agenda_eventos como PENDENTE. Nada vira OS sem
# confirmação humana no Admin. Idempotente por (uid_google, mes_ref).
# ══════════════════════════════════════════════════════════════════

def _agenda_parse_ics(texto_ics: str):
    """Extrai eventos (uid, titulo, descricao, data_ini, data_fim) de um ICS.
    Usa a lib icalendar; DTEND de evento de dia inteiro é exclusivo → -1 dia.
    RRULE (recorrência) fora do escopo v1: trata como ocorrência única."""
    from icalendar import Calendar as _ICal
    eventos = []
    cal = _ICal.from_ical(texto_ics)
    for comp in cal.walk("VEVENT"):
        try:
            uid = str(comp.get("UID") or "").strip()
            titulo = str(comp.get("SUMMARY") or "").strip()
            desc = str(comp.get("DESCRIPTION") or "").strip()
            dtstart = comp.get("DTSTART")
            if not uid or not dtstart:
                continue
            ini = dtstart.dt
            fim_prop = comp.get("DTEND")
            fim = fim_prop.dt if fim_prop else ini
            all_day = not hasattr(ini, "hour")
            d_ini = ini if all_day else ini.date()
            d_fim = fim if not hasattr(fim, "hour") else fim.date()
            if all_day and fim_prop is not None:
                d_fim = d_fim - timedelta(days=1)  # DTEND exclusivo
            if d_fim < d_ini:
                d_fim = d_ini
            eventos.append({"uid": uid, "titulo": titulo, "descricao": desc,
                            "data_ini": d_ini, "data_fim": d_fim})
        except Exception:
            continue  # evento malformado não derruba a sincronização
    return eventos


def _agenda_segmentos_mensais(d_ini, d_fim):
    """Fatia [d_ini, d_fim] em segmentos por mês civil.
    Retorna lista de (mes_ref=1º dia do mês, seg_ini, seg_fim)."""
    segs = []
    cursor = d_ini
    while cursor <= d_fim:
        mes_ref = cursor.replace(day=1)
        ultimo = date(cursor.year, cursor.month,
                      calendar.monthrange(cursor.year, cursor.month)[1])
        seg_fim = min(ultimo, d_fim)
        segs.append((mes_ref, cursor, seg_fim))
        cursor = ultimo + timedelta(days=1)
    return segs


@router.post("/operacional/api/agenda/sincronizar")
async def op_agenda_sincronizar(payload=Depends(verificar_gestor)):
    """Varre as agendas ICS cadastradas nos equipamentos e grava/atualiza
    os eventos pendentes. Só meses do atual em diante entram."""
    import asyncio, requests as _rq
    equips = await ajard_query(
        """SELECT id, codigo, agenda_ics_url FROM operacional.equipamentos
           WHERE ativo=true AND agenda_ics_url IS NOT NULL AND agenda_ics_url <> ''"""
    )
    hoje_local = (datetime.utcnow() - timedelta(hours=3)).date()
    corte = hoje_local.replace(day=1)
    loop = asyncio.get_event_loop()
    resumo = {"agendas": 0, "eventos_novos": 0, "eventos_atualizados": 0, "erros": []}
    for eq in (equips or []):
        url = (eq.get("agenda_ics_url") or "").strip()
        try:
            resp = await loop.run_in_executor(
                None, lambda u=url: _rq.get(u, timeout=20))
            if resp.status_code != 200:
                resumo["erros"].append(f"{eq['codigo']}: HTTP {resp.status_code}")
                continue
            eventos = _agenda_parse_ics(resp.text)
            resumo["agendas"] += 1
        except Exception as e:
            resumo["erros"].append(f"{eq['codigo']}: {str(e)[:120]}")
            continue
        for ev in eventos:
            for mes_ref, seg_ini, seg_fim in _agenda_segmentos_mensais(ev["data_ini"], ev["data_fim"]):
                if seg_fim < corte:
                    continue  # mês já passado — não importa histórico
                existente = await ajard_query(
                    "SELECT id, status FROM operacional.agenda_eventos WHERE uid_google=%s AND mes_ref=%s",
                    (ev["uid"], mes_ref), fetch="one"
                )
                if existente:
                    if existente.get("status") == "pendente":
                        await ajard_query(
                            """UPDATE operacional.agenda_eventos
                               SET titulo=%s, descricao=%s, data_inicio=%s, data_fim=%s,
                                   equipamento_id=%s, atualizado_em=now()
                               WHERE id=%s""",
                            (ev["titulo"], ev["descricao"] or None, seg_ini, seg_fim,
                             eq["id"], existente["id"]), fetch="none"
                        )
                        resumo["eventos_atualizados"] += 1
                else:
                    await ajard_query(
                        """INSERT INTO operacional.agenda_eventos
                           (uid_google, mes_ref, equipamento_id, titulo, descricao,
                            data_inicio, data_fim, status)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,'pendente')""",
                        (ev["uid"], mes_ref, eq["id"], ev["titulo"],
                         ev["descricao"] or None, seg_ini, seg_fim), fetch="none"
                    )
                    resumo["eventos_novos"] += 1
    return resumo


@router.get("/operacional/api/agenda/eventos")
async def op_agenda_listar(status: Optional[str] = "pendente",
                           payload=Depends(verificar_gestor)):
    rows = await ajard_query(
        """SELECT ae.id, ae.uid_google, ae.mes_ref, ae.titulo, ae.descricao,
                  ae.data_inicio, ae.data_fim, ae.status, ae.os_id,
                  ae.equipamento_id, eq.codigo AS equipamento_codigo,
                  eq.descricao AS equipamento_descricao
           FROM operacional.agenda_eventos ae
           LEFT JOIN operacional.equipamentos eq ON eq.id = ae.equipamento_id
           WHERE ae.status = %s
           ORDER BY ae.mes_ref, eq.codigo, ae.data_inicio""",
        (status,)
    )
    return [dict(r) for r in (rows or [])]


@router.post("/operacional/api/agenda/eventos/{ev_id}/ignorar")
async def op_agenda_ignorar(ev_id: str, payload=Depends(verificar_gestor)):
    row = await ajard_query(
        """UPDATE operacional.agenda_eventos SET status='ignorado', atualizado_em=now()
           WHERE id=%s AND status='pendente' RETURNING id""",
        (ev_id,), fetch="one"
    )
    if not row:
        raise HTTPException(status_code=404, detail="Evento não encontrado ou já tratado")
    return {"ok": True}


@router.post("/operacional/api/agenda/eventos/{ev_id}/confirmar")
async def op_agenda_confirmar(ev_id: str, request: Request,
                              payload=Depends(verificar_gestor)):
    """Vincula o evento à OS criada pela gestão (fluxo: modal pré-preenchido
    → salvarOS → esta rota marca o evento como confirmado)."""
    d = await request.json()
    os_id = d.get("os_id")
    if not os_id:
        raise HTTPException(status_code=400, detail="os_id é obrigatório")
    row = await ajard_query(
        """UPDATE operacional.agenda_eventos
           SET status='confirmado', os_id=%s, atualizado_em=now()
           WHERE id=%s AND status='pendente' RETURNING id""",
        (os_id, ev_id), fetch="one"
    )
    if not row:
        raise HTTPException(status_code=404, detail="Evento não encontrado ou já tratado")
    return {"ok": True}
